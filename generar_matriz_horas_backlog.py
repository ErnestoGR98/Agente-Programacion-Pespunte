"""
Genera Excel: Matriz de Horas por Etapa - Backlog Andrea (Sem 15-21)
PROPUESTA de distribucion optimizada por semana
USA FORMULAS en Excel para que sea editable
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

# === Catalogo: sec/par por etapa (de Supabase) ===
CATALOGO = {
    "68127 NE/RO SLI":    {"PRELIMINARES": 108, "ROBOT": 263, "POST": 72, "N/A PRELIMINAR": 36},
    "64197 NE SLI":        {"ROBOT": 47, "POST": 108, "N/A PRELIMINAR": 36},
    "65568 RO/HU SLI":    {"PRELIMINARES": 108, "ROBOT": 91, "POST": 108},
    "77525 NE TEX":        {"POST": 108, "MAQUILA": 468},
    "94750 AA SLI":        {"PRELIMINARES": 72, "ROBOT": 100, "POST": 36, "MAQUILA": 120},
    "61747 GC SLI":        {"PRELIMINARES": 216, "ROBOT": 60, "POST": 180, "MAQUILA": 424},
    "65422 HU SLI":        {"PRELIMINARES": 44, "ROBOT": 155, "POST": 290, "N/A PRELIMINAR": 42},
    "88186 NE SLI":        {"PRELIMINARES": 144, "ROBOT": 145, "POST": 108, "N/A PRELIMINAR": 36},
    "62100 BL SLI":        {"PRELIMINARES": 134, "ROBOT": 141, "POST": 658},
    "61748 KR SLI":        {"PRELIMINARES": 180, "ROBOT": 180, "POST": 240, "N/A PRELIMINAR": 60},
    "65413 NE/CG/VI SLI": {"ROBOT": 125, "POST": 144},
}

# Demanda total del backlog original (pares)
DEMANDA = {
    "68127 NE/RO SLI":    3900,
    "64197 NE SLI":        4100,
    "65568 RO/HU SLI":    3300,
    "77525 NE TEX":        3500,
    "94750 AA SLI":        3000,
    "61747 GC SLI":        1800,
    "65422 HU SLI":        2000,
    "88186 NE SLI":        1800,
    "62100 BL SLI":        1800,
    "61748 KR SLI":        1500,
    "65413 NE/CG/VI SLI": 1600,
    "93347 CA SLI":        1500,
    "69906 BL SLI":        1200,
    "93349 RS SLI":        1200,
    "93346 NE SLI":        1200,
}

MODELOS_ORDER = list(DEMANDA.keys())
MISSING = {"93347 CA SLI", "69906 BL SLI", "93349 RS SLI", "93346 NE SLI"}
SEMANAS = [15, 16, 17, 18, 19, 20, 21]

ETAPAS_FULL = ["PRELIMINARES", "ROBOT", "POST"]
ETAPAS_SIMPLE = ["N/A PRELIMINAR", "MAQUILA"]
ETAPAS_ALL = ETAPAS_FULL + ETAPAS_SIMPLE

HEADCOUNT_DEFAULT = {"PRELIMINARES": 4, "ROBOT": 6, "POST": 8}
JORNADA_DEFAULT = 9
DIAS_SEMANA = 6  # Lun-Sab

# Capacidad semanal por etapa (hrs)
CAP = {e: HEADCOUNT_DEFAULT[e] * JORNADA_DEFAULT * DIAS_SEMANA for e in ETAPAS_FULL}
# PRELIM=216, ROBOT=324, POST=432

# ==============================================================================
# ALGORITMO DE DISTRIBUCION
# ==============================================================================
def calc_stage_hrs(modelo, pares):
    """Calcula horas por etapa para un modelo y cantidad de pares"""
    if modelo in MISSING or modelo not in CATALOGO:
        return {}
    return {e: pares * spp / 3600 for e, spp in CATALOGO[modelo].items()}

def optimize_distribution():
    """
    Distribuye pares por semana agrupando modelos en pocas semanas (lotes grandes).

    Heuristicas:
    1. Cada modelo aparece en 1-3 semanas max (lotes grandes, pocos cambios)
    2. Lote minimo 300 pares (no hacer lotes chiquitos de 100)
    3. Max 5-7 modelos con catalogo por semana (mezcla controlada)
    4. Balance de horas por etapa entre semanas (no saturar un stage)
    5. Modelos sin catalogo se concentran en 2-3 semanas
    """
    dist = {m: {s: 0 for s in SEMANAS} for m in MODELOS_ORDER}
    week_hrs = {s: {e: 0.0 for e in ETAPAS_ALL} for s in SEMANAS}
    week_pares = {s: 0 for s in SEMANAS}
    week_models = {s: set() for s in SEMANAS}  # modelos asignados por semana

    MAX_MODELS_PER_WEEK = 7   # mezcla maxima de modelos con catalogo
    MIN_BATCH = 300           # lote minimo por modelo por semana
    TARGET_UTIL = 0.70        # usar ~70% de capacidad como target

    # Calcular bottleneck de cada modelo (que stage lo limita mas)
    def model_bottleneck_hrs(modelo, pares):
        """Hrs del stage mas critico (relativo a capacidad)"""
        if modelo in MISSING:
            return 0
        cat = CATALOGO[modelo]
        ratios = []
        for e in ETAPAS_FULL:
            if e in cat:
                h = pares * cat[e] / 3600
                ratios.append(h / CAP[e])
        return max(ratios) if ratios else 0

    def add_to_week(modelo, pares, sem):
        """Asigna pares a una semana y actualiza contadores"""
        dist[modelo][sem] += pares
        week_pares[sem] += pares
        week_models[sem].add(modelo)
        cat = CATALOGO.get(modelo, {})
        for e, spp in cat.items():
            week_hrs[sem][e] = week_hrs[sem].get(e, 0) + pares * spp / 3600

    def week_bottleneck(sem):
        """Max utilizacion de cualquier stage en una semana"""
        return max(week_hrs[sem].get(e, 0) / CAP[e] for e in ETAPAS_FULL)

    def week_can_fit(modelo, pares, sem):
        """Checa si un modelo cabe en una semana sin exceder limites"""
        # Ya esta asignado ahi? ok, no cuenta como modelo nuevo
        if modelo not in week_models[sem]:
            cat_models = [m for m in week_models[sem] if m not in MISSING]
            if len(cat_models) >= MAX_MODELS_PER_WEEK and modelo not in MISSING:
                return False
        # Checar que no sature ningun stage
        cat = CATALOGO.get(modelo, {})
        for e in ETAPAS_FULL:
            if e in cat:
                new_hrs = week_hrs[sem].get(e, 0) + pares * cat[e] / 3600
                if new_hrs > CAP[e] * 0.85:  # max 85%
                    return False
        return True

    # === PASO 1: Clasificar modelos ===
    # Separar por volumen: grandes (>2500), medianos (1500-2500), chicos (<1500)
    modelos_cat = [m for m in MODELOS_ORDER if m not in MISSING]
    modelos_miss = [m for m in MODELOS_ORDER if m in MISSING]

    # Ordenar por "peso" (bottleneck ratio) descendente - los mas pesados primero
    modelos_cat.sort(key=lambda m: model_bottleneck_hrs(m, DEMANDA[m]), reverse=True)

    remaining = dict(DEMANDA)

    # === PASO 2: Asignar modelos pesados (>2500 pares) en 2 semanas ===
    for m in modelos_cat:
        total = remaining[m]
        if total <= 0:
            continue

        # Decidir en cuantas semanas: basado en volumen y peso
        bottleneck = model_bottleneck_hrs(m, total)
        if total > 3000 or bottleneck > TARGET_UTIL * 1.5:
            n_weeks = 3
        elif total > 1500 or bottleneck > TARGET_UTIL:
            n_weeks = 2
        else:
            n_weeks = 1

        n_weeks = min(n_weeks, len(SEMANAS))

        # Elegir las n_weeks semanas con menor bottleneck que puedan recibir este modelo
        candidates = sorted(SEMANAS, key=lambda s: (week_bottleneck(s), week_pares[s]))

        chosen = []
        for s in candidates:
            if len(chosen) >= n_weeks:
                break
            # Verificar que cabe (al menos lote minimo)
            test_pares = max(MIN_BATCH, total // n_weeks)
            if week_can_fit(m, test_pares, s):
                chosen.append(s)

        # Si no encontramos suficientes, forzar
        if len(chosen) < n_weeks:
            for s in candidates:
                if s not in chosen:
                    chosen.append(s)
                if len(chosen) >= n_weeks:
                    break

        chosen.sort()

        # Repartir equitativamente en lotes de 100, redondeando
        per_week = (total // len(chosen) // 100) * 100
        per_week = max(MIN_BATCH, per_week)

        for i, s in enumerate(chosen):
            if i == len(chosen) - 1:
                # Ultima: todo lo restante
                assign = remaining[m]
            else:
                assign = min(per_week, remaining[m])

            if assign > 0:
                add_to_week(m, assign, s)
                remaining[m] -= assign

        # Si quedo sobrante (por redondeo)
        if remaining[m] > 0:
            s = chosen[-1]
            add_to_week(m, remaining[m], s)
            remaining[m] = 0

    # === PASO 3: Modelos sin catalogo en 2-3 semanas ===
    for m in modelos_miss:
        total = remaining[m]
        if total <= 0:
            continue
        n_weeks = 2 if total <= 1200 else 3
        candidates = sorted(SEMANAS, key=lambda s: week_pares[s])
        chosen = candidates[:n_weeks]
        chosen.sort()

        per_week = (total // len(chosen) // 100) * 100
        for i, s in enumerate(chosen):
            if i == len(chosen) - 1:
                assign = remaining[m]
            else:
                assign = min(per_week, remaining[m])
            if assign > 0:
                add_to_week(m, assign, s)
                remaining[m] -= assign
        if remaining[m] > 0:
            add_to_week(m, remaining[m], chosen[-1])
            remaining[m] = 0

    return dist, week_hrs, week_pares

# === Backlog MANUAL (desglose del usuario) ===
BACKLOG_MANUAL = {
    "68127 NE/RO SLI":    {18: 1300, 20: 2000, 21: 600},
    "64197 NE SLI":        {16: 3200},
    "65568 RO/HU SLI":    {15: 400, 16: 2900},
    "77525 NE TEX":        {15: 500, 17: 3000},
    "94750 AA SLI":        {15: 500, 17: 2500},
    "61747 GC SLI":        {15: 300, 18: 1500},
    "65422 HU SLI":        {15: 300, 16: 1700},
    "88186 NE SLI":        {15: 300, 18: 1500},
    "62100 BL SLI":        {15: 300, 16: 300, 17: 300, 18: 300, 19: 300, 20: 300},
    "61748 KR SLI":        {15: 300, 17: 1200},
    "65413 NE/CG/VI SLI": {15: 700, 18: 900},
    "93347 CA SLI":        {19: 1500},
    "69906 BL SLI":        {15: 300, 16: 300, 17: 300, 18: 300},
    "93349 RS SLI":        {19: 400, 20: 400, 21: 400},
    "93346 NE SLI":        {19: 400, 20: 400, 21: 400},
}

# Ejecutar optimizacion
BACKLOG_PROPUESTA, week_hrs_calc, week_pares_calc = optimize_distribution()

# Imprimir resumen para debug
print("=" * 80)
print("PROPUESTA DE DISTRIBUCION - PARES POR SEMANA")
print("=" * 80)
print(f"{'Modelo':<22} " + " ".join(f"{'S'+str(s):>6}" for s in SEMANAS) + f" {'Total':>7}")
print("-" * 80)
total_row = {s: 0 for s in SEMANAS}
for m in MODELOS_ORDER:
    vals = [BACKLOG_PROPUESTA[m].get(s, 0) for s in SEMANAS]
    total_row = {s: total_row[s] + BACKLOG_PROPUESTA[m].get(s, 0) for s in SEMANAS}
    total = sum(vals)
    line = f"{m:<22} " + " ".join(f"{v:>6}" for v in vals) + f" {total:>7}"
    print(line)
print("-" * 80)
print(f"{'TOTAL':<22} " + " ".join(f"{total_row[s]:>6}" for s in SEMANAS) +
      f" {sum(total_row.values()):>7}")

print("\n" + "=" * 80)
print("HORAS POR ETAPA POR SEMANA")
print("=" * 80)
print(f"{'Semana':<10} " + " ".join(f"{e[:6]:>10}" for e in ETAPAS_FULL) + f" {'TOTAL':>10}")
print(f"{'Capacidad':<10} " + " ".join(f"{CAP[e]:>10.0f}" for e in ETAPAS_FULL))
print("-" * 60)
for s in SEMANAS:
    hrs = [week_hrs_calc[s].get(e, 0) for e in ETAPAS_FULL]
    print(f"Sem {s:<6} " + " ".join(f"{h:>10.1f}" for h in hrs) + f" {sum(hrs):>10.1f}")
print("-" * 60)
for e in ETAPAS_FULL:
    total_e = sum(week_hrs_calc[s].get(e, 0) for s in SEMANAS)
    pct = total_e / (CAP[e] * len(SEMANAS)) * 100
    print(f"  {e}: {total_e:.0f} hrs / {CAP[e] * len(SEMANAS)} cap = {pct:.0f}% utilizacion")


# ==============================================================================
# GENERACION DE EXCEL CON FORMULAS
# ==============================================================================
ETAPA_FILLS = {
    "PRELIMINARES": "F59E0B", "ROBOT": "10B981", "POST": "EC4899",
    "N/A PRELIMINAR": "94A3B8", "MAQUILA": "8B5CF6",
}
ETAPA_LIGHT = {
    "PRELIMINARES": "FEF3C7", "ROBOT": "D1FAE5", "POST": "FCE7F3",
    "N/A PRELIMINAR": "E2E8F0", "MAQUILA": "EDE9FE",
}

HEADER_FILL  = PatternFill(start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FILL   = PatternFill(start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid")
TOTAL_FONT   = Font(bold=True, color="FFFFFF", size=10)
MISSING_FILL = PatternFill(start_color="FFFECACA", end_color="FFFECACA", fill_type="solid")
TITLE_FONT   = Font(bold=True, size=14, color="1E3A5F")
PARAM_FONT   = Font(bold=True, size=11, color="1E3A5F")
PARAM_VAL_FONT = Font(bold=True, size=12, color="C0392B")
NUM_FONT     = Font(size=10)
BOLD_FONT    = Font(bold=True, size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT_CENTER  = Alignment(horizontal="left", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
)
PARAM_FILL = PatternFill(start_color="FFFFF3CD", end_color="FFFFF3CD", fill_type="solid")

def sc(cell, font=NUM_FONT, align=CENTER, border=THIN_BORDER, fill=None, fmt=None):
    cell.font = font; cell.alignment = align; cell.border = border
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt

def efill(etapa, light=True):
    c = ETAPA_LIGHT[etapa] if light else ETAPA_FILLS[etapa]
    return PatternFill(start_color="FF" + c, end_color="FF" + c, fill_type="solid")

def cl(col):
    return get_column_letter(col)

# Layout: A=IMG, B=Modelo, C=Pares,
# PRELIM: D=seg/par, E=horas, F=personas, G=dias
# ROBOT:  H=seg/par, I=horas, J=personas, K=dias
# POST:   L=seg/par, M=horas, N=personas, O=dias
# N/A:    P=seg/par, Q=horas
# MAQ:    R=seg/par, S=horas
# TOTAL:  T
IMG_COL = 1
MODEL_COL = 2
PARES_COL = 3
COL = {}
c = 4
for etapa in ETAPAS_FULL:
    COL[etapa] = {"spp": c, "hrs": c+1, "pers": c+2, "dias": c+3}
    c += 4
for etapa in ETAPAS_SIMPLE:
    COL[etapa] = {"spp": c, "hrs": c+1}
    c += 2
COL["TOTAL"] = c
LAST_COL = c
PARAM_ROW = 2

# Imagenes
IMG_DIR = "img_cache"

def get_img_path(modelo_text):
    import re
    m = re.match(r"(\d{5})", str(modelo_text))
    if m:
        path = os.path.join(IMG_DIR, f"{m.group(1)}.png")
        if os.path.exists(path):
            return path
    return None


def build_sheet(ws, title_text, modelos_pares, pares_label="Pares"):
    """Construye hoja con formulas de Excel. Col A=imagen, B=modelo, C=pares"""
    PC = cl(PARES_COL)  # "C"

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)
    ws.cell(row=1, column=1).value = title_text
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Parametros editables (fila 2)
    ws.cell(row=PARAM_ROW, column=MODEL_COL).value = "JORNADA HRS →"
    ws.cell(row=PARAM_ROW, column=MODEL_COL).font = PARAM_FONT
    ws.cell(row=PARAM_ROW, column=MODEL_COL).alignment = LEFT_CENTER
    sc(ws.cell(row=PARAM_ROW, column=PARES_COL, value=JORNADA_DEFAULT),
       font=PARAM_VAL_FONT, fill=PARAM_FILL, fmt="0")

    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        ws.cell(row=PARAM_ROW, column=c_["spp"]).value = "HC →"
        ws.cell(row=PARAM_ROW, column=c_["spp"]).font = PARAM_FONT
        sc(ws.cell(row=PARAM_ROW, column=c_["pers"], value=HEADCOUNT_DEFAULT[etapa]),
           font=PARAM_VAL_FONT, fill=PARAM_FILL, fmt="0")

    # Headers fila 3-4
    ROW_H1, ROW_H2 = 3, 4
    sc(ws.cell(row=ROW_H1, column=IMG_COL), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws.cell(row=ROW_H2, column=IMG_COL), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws.cell(row=ROW_H1, column=MODEL_COL, value="Modelo"), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws.cell(row=ROW_H2, column=MODEL_COL), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws.cell(row=ROW_H1, column=PARES_COL, value=pares_label), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws.cell(row=ROW_H2, column=PARES_COL), font=HEADER_FONT, fill=HEADER_FILL)

    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        ws.merge_cells(start_row=ROW_H1, start_column=c_["spp"], end_row=ROW_H1, end_column=c_["dias"])
        sc(ws.cell(row=ROW_H1, column=c_["spp"], value=etapa), font=HEADER_FONT, fill=HEADER_FILL)
        for cc in range(c_["spp"], c_["dias"]+1):
            ws.cell(row=ROW_H1, column=cc).border = THIN_BORDER
        for cc, lbl in [(c_["spp"], "Seg/Par"), (c_["hrs"], "HORAS"), (c_["pers"], "PERSONAS"), (c_["dias"], "DIAS")]:
            sc(ws.cell(row=ROW_H2, column=cc, value=lbl), font=HEADER_FONT, fill=HEADER_FILL)

    for etapa in ETAPAS_SIMPLE:
        c_ = COL[etapa]
        ws.merge_cells(start_row=ROW_H1, start_column=c_["spp"], end_row=ROW_H1, end_column=c_["hrs"])
        sc(ws.cell(row=ROW_H1, column=c_["spp"], value=etapa), font=HEADER_FONT, fill=HEADER_FILL)
        for cc in [c_["spp"], c_["hrs"]]:
            ws.cell(row=ROW_H1, column=cc).border = THIN_BORDER
        sc(ws.cell(row=ROW_H2, column=c_["spp"], value="Seg/Par"), font=HEADER_FONT, fill=HEADER_FILL)
        sc(ws.cell(row=ROW_H2, column=c_["hrs"], value="HORAS"), font=HEADER_FONT, fill=HEADER_FILL)

    sc(ws.cell(row=ROW_H1, column=COL["TOTAL"], value="TOTAL"), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws.cell(row=ROW_H2, column=COL["TOTAL"], value="HRS"), font=HEADER_FONT, fill=HEADER_FILL)

    # Color bar fila 5
    ROW_COLOR = 5
    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        for cc in range(c_["spp"], c_["dias"]+1):
            ws.cell(row=ROW_COLOR, column=cc).fill = efill(etapa, light=False)
            ws.cell(row=ROW_COLOR, column=cc).border = THIN_BORDER
    for etapa in ETAPAS_SIMPLE:
        c_ = COL[etapa]
        for cc in [c_["spp"], c_["hrs"]]:
            ws.cell(row=ROW_COLOR, column=cc).fill = efill(etapa, light=False)
            ws.cell(row=ROW_COLOR, column=cc).border = THIN_BORDER
    for cc in [IMG_COL, MODEL_COL, PARES_COL, COL["TOTAL"]]:
        ws.cell(row=ROW_COLOR, column=cc).border = THIN_BORDER

    # Datos (fila 6+)
    DATA_START = 6
    jornada_ref = f"${PC}${PARAM_ROW}"
    row = DATA_START

    for modelo, pares in modelos_pares:
        is_miss = modelo in MISSING

        # Imagen
        img_path = get_img_path(modelo)
        if img_path:
            img = XlImage(img_path)
            img.width = 55; img.height = 38
            ws.add_image(img, f"A{row}")
            ws.row_dimensions[row].height = 33

        sc(ws.cell(row=row, column=MODEL_COL, value=modelo), font=BOLD_FONT,
           fill=MISSING_FILL if is_miss else None)
        ws.cell(row=row, column=MODEL_COL).alignment = LEFT_CENTER
        sc(ws.cell(row=row, column=PARES_COL, value=pares), fmt="#,##0",
           fill=MISSING_FILL if is_miss else None)

        if is_miss:
            first = COL[ETAPAS_FULL[0]]["spp"]
            last = COL[ETAPAS_SIMPLE[-1]]["hrs"]
            for cc in range(first, LAST_COL + 1):
                sc(ws.cell(row=row, column=cc), fill=MISSING_FILL)
            ws.merge_cells(start_row=row, start_column=first, end_row=row, end_column=last)
            ws.cell(row=row, column=first).value = "SIN CATALOGO"
            ws.cell(row=row, column=first).font = Font(italic=True, color="991B1B", size=10)
            ws.cell(row=row, column=first).alignment = CENTER
            ws.cell(row=row, column=COL["TOTAL"]).value = "—"
            row += 1
            continue

        cat = CATALOGO[modelo]
        r = str(row)
        hrs_cells = []

        for etapa in ETAPAS_FULL:
            c_ = COL[etapa]
            spp_c = cl(c_["spp"]); hrs_c = cl(c_["hrs"])
            pers_c = cl(c_["pers"])
            hc_ref = f"${cl(c_['pers'])}${PARAM_ROW}"

            if etapa in cat:
                sc(ws.cell(row=row, column=c_["spp"], value=cat[etapa]), fill=efill(etapa), fmt="0")
                ws.cell(row=row, column=c_["hrs"]).value = f"={PC}{r}*{spp_c}{r}/3600"
                sc(ws.cell(row=row, column=c_["hrs"]), fill=efill(etapa), fmt="#,##0.0")
                ws.cell(row=row, column=c_["pers"]).value = f"={hrs_c}{r}/{hc_ref}"
                sc(ws.cell(row=row, column=c_["pers"]), fill=efill(etapa), fmt="#,##0.0")
                ws.cell(row=row, column=c_["dias"]).value = f"={pers_c}{r}/{jornada_ref}"
                sc(ws.cell(row=row, column=c_["dias"]), fill=efill(etapa), fmt="#,##0.0")
                hrs_cells.append(f"{hrs_c}{r}")
            else:
                for cc in [c_["spp"], c_["hrs"], c_["pers"], c_["dias"]]:
                    sc(ws.cell(row=row, column=cc, value="—"))

        for etapa in ETAPAS_SIMPLE:
            c_ = COL[etapa]
            spp_c = cl(c_["spp"]); hrs_c = cl(c_["hrs"])
            if etapa in cat:
                sc(ws.cell(row=row, column=c_["spp"], value=cat[etapa]), fill=efill(etapa), fmt="0")
                ws.cell(row=row, column=c_["hrs"]).value = f"={PC}{r}*{spp_c}{r}/3600"
                sc(ws.cell(row=row, column=c_["hrs"]), fill=efill(etapa), fmt="#,##0.0")
                hrs_cells.append(f"{hrs_c}{r}")
            else:
                sc(ws.cell(row=row, column=c_["spp"], value="—"))
                sc(ws.cell(row=row, column=c_["hrs"], value="—"))

        if hrs_cells:
            ws.cell(row=row, column=COL["TOTAL"]).value = "=" + "+".join(hrs_cells)
            sc(ws.cell(row=row, column=COL["TOTAL"]), font=BOLD_FONT, fmt="#,##0.0")
        row += 1

    # Fila TOTAL con SUM
    tr = row
    for cc in range(1, LAST_COL + 1):
        sc(ws.cell(row=tr, column=cc), font=TOTAL_FONT, fill=TOTAL_FILL)
    ws.cell(row=tr, column=MODEL_COL).value = "TOTAL"
    ws.cell(row=tr, column=PARES_COL).value = f"=SUM({PC}{DATA_START}:{PC}{tr-1})"
    ws.cell(row=tr, column=PARES_COL).number_format = "#,##0"

    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        for cc in [c_["hrs"], c_["pers"], c_["dias"]]:
            ws.cell(row=tr, column=cc).value = f"=SUM({cl(cc)}{DATA_START}:{cl(cc)}{tr-1})"
            ws.cell(row=tr, column=cc).number_format = "#,##0.0"
    for etapa in ETAPAS_SIMPLE:
        c_ = COL[etapa]
        ws.cell(row=tr, column=c_["hrs"]).value = f"=SUM({cl(c_['hrs'])}{DATA_START}:{cl(c_['hrs'])}{tr-1})"
        ws.cell(row=tr, column=c_["hrs"]).number_format = "#,##0.0"
    ws.cell(row=tr, column=COL["TOTAL"]).value = f"=SUM({cl(COL['TOTAL'])}{DATA_START}:{cl(COL['TOTAL'])}{tr-1})"
    ws.cell(row=tr, column=COL["TOTAL"]).number_format = "#,##0.0"

    # Anchos
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions[cl(MODEL_COL)].width = 22
    ws.column_dimensions[cl(PARES_COL)].width = 11
    for col in range(PARES_COL + 1, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = f"A{DATA_START}"

    return DATA_START, tr


def generate_workbook(backlog_data, output_file, title_prefix, subtitle=""):
    """Genera un workbook completo a partir de un backlog dado"""
    wb = openpyxl.Workbook()

    # === HOJA 1: Backlog (pares por semana) con col A=img ===
    ws_bl = wb.active
    ws_bl.title = "Backlog"

    # Col A=img, B=modelo, C-I=semanas, J=total
    BL_MODEL = 2; BL_FIRST_SEM = 3; BL_TOTAL = BL_FIRST_SEM + len(SEMANAS)

    ws_bl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=BL_TOTAL)
    ws_bl.cell(row=1, column=1).value = f"{title_prefix} — BACKLOG ANDREA (SEM 15-21)"
    ws_bl.cell(row=1, column=1).font = TITLE_FONT
    ws_bl.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    if subtitle:
        ws_bl.merge_cells(start_row=2, start_column=1, end_row=2, end_column=BL_TOTAL)
        ws_bl.cell(row=2, column=1).value = subtitle
        ws_bl.cell(row=2, column=1).font = Font(italic=True, size=10, color="64748B")
        ws_bl.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    bl_headers = ["", "Modelo"] + [f"Sem {s}" for s in SEMANAS] + ["Total"]
    for col, h in enumerate(bl_headers, 1):
        sc(ws_bl.cell(row=4, column=col, value=h), font=HEADER_FONT, fill=HEADER_FILL)

    row = 5
    for m in MODELOS_ORDER:
        # Imagen
        img_path = get_img_path(m)
        if img_path:
            img = XlImage(img_path)
            img.width = 55; img.height = 38
            ws_bl.add_image(img, f"A{row}")
            ws_bl.row_dimensions[row].height = 33

        sc(ws_bl.cell(row=row, column=BL_MODEL, value=m), font=BOLD_FONT,
           fill=MISSING_FILL if m in MISSING else None)
        ws_bl.cell(row=row, column=BL_MODEL).alignment = LEFT_CENTER

        for i, s in enumerate(SEMANAS):
            val = backlog_data[m].get(s, 0)
            cell = ws_bl.cell(row=row, column=BL_FIRST_SEM + i, value=val if val > 0 else "")
            sc(cell, fmt="#,##0", fill=MISSING_FILL if m in MISSING else None)

        ws_bl.cell(row=row, column=BL_TOTAL).value = \
            f"=SUM({cl(BL_FIRST_SEM)}{row}:{cl(BL_FIRST_SEM + len(SEMANAS) - 1)}{row})"
        sc(ws_bl.cell(row=row, column=BL_TOTAL), font=BOLD_FONT, fmt="#,##0",
           fill=MISSING_FILL if m in MISSING else None)
        row += 1

    for cc in range(1, BL_TOTAL + 1):
        sc(ws_bl.cell(row=row, column=cc), font=TOTAL_FONT, fill=TOTAL_FILL)
    ws_bl.cell(row=row, column=BL_MODEL).value = "TOTAL"
    for col in range(BL_FIRST_SEM, BL_TOTAL + 1):
        ws_bl.cell(row=row, column=col).value = f"=SUM({cl(col)}5:{cl(col)}{row-1})"
        ws_bl.cell(row=row, column=col).number_format = "#,##0"

    ws_bl.column_dimensions["A"].width = 9
    ws_bl.column_dimensions[cl(BL_MODEL)].width = 22
    for col in range(BL_FIRST_SEM, BL_TOTAL + 1):
        ws_bl.column_dimensions[cl(col)].width = 10
    ws_bl.freeze_panes = "A5"

    # === HOJA: Total General ===
    ws_tot = wb.create_sheet("Total General")
    modelos_total = [(m, sum(backlog_data[m].values())) for m in MODELOS_ORDER]
    build_sheet(ws_tot, f"MATRIZ DE HORAS POR ETAPA — TOTAL ACUMULADO (SEM 15-21)",
                modelos_total, "Total Pares")

    # === HOJAS POR SEMANA ===
    for sem in SEMANAS:
        ws_s = wb.create_sheet(f"Sem {sem}")
        modelos_sem = [(m, backlog_data[m][sem]) for m in MODELOS_ORDER
                       if backlog_data[m].get(sem, 0) > 0]
        build_sheet(ws_s, f"HORAS POR ETAPA — SEMANA {sem}", modelos_sem, "Pares Sem")

    # === HOJA RESUMEN SEMANAL ===
    ws_r = wb.create_sheet("Resumen Semanal")
    ws_r.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)
    ws_r["A1"] = "RESUMEN SEMANAL — HORAS POR ETAPA"
    ws_r["A1"].font = TITLE_FONT
    ws_r["A1"].alignment = Alignment(horizontal="center")

    # Resumen: no tiene col imagen, usa col 1=Semana, col 2=Pares directamente
    # pero las etapas usan las mismas posiciones COL[]
    PC_R = cl(PARES_COL)  # "C" for pares in resumen
    ws_r.cell(row=PARAM_ROW, column=MODEL_COL).value = "JORNADA HRS →"
    ws_r.cell(row=PARAM_ROW, column=MODEL_COL).font = PARAM_FONT
    ws_r.cell(row=PARAM_ROW, column=MODEL_COL).alignment = LEFT_CENTER
    sc(ws_r.cell(row=PARAM_ROW, column=PARES_COL, value=JORNADA_DEFAULT),
       font=PARAM_VAL_FONT, fill=PARAM_FILL, fmt="0")
    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        ws_r.cell(row=PARAM_ROW, column=c_["spp"]).value = "HC →"
        ws_r.cell(row=PARAM_ROW, column=c_["spp"]).font = PARAM_FONT
        sc(ws_r.cell(row=PARAM_ROW, column=c_["pers"], value=HEADCOUNT_DEFAULT[etapa]),
           font=PARAM_VAL_FONT, fill=PARAM_FILL, fmt="0")

    for r_h in [3, 4]:
        sc(ws_r.cell(row=r_h, column=IMG_COL), font=HEADER_FONT, fill=HEADER_FILL)
        sc(ws_r.cell(row=r_h, column=MODEL_COL, value="Semana" if r_h == 3 else ""), font=HEADER_FONT, fill=HEADER_FILL)
        sc(ws_r.cell(row=r_h, column=PARES_COL, value="Pares" if r_h == 3 else ""), font=HEADER_FONT, fill=HEADER_FILL)
    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        ws_r.merge_cells(start_row=3, start_column=c_["spp"], end_row=3, end_column=c_["dias"])
        sc(ws_r.cell(row=3, column=c_["spp"], value=etapa), font=HEADER_FONT, fill=HEADER_FILL)
        for cc in range(c_["spp"], c_["dias"]+1):
            ws_r.cell(row=3, column=cc).border = THIN_BORDER
        for cc, lbl in [(c_["spp"], ""), (c_["hrs"], "HORAS"), (c_["pers"], "PERSONAS"), (c_["dias"], "DIAS")]:
            sc(ws_r.cell(row=4, column=cc, value=lbl), font=HEADER_FONT, fill=HEADER_FILL)
    for etapa in ETAPAS_SIMPLE:
        c_ = COL[etapa]
        ws_r.merge_cells(start_row=3, start_column=c_["spp"], end_row=3, end_column=c_["hrs"])
        sc(ws_r.cell(row=3, column=c_["spp"], value=etapa), font=HEADER_FONT, fill=HEADER_FILL)
        for cc in [c_["spp"], c_["hrs"]]:
            ws_r.cell(row=3, column=cc).border = THIN_BORDER
        sc(ws_r.cell(row=4, column=c_["spp"], value=""), font=HEADER_FONT, fill=HEADER_FILL)
        sc(ws_r.cell(row=4, column=c_["hrs"], value="HORAS"), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws_r.cell(row=3, column=COL["TOTAL"], value="TOTAL"), font=HEADER_FONT, fill=HEADER_FILL)
    sc(ws_r.cell(row=4, column=COL["TOTAL"], value="HRS"), font=HEADER_FONT, fill=HEADER_FILL)

    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        for cc in range(c_["spp"], c_["dias"]+1):
            ws_r.cell(row=5, column=cc).fill = efill(etapa, light=False)
            ws_r.cell(row=5, column=cc).border = THIN_BORDER
    for etapa in ETAPAS_SIMPLE:
        c_ = COL[etapa]
        for cc in [c_["spp"], c_["hrs"]]:
            ws_r.cell(row=5, column=cc).fill = efill(etapa, light=False)
            ws_r.cell(row=5, column=cc).border = THIN_BORDER
    for cc in [IMG_COL, MODEL_COL, PARES_COL, COL["TOTAL"]]:
        ws_r.cell(row=5, column=cc).border = THIN_BORDER

    jornada_ref = f"${PC_R}${PARAM_ROW}"
    row = 6
    for sem in SEMANAS:
        sheet_name = f"Sem {sem}"
        ws_sem = wb[sheet_name]
        total_row_sem = None
        for r in range(6, 30):
            v = ws_sem.cell(row=r, column=MODEL_COL).value
            if v and "TOTAL" in str(v):
                total_row_sem = r
                break

        sc(ws_r.cell(row=row, column=MODEL_COL, value=f"Sem {sem}"), font=BOLD_FONT)
        ws_r.cell(row=row, column=PARES_COL).value = f"='{sheet_name}'!{PC_R}{total_row_sem}"
        sc(ws_r.cell(row=row, column=PARES_COL), fmt="#,##0")

        hrs_cells = []
        for etapa in ETAPAS_FULL:
            c_ = COL[etapa]
            hrs_c = cl(c_["hrs"]); pers_c = cl(c_["pers"])
            hc_ref = f"${cl(c_['pers'])}${PARAM_ROW}"
            ws_r.cell(row=row, column=c_["spp"]).border = THIN_BORDER
            ws_r.cell(row=row, column=c_["hrs"]).value = f"='{sheet_name}'!{hrs_c}{total_row_sem}"
            sc(ws_r.cell(row=row, column=c_["hrs"]), fill=efill(etapa), fmt="#,##0.0")
            ws_r.cell(row=row, column=c_["pers"]).value = f"={hrs_c}{row}/{hc_ref}"
            sc(ws_r.cell(row=row, column=c_["pers"]), fill=efill(etapa), fmt="#,##0.0")
            ws_r.cell(row=row, column=c_["dias"]).value = f"={pers_c}{row}/{jornada_ref}"
            sc(ws_r.cell(row=row, column=c_["dias"]), fill=efill(etapa), fmt="#,##0.0")
            hrs_cells.append(f"{hrs_c}{row}")

        for etapa in ETAPAS_SIMPLE:
            c_ = COL[etapa]
            hrs_c = cl(c_["hrs"])
            ws_r.cell(row=row, column=c_["spp"]).border = THIN_BORDER
            ws_r.cell(row=row, column=c_["hrs"]).value = f"='{sheet_name}'!{hrs_c}{total_row_sem}"
            sc(ws_r.cell(row=row, column=c_["hrs"]), fill=efill(etapa), fmt="#,##0.0")
            hrs_cells.append(f"{hrs_c}{row}")

        ws_r.cell(row=row, column=COL["TOTAL"]).value = "=" + "+".join(hrs_cells)
        sc(ws_r.cell(row=row, column=COL["TOTAL"]), font=BOLD_FONT, fmt="#,##0.0")
        row += 1

    tr = row
    for cc in range(1, LAST_COL + 1):
        sc(ws_r.cell(row=tr, column=cc), font=TOTAL_FONT, fill=TOTAL_FILL)
    ws_r.cell(row=tr, column=MODEL_COL).value = "TOTAL"
    ws_r.cell(row=tr, column=PARES_COL).value = f"=SUM({PC_R}6:{PC_R}{tr-1})"
    ws_r.cell(row=tr, column=PARES_COL).number_format = "#,##0"
    for etapa in ETAPAS_FULL:
        c_ = COL[etapa]
        for cc in [c_["hrs"], c_["pers"], c_["dias"]]:
            ws_r.cell(row=tr, column=cc).value = f"=SUM({cl(cc)}6:{cl(cc)}{tr-1})"
            ws_r.cell(row=tr, column=cc).number_format = "#,##0.0"
    for etapa in ETAPAS_SIMPLE:
        c_ = COL[etapa]
        ws_r.cell(row=tr, column=c_["hrs"]).value = f"=SUM({cl(c_['hrs'])}6:{cl(c_['hrs'])}{tr-1})"
        ws_r.cell(row=tr, column=c_["hrs"]).number_format = "#,##0.0"
    ws_r.cell(row=tr, column=COL["TOTAL"]).value = f"=SUM({cl(COL['TOTAL'])}6:{cl(COL['TOTAL'])}{tr-1})"
    ws_r.cell(row=tr, column=COL["TOTAL"]).number_format = "#,##0.0"

    ws_r.column_dimensions["A"].width = 4
    ws_r.column_dimensions[cl(MODEL_COL)].width = 12
    ws_r.column_dimensions[cl(PARES_COL)].width = 11
    for col in range(PARES_COL + 1, LAST_COL + 1):
        ws_r.column_dimensions[get_column_letter(col)].width = 11
    ws_r.freeze_panes = "A6"

    # === HOJA: Seg/Par (con col imagen) ===
    ws2 = wb.create_sheet("Seg por Par")
    h2 = ["", "Modelo", "Total Seg/Par"] + ETAPAS_ALL + ["Pares/Hr"]
    NUM_H2 = len(h2)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_H2)
    ws2.cell(row=1, column=1).value = "SEGUNDOS POR PAR POR ETAPA (DATOS CATALOGO)"
    ws2.cell(row=1, column=1).font = TITLE_FONT
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for col, h in enumerate(h2, 1):
        sc(ws2.cell(row=3, column=col, value=h), font=HEADER_FONT, fill=HEADER_FILL)
        ws2.cell(row=3, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 4
    SPP_MODEL = 2; SPP_TOTAL = 3; SPP_FIRST_ETAPA = 4
    for modelo in MODELOS_ORDER:
        if modelo in MISSING:
            continue
        cat = CATALOGO[modelo]
        # Imagen
        img_path = get_img_path(modelo)
        if img_path:
            img = XlImage(img_path)
            img.width = 55; img.height = 38
            ws2.add_image(img, f"A{row}")
            ws2.row_dimensions[row].height = 33

        sc(ws2.cell(row=row, column=SPP_MODEL, value=modelo), font=BOLD_FONT)
        ws2.cell(row=row, column=SPP_MODEL).alignment = LEFT_CENTER
        spp_cells = []
        for col, etapa in enumerate(ETAPAS_ALL, SPP_FIRST_ETAPA):
            cell = ws2.cell(row=row, column=col)
            if etapa in cat:
                cell.value = cat[etapa]
                sc(cell, fill=efill(etapa))
                spp_cells.append(f"{cl(col)}{row}")
            else:
                cell.value = 0
                sc(cell)
        ws2.cell(row=row, column=SPP_TOTAL).value = "=" + "+".join(spp_cells) if spp_cells else 0
        sc(ws2.cell(row=row, column=SPP_TOTAL), fmt="0")
        ws2.cell(row=row, column=NUM_H2).value = f"=IF({cl(SPP_TOTAL)}{row}>0,3600/{cl(SPP_TOTAL)}{row},0)"
        sc(ws2.cell(row=row, column=NUM_H2), font=BOLD_FONT, fmt="#,##0.0")
        row += 1

    ws2.column_dimensions["A"].width = 9
    ws2.column_dimensions[cl(SPP_MODEL)].width = 24
    ws2.column_dimensions[cl(SPP_TOTAL)].width = 14
    for col in range(SPP_FIRST_ETAPA, NUM_H2 + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # Guardar
    wb.save(output_file)
    return output_file


# ==============================================================================
# GENERAR AMBOS ARCHIVOS
# ==============================================================================

# 1) Excel con desglose MANUAL
f1 = generate_workbook(
    BACKLOG_MANUAL,
    "Matriz_Horas_Backlog_Andrea_Manual.xlsx",
    "DESGLOSE MANUAL",
    "Distribucion de pares definida manualmente"
)
print(f"Archivo generado: {f1}")

# 2) Excel con desglose PROPUESTA optimizada
f2 = generate_workbook(
    BACKLOG_PROPUESTA,
    "Matriz_Horas_Backlog_Andrea_Propuesta.xlsx",
    "PROPUESTA OPTIMIZADA",
    "Distribucion balanceando carga PRELIM/ROBOT/POST por semana"
)
print(f"Archivo generado: {f2}")

# 3) Agregar hoja "Backlog Original" del archivo fuente a ambos
from copy import copy
SOURCE_FILE = "BACKLOG ANDREA SEM14 (1).xlsx"
if os.path.exists(SOURCE_FILE):
    source = openpyxl.load_workbook(SOURCE_FILE)
    ws_src = source["Hoja1 (2)"]
    for target_file in [f1, f2]:
        wb = openpyxl.load_workbook(target_file)
        ws_new = wb.create_sheet("Backlog Original", 0)
        for row in range(1, ws_src.max_row + 1):
            for col in range(1, ws_src.max_column + 1):
                src_cell = ws_src.cell(row=row, column=col)
                dst_cell = ws_new.cell(row=row, column=col, value=src_cell.value)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
        for merge in ws_src.merged_cells.ranges:
            ws_new.merge_cells(str(merge))
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = dim.width
        for row_num, dim in ws_src.row_dimensions.items():
            ws_new.row_dimensions[row_num].height = dim.height
        wb.save(target_file)
        print(f"  + Backlog Original agregado a {target_file}")

print(f"\nAmbos con formulas editables, columna de imagenes y Backlog Original")
