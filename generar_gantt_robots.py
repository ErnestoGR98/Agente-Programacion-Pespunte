"""
Genera Excel con Gantt secuencial para programacion de robots faltantes.
Fecha de inicio: 16 de abril de 2026.
1 programador, 1 programa a la vez, ~1 hora por programa.
Horario: 8:00-16:00 (8h/dia), Lunes a Sabado.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from datetime import datetime, timedelta

# ── Programas faltantes (impacto alto) ──
FALTANTES = [
    ("61748", 6, "COSTURA DE CHINELA EXTERNA", "2A-3020-M1"),
    ("94750", 4, "COSTURA DE TALON", "2A-3020-M1"),
    ("61748", 5, "COSTURA DE CHINELA INTERNA", "2A-3020-M2"),
    ("61748", 6, "COSTURA DE CHINELA EXTERNA", "2A-3020-M2"),
    ("65413", 2, "COSTURA DE LATIGO", "2A-3020-M2"),
    ("65422", 1, "COSER CHINELAS CHICAS", "2A-3020-M2"),
    ("65422", 8, "COSTURA DE CHINELA LASER", "2A-3020-M2"),
    ("65568", 1, "COSTURA CHINELA", "6040-M4"),
    ("94750", 4, "COSTURA DE TALON", "6040-M4"),
    ("65413", 3, "COSTURA DE HEBILLA", "6040-M5"),
    ("65422", 8, "COSTURA DE CHINELA LASER", "6040-M5"),
    ("65568", 1, "COSTURA CHINELA", "6040-M5"),
    ("94750", 4, "COSTURA DE TALON", "6040-M5"),
    ("64197", 1, "COSTURA DE CHINELA", "M048-CHACHE"),
    ("65413", 1, "COSTURA DE CHINELA", "M048-CHACHE"),
    ("88186", 1, "COSTURA DE CHINELA", "M048-CHACHE"),
    ("88186", 4, "COSTURA DE GANCHOS", "M048-CHACHE"),
    ("61747", 4, "COSTURA DE FELPA Y GANCHO", "M049-CHACHE"),
    ("61748", 4, "COSTURA DE TALON", "M049-CHACHE"),
    ("64197", 1, "COSTURA DE CHINELA", "M049-CHACHE"),
    ("88186", 4, "COSTURA DE GANCHOS", "M049-CHACHE"),
    ("94750", 4, "COSTURA DE TALON", "M049-CHACHE"),
]

por_robot = OrderedDict()
for modelo, fracc, op, robot in FALTANTES:
    por_robot.setdefault(robot, []).append((modelo, fracc, op))

# ── Config ──
FECHA_INICIO = datetime(2026, 4, 16)  # Jueves 16 de abril 2026
HORAS_POR_DIA = 8
HORA_INICIO = 8  # 8:00 AM
DIAS_LABORALES = {0, 1, 2, 3, 4, 5}  # Lun(0)-Sab(5)
DIAS_NOMBRE = {0: "Lun", 1: "Mar", 2: "Mie", 3: "Jue", 4: "Vie", 5: "Sab"}

total_horas = len(FALTANTES)


def calcular_fechas(fecha_inicio, num_tareas):
    """Calcula fecha y hora de inicio/fin de cada tarea respetando horario laboral."""
    fechas = []
    current = fecha_inicio
    # Avanzar al primer dia laboral si es necesario
    while current.weekday() not in DIAS_LABORALES:
        current += timedelta(days=1)

    hora_slot = 0  # slot dentro del dia (0..7)
    for i in range(num_tareas):
        inicio = current.replace(hour=HORA_INICIO + hora_slot, minute=0)
        fin = current.replace(hour=HORA_INICIO + hora_slot + 1, minute=0)
        fechas.append((inicio, fin))

        hora_slot += 1
        if hora_slot >= HORAS_POR_DIA:
            hora_slot = 0
            current += timedelta(days=1)
            while current.weekday() not in DIAS_LABORALES:
                current += timedelta(days=1)

    return fechas


fechas = calcular_fechas(FECHA_INICIO, total_horas)

# Calcular dias unicos para el Gantt
dias_unicos = list(OrderedDict.fromkeys(f.date() for f, _ in fechas))
total_dias = len(dias_unicos)

COLORES_ROBOT = {
    "2A-3020-M1": "4472C4",
    "2A-3020-M2": "5B9BD5",
    "6040-M4":    "70AD47",
    "6040-M5":    "A9D18E",
    "M048-CHACHE": "ED7D31",
    "M049-CHACHE": "FFC000",
}

# Estilos
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
LIGHT_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def get_font_color(hex_color):
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    lum = (r * 299 + g * 587 + b * 114) / 1000
    return "FFFFFF" if lum < 150 else "000000"


wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# HOJA 1: GANTT SECUENCIAL CON FECHAS
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Gantt Secuencial"

# Titulo
ws.cell(row=1, column=1,
        value="GANTT SECUENCIAL - PROGRAMACION DE ROBOTS FALTANTES").font = Font(bold=True, size=14)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + total_dias * HORAS_POR_DIA)

fecha_inicio_str = FECHA_INICIO.strftime("%d/%m/%Y")
fecha_fin_str = fechas[-1][1].strftime("%d/%m/%Y")
ws.cell(row=2, column=1,
        value=f"Inicio: {fecha_inicio_str} | Fin: {fecha_fin_str} | "
              f"{total_horas} programas x 1 hora = {total_horas} hrs = {total_dias} dias laborales (8h/dia)"
        ).font = Font(size=11, italic=True)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5 + total_dias * HORAS_POR_DIA)

# Fila 4: Headers de dia con fecha real
for d_idx, dia_date in enumerate(dias_unicos):
    start_col = 6 + d_idx * HORAS_POR_DIA
    end_col = start_col + HORAS_POR_DIA - 1
    dia_nombre = DIAS_NOMBRE.get(dia_date.weekday(), "")
    label = f"{dia_nombre} {dia_date.strftime('%d/%m/%Y')}"
    cell = ws.cell(row=4, column=start_col, value=label)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    cell.alignment = CENTER
    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)

# Fila 5: Headers tabla + horas
table_headers = ["#", "ROBOT", "MODELO", "FRACC", "OPERACION"]
for c, h in enumerate(table_headers, 1):
    cell = ws.cell(row=5, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for d_idx in range(total_dias):
    for h in range(HORAS_POR_DIA):
        col = 6 + d_idx * HORAS_POR_DIA + h
        hora_real = HORA_INICIO + h
        cell = ws.cell(row=5, column=col, value=f"{hora_real}:00")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

# Data rows
for i, (modelo, fracc, op, robot) in enumerate(FALTANTES):
    r = 6 + i
    color = COLORES_ROBOT.get(robot, "BDD7EE")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font_color = get_font_color(color)

    ws.cell(row=r, column=1, value=i + 1).font = NORMAL_FONT
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).border = THIN_BORDER

    cell_robot = ws.cell(row=r, column=2, value=robot)
    cell_robot.font = Font(bold=True, color=font_color)
    cell_robot.fill = fill
    cell_robot.border = THIN_BORDER

    ws.cell(row=r, column=3, value=modelo).font = NORMAL_FONT
    ws.cell(row=r, column=3).alignment = CENTER
    ws.cell(row=r, column=3).border = THIN_BORDER

    ws.cell(row=r, column=4, value=fracc).font = NORMAL_FONT
    ws.cell(row=r, column=4).alignment = CENTER
    ws.cell(row=r, column=4).border = THIN_BORDER

    ws.cell(row=r, column=5, value=op).font = NORMAL_FONT
    ws.cell(row=r, column=5).border = THIN_BORDER

    # Calcular posicion en el gantt
    inicio_tarea, fin_tarea = fechas[i]
    dia_idx = dias_unicos.index(inicio_tarea.date())
    hora_idx = inicio_tarea.hour - HORA_INICIO
    gantt_col = 6 + dia_idx * HORAS_POR_DIA + hora_idx

    # Pintar todas las celdas del gantt (fondo vacio)
    for h in range(total_dias * HORAS_POR_DIA):
        col = 6 + h
        cell = ws.cell(row=r, column=col)
        cell.border = LIGHT_BORDER

    # Barra del gantt
    cell = ws.cell(row=r, column=gantt_col)
    cell.fill = fill
    cell.font = Font(bold=True, size=7, color=font_color)
    cell.value = f"{modelo}-F{fracc}"
    cell.alignment = CENTER

# ── Resumen por robot ──
row_num = 6 + len(FALTANTES) + 2
ws.cell(row=row_num, column=1, value="RESUMEN POR ROBOT:").font = Font(bold=True, size=12)
row_num += 1
for robot, tareas in por_robot.items():
    color = COLORES_ROBOT.get(robot, "BDD7EE")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font_color = get_font_color(color)
    cell = ws.cell(row=row_num, column=1)
    cell.fill = fill
    cell_name = ws.cell(row=row_num, column=2, value=robot)
    cell_name.font = Font(bold=True, color=font_color)
    cell_name.fill = fill
    ws.cell(row=row_num, column=3, value=f"{len(tareas)} programa(s) = {len(tareas)} hora(s)").font = NORMAL_FONT
    row_num += 1

row_num += 1
ws.cell(row=row_num, column=1, value="TOTAL:").font = Font(bold=True, size=12)
ws.cell(row=row_num, column=3,
        value=f"{total_horas} programas = {total_horas} horas = {total_dias} dias laborales").font = BOLD_FONT

# Column widths
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 7
ws.column_dimensions["E"].width = 32
for h in range(total_dias * HORAS_POR_DIA):
    ws.column_dimensions[get_column_letter(6 + h)].width = 12

ws.freeze_panes = "F6"

# ═══════════════════════════════════════════════
# HOJA 2: CHECKLIST CON FECHAS REALES
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("Checklist")

ws2.cell(row=1, column=1,
         value="CHECKLIST - PROGRAMAS FALTANTES CON FECHA Y HORA").font = Font(bold=True, size=14)
ws2.cell(row=2, column=1,
         value=f"Inicio: {fecha_inicio_str} | {total_horas} programas | {total_dias} dias laborales").font = Font(
    size=11, italic=True)

headers_ck = ["#", "ROBOT", "MODELO", "FRACC", "OPERACION", "FECHA", "DIA", "HORA INICIO", "HORA FIN", "STATUS"]
for c, h in enumerate(headers_ck, 1):
    cell = ws2.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, (modelo, fracc, op, robot) in enumerate(FALTANTES):
    r = 5 + i
    inicio, fin = fechas[i]
    dia_nombre = DIAS_NOMBRE.get(inicio.weekday(), "")
    color = COLORES_ROBOT.get(robot, "BDD7EE")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font_color = get_font_color(color)

    ws2.cell(row=r, column=1, value=i + 1).font = NORMAL_FONT
    ws2.cell(row=r, column=1).alignment = CENTER

    cell_robot = ws2.cell(row=r, column=2, value=robot)
    cell_robot.font = Font(bold=True, color=font_color)
    cell_robot.fill = fill

    ws2.cell(row=r, column=3, value=modelo).font = NORMAL_FONT
    ws2.cell(row=r, column=3).alignment = CENTER
    ws2.cell(row=r, column=4, value=fracc).font = NORMAL_FONT
    ws2.cell(row=r, column=4).alignment = CENTER
    ws2.cell(row=r, column=5, value=op).font = NORMAL_FONT

    ws2.cell(row=r, column=6, value=inicio.strftime("%d/%m/%Y")).font = NORMAL_FONT
    ws2.cell(row=r, column=6).alignment = CENTER
    ws2.cell(row=r, column=7, value=dia_nombre).font = NORMAL_FONT
    ws2.cell(row=r, column=7).alignment = CENTER
    ws2.cell(row=r, column=8, value=inicio.strftime("%H:%M")).font = NORMAL_FONT
    ws2.cell(row=r, column=8).alignment = CENTER
    ws2.cell(row=r, column=9, value=fin.strftime("%H:%M")).font = NORMAL_FONT
    ws2.cell(row=r, column=9).alignment = CENTER
    ws2.cell(row=r, column=10, value="PENDIENTE").font = Font(size=10, color="FF0000")
    ws2.cell(row=r, column=10).alignment = CENTER

    for c in range(1, 11):
        ws2.cell(row=r, column=c).border = THIN_BORDER

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 7
ws2.column_dimensions["E"].width = 30
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 6
ws2.column_dimensions["H"].width = 12
ws2.column_dimensions["I"].width = 10
ws2.column_dimensions["J"].width = 14

OUTPUT = "Gantt_Programacion_Robots.xlsx"
wb.save(OUTPUT)
print(f"\nArchivo generado: {OUTPUT}")
print(f"  {total_horas} programas faltantes")
print(f"  Inicio: {fechas[0][0].strftime('%d/%m/%Y %H:%M')}")
print(f"  Fin:    {fechas[-1][1].strftime('%d/%m/%Y %H:%M')}")
print(f"  {total_dias} dias laborales")
print("\nDetalle:")
for i, (modelo, fracc, op, robot) in enumerate(FALTANTES):
    inicio, fin = fechas[i]
    dia_nombre = DIAS_NOMBRE.get(inicio.weekday(), "")
    print(f"  {i+1:2d}. {robot:<14s} {modelo}-F{fracc} {op:<30s} "
          f"{dia_nombre} {inicio.strftime('%d/%m/%Y %H:%M')}-{fin.strftime('%H:%M')}")
