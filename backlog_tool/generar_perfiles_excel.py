"""Genera Excel con clasificacion de modelos en 3 perfiles (A/B/C) + imagenes."""
import io
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

OUT = os.path.join(os.path.dirname(__file__), "outputs", "perfiles_modelos.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# (perfil, modelo, codigo, ICO, PREL, Robot, POST, NA, Maquila, etapa_cuello, regla, img_url)
DATA = [
    ("A", "62100", "62100 BL SLI", 15.55, 2.23, 2.35, 10.97, 0.00, 0.00, "POST extremo",
     "Cuello conveyor POST. Solo en dias dedicados.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/62100_BL.jpeg"),
    ("A", "61748", "61748 KR SLI", 11.00, 3.00, 3.00, 4.00, 1.00, 0.00, "Robot pesado",
     "3 ops robot largos. No mezclar con 68127.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/61748_KR.jpeg"),
    ("A", "65422", "65422 HU SLI", 8.85, 0.73, 2.58, 4.83, 0.70, 0.00, "Robot + POST",
     "Doble carga. No mezclar con otra ancla POST.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/65422_HU.png"),
    ("A", "68127", "68127 NE SLI", 7.98, 1.80, 4.38, 1.20, 0.60, 0.00, "Robot extremo (7 ops)",
     "Satura robots 3020/6040. No mezclar con 61748.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/68127.avif"),
    ("A", "88186", "88186 NE SLI", 7.22, 2.40, 2.42, 1.80, 0.60, 0.00, "Robot",
     "Robot pesado de fondo.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/88186_NE.jpeg"),

    ("B", "61747", "61747 GC SLI", 7.60, 3.60, 1.00, 3.00, 0.00, 7.07, "POST + Maquila",
     "Parte sale a maquila. Buen acompanante de ancla robot.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/61747_GC.jpeg"),
    ("B", "65568", "65568 RO/HU SLI", 5.12, 1.80, 1.52, 1.80, 0.00, 0.00, "Balanceado",
     "Distribucion equilibrada. Flexible para cualquier dia.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/65568.jpeg"),
    ("B", "69906", "69906 BL SLI", 5.00, 1.50, 1.50, 2.00, 0.00, 0.00, "Estimado (sin catalogo)",
     "Cargar catalogo antes de optimizar.",
     None),
    ("B", "65413", "65413 NE/CG/VI", 4.48, 0.00, 2.08, 2.40, 0.00, 5.12, "Robot + POST ligero",
     "Robot ligero, parte en maquila.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/65413.avif"),
    ("B", "94750", "94750 AA SLI", 3.47, 1.20, 1.67, 0.60, 0.00, 2.00, "Robot + Maquila",
     "Carga interna baja por maquila. Acompanante ideal.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/94750.jpeg"),

    ("C", "64197", "64197 NE/RO SLI", 3.18, 0.00, 0.78, 1.80, 0.60, 0.00, "Simple",
     "Distribuir libre. Filler de fondo.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/64197_NE.jpeg"),
    ("C", "77525", "77525 NE TEX", 1.80, 0.00, 0.00, 1.80, 0.00, 7.80, "95% Maquila externa",
     "Casi cero HC interno. Meter donde haya hueco.",
     "https://folmyddedsdzlbegumbo.supabase.co/storage/v1/object/public/modelos/77525.avif"),
]

PERFIL_INFO = {
    "A": ("ANCLAS", "FFEF4444", "Definen el dia. Maximo 1 ancla POST + 1 ancla Robot por dia. Lotes >=300, skill alto."),
    "B": ("BALANCERS", "FFF59E0B", "Mid-load. Acompanan a un ancla si usan otra etapa. Lotes 200-400."),
    "C": ("FILLERS", "FF10B981", "Carga interna minima. Distribuir libre 50-200/dia. Buffer ante imprevistos."),
}

def fetch_image(url):
    if not url:
        return None
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        img = PILImage.open(io.BytesIO(r.content))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        # Resize to fit cell
        img.thumbnail((140, 140))
        out = io.BytesIO()
        img.save(out, format="PNG")
        out.seek(0)
        return out
    except Exception as e:
        print(f"  ! falla {url}: {e}")
        return None

wb = Workbook()
ws = wb.active
ws.title = "Perfiles"

# Title
ws.merge_cells("A1:L1")
ws["A1"] = "Perfilamiento de modelos - Sem 15-16"
ws["A1"].font = Font(size=16, bold=True, color="FFFFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="FF1E3A8A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Headers
headers = ["Perfil", "Foto", "Modelo", "Codigo", "ICO", "PREL", "Robot", "POST", "N/A", "Maquila", "Etapa cuello", "Regla operativa"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor="FF334155")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(bottom=Side(style="medium"))
ws.row_dimensions[3].height = 28

# Column widths
widths = [10, 22, 10, 18, 8, 8, 8, 8, 8, 10, 22, 38]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

thin = Side(style="thin", color="FFCBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row = 4
for perfil, modelo, codigo, ico, prel, robot, post, na, maq, cuello, regla, url in DATA:
    nombre, color, _ = PERFIL_INFO[perfil]
    fill = PatternFill("solid", fgColor=color)
    light = PatternFill("solid", fgColor="FFF8FAFC" if row % 2 == 0 else "FFFFFFFF")

    ws.cell(row=row, column=1, value=f"{perfil} - {nombre}").fill = fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFFFF", size=10)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Image
    print(f"Descargando {modelo}...")
    img_data = fetch_image(url)
    if img_data:
        try:
            xl_img = XLImage(img_data)
            xl_img.width = 130
            xl_img.height = 110
            ws.add_image(xl_img, f"B{row}")
        except Exception as e:
            ws.cell(row=row, column=2, value="(error img)")
    else:
        ws.cell(row=row, column=2, value="(sin imagen)")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=modelo).font = Font(bold=True)
    ws.cell(row=row, column=4, value=codigo)
    ws.cell(row=row, column=5, value=ico).font = Font(bold=True)
    ws.cell(row=row, column=6, value=prel)
    ws.cell(row=row, column=7, value=robot)
    ws.cell(row=row, column=8, value=post)
    ws.cell(row=row, column=9, value=na)
    ws.cell(row=row, column=10, value=maq)
    ws.cell(row=row, column=11, value=cuello)
    ws.cell(row=row, column=12, value=regla)

    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        if col >= 3:
            cell.fill = light
        if col >= 5 and col <= 10:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"
        elif col >= 11:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif col == 3 or col == 4:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row].height = 90
    row += 1

# Leyenda de perfiles abajo
row += 2
ws.cell(row=row, column=1, value="LEYENDA DE PERFILES").font = Font(bold=True, size=12)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
row += 1
for perfil, (nombre, color, desc) in PERFIL_INFO.items():
    c = ws.cell(row=row, column=1, value=f"{perfil} - {nombre}")
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    d = ws.cell(row=row, column=2, value=desc)
    d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

# Regla de oro
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
c = ws.cell(row=row, column=1, value="REGLA DE ORO: 1 Ancla + 1-2 Balancers + 1 Filler ~= 90% utilizacion sin cuellos")
c.font = Font(bold=True, size=11, color="FFFFFFFF")
c.fill = PatternFill("solid", fgColor="FF1E3A8A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 26

ws.freeze_panes = "A4"

wb.save(OUT)
print(f"\nOK: {OUT}")
