"""
Generador de propuesta optimizada de distribución semanal de un backlog.

USO:
    python scripts/generar_propuesta_backlog.py --input <archivo.xlsx>

OPCIONES:
    --input PATH            Excel del backlog (formato Andrea)
    --output PATH           Salida (default: <input>_OPTIMIZADO.xlsx)
    --template PATH         Template Excel con estructura/imágenes/fórmulas
                            (default: Matriz_Horas_Backlog_Andrea_Propuesta.xlsx)
    --max-modelos N         Máx modelos activos por semana (default: 5)
    --semanas A-B           Rango de semanas (default: lee del Excel)
    --excluir LIST          Modelos a excluir, separados por coma
    --fijar M:S1,S2         Fijar modelo M en semanas S1,S2 (repetible)
    --no-aislar M           No aislar el modelo robot-restringido M
    --max-sem N             Máx semanas por modelo (default: 3)
    --no-overrides          Ignorar overrides.yaml
    --dry-run               No escribir Excel, solo imprimir distribución

REGLAS APLICADAS (playbook):
    R1) Lotes contiguos: <=1500 -> 1 sem, <=2200 -> 2 sem, >2200 -> 3 sem
    R2) Mezcla baja: max 5 modelos activos por semana
    R3) Capacidad robot no excedida (9 robots PESPUNTE x 50h/sem = 450 h)
    R4) Modelos con robots restringidos se aíslan primero
    R5) Modelos neutros (sin catálogo o solo MAQUILA) como relleno

PRECEDENCIA DE OVERRIDES:
    1. Flags CLI
    2. scripts/overrides.yaml (si existe y activo)
    3. Restricciones activas en Supabase (FECHA_LIMITE, ROBOT_NO_DISPONIBLE, etc.)
    4. Defaults del playbook
"""
import argparse
import json
import math
import os
import sys
import urllib.request
import urllib.error
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACIÓN SUPABASE
# ============================================================
SUPABASE_URL = "https://folmyddedsdzlbegumbo.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZvbG15ZGRlZHNkemxiZWd1bWJvIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzE2MDcwNDYsImV4cCI6MjA4NzE4MzA0Nn0.RwCvujunSKA-zc704OlrEDkWlpWkOEIaavQAMD6-ewU"
HEADERS = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {ANON_KEY}",
    "Content-Type": "application/json",
}

# ============================================================
# DEFAULTS DEL PLAYBOOK
# ============================================================
HRS_SEM_POR_ROBOT = 50         # L-V 9h*5 + Sab 5h
MAX_MOD_SEM_DEF = 5
MAX_SEM_MOD_DEF = 3
ROBOT_RESTRINGIDO_THRESHOLD = 3  # operación con <=3 robots compatibles = restringida

# ============================================================
# UTILIDADES SUPABASE (HTTP plano, sin sdk)
# ============================================================
def sb_get(table, params=""):
    url = f"{SUPABASE_URL}/rest/v1/{table}?{params}"
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f"[ERROR Supabase] {e.code} en {table}: {e.read().decode()[:200]}", file=sys.stderr)
        return []
    except Exception as e:
        print(f"[ERROR red] {e}", file=sys.stderr)
        return []

def fetch_robots_activos():
    """Robots PESPUNTE activos compatibles con la familia general."""
    rows = sb_get("robots", "select=id,nombre,estado,area")
    return [r for r in rows if r["estado"] == "ACTIVO" and r["area"] == "PESPUNTE"]

def fetch_catalogo():
    """{modelo_num: {alternativas, seg_par por etapa, robots_restringido, imagen_url}}"""
    modelos = sb_get("catalogo_modelos", "select=id,modelo_num,alternativas,total_sec_per_pair,imagen_url,alternativas_imagenes")
    out = {}
    for m in modelos:
        out[m["modelo_num"]] = {
            "id": m["id"],
            "alternativas": m.get("alternativas") or [],
            "total_sec_per_pair": m.get("total_sec_per_pair") or 0,
            "imagen_url": m.get("imagen_url") or "",
            "alternativas_imagenes": m.get("alternativas_imagenes") or {},
            "PRE": 0, "ROB": 0, "POST": 0, "NA": 0, "MAQ": 0,
            "ops": [],
            "robot_restringido": False,
        }
    # operaciones
    ops = sb_get("catalogo_operaciones", "select=id,modelo_id,etapa,fraccion,operacion,recurso,sec_per_pair")
    op_by_modelo = {}
    for op in ops:
        op_by_modelo.setdefault(op["modelo_id"], []).append(op)
    # robots por operación
    cor = sb_get("catalogo_operacion_robots", "select=operacion_id,robot_id")
    robots_per_op = {}
    for c in cor:
        robots_per_op.setdefault(c["operacion_id"], []).append(c["robot_id"])

    PROC_MAP = {
        "MAQUILA": "MAQ",
        "PLANA": "POST", "POSTE": "POST", "PLANA,POSTE": "POST",
        "MESA": "PRE",
        "ROBOT": "ROB",
    }
    for num, m in out.items():
        for op in op_by_modelo.get(m["id"], []):
            recurso = (op.get("recurso") or "").upper()
            seg = op.get("sec_per_pair") or 0
            etapa_key = None
            if "ROBOT" in recurso:
                etapa_key = "ROB"
            elif "MAQUILA" in recurso:
                etapa_key = "MAQ"
            elif "POST" in (op.get("etapa") or "").upper() or "POSTE" in recurso:
                etapa_key = "POST"
            elif "PLANA" in recurso:
                etapa_key = "POST"
            elif "MESA" in recurso:
                etapa_key = "PRE"
            else:
                etapa_key = "PRE"
            m[etapa_key] += seg
            m["ops"].append(op)
            # restricción robots
            if etapa_key == "ROB":
                rids = robots_per_op.get(op["id"], [])
                if 0 < len(rids) <= ROBOT_RESTRINGIDO_THRESHOLD:
                    m["robot_restringido"] = True
    return out

def fetch_restricciones_activas():
    rows = sb_get("restricciones", "select=tipo,modelo_num,semana,parametros&activa=eq.true")
    return rows

# ============================================================
# OVERRIDES YAML
# ============================================================
def load_overrides():
    # Está en la misma carpeta que este script (backlog_tool/)
    p = Path(__file__).parent / "overrides.yaml"
    if not p.exists():
        return {}
    try:
        import yaml
    except ImportError:
        print("[WARN] PyYAML no instalado, ignorando overrides.yaml. pip install pyyaml", file=sys.stderr)
        return {}
    data = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    if not data.get("activo", True):
        return {}
    fecha_exp = data.get("fecha_expira")
    if fecha_exp:
        if isinstance(fecha_exp, str):
            fecha_exp = datetime.fromisoformat(fecha_exp).date()
        if date.today() > fecha_exp:
            print(f"[WARN] overrides.yaml expirado el {fecha_exp}, ignorando.", file=sys.stderr)
            return {}
    return data.get("overrides", {})

# ============================================================
# LECTURA DEL BACKLOG INPUT
# ============================================================
def parse_backlog(path):
    """Devuelve [(modelo_str, total_pares)] y lista de semanas.

    Soporta 2 formatos:
    A) Simple: celdas 'Semana inicio'/'Semana fin' + columna 'Modelo' + 'Total Pares'
    B) Legacy: matriz modelo×semana (suma cada fila para obtener total)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    # === Formato A: simple ===
    for sn in wb.sheetnames:
        ws = wb[sn]
        sem_ini = sem_fin = None
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str): continue
                low = v.lower().strip()
                if "semana" in low and "inicio" in low:
                    nxt = ws.cell(row=r, column=c+1).value
                    if isinstance(nxt, (int, float)): sem_ini = int(nxt)
                elif "semana" in low and "fin" in low:
                    nxt = ws.cell(row=r, column=c+1).value
                    if isinstance(nxt, (int, float)): sem_fin = int(nxt)
        if sem_ini and sem_fin and sem_fin >= sem_ini:
            # buscar cabecera "Modelo" en alguna fila
            for r in range(1, min(ws.max_row + 1, 20)):
                for c in range(1, min(ws.max_column + 1, 10)):
                    v = ws.cell(row=r, column=c).value
                    if not (isinstance(v, str) and v.strip().lower() == "modelo"):
                        continue
                    # Detectar columnas: revisar las siguientes 1-3 columnas
                    # buscando "alternativas" / "alternativa" y "total pares"
                    col_modelo = c
                    col_alt = None
                    col_total = None
                    for off in range(1, 5):
                        h = ws.cell(row=r, column=c + off).value
                        if not isinstance(h, str): continue
                        hl = h.strip().lower()
                        if "alternativ" in hl and col_alt is None:
                            col_alt = c + off
                        elif "total" in hl and "par" in hl and col_total is None:
                            col_total = c + off
                    if col_total is None:
                        continue  # no es la cabecera buscada
                    # Leer filas
                    modelos = []
                    rr = r + 1
                    while rr <= ws.max_row + 30:
                        nm_raw = ws.cell(row=rr, column=col_modelo).value
                        tot_raw = ws.cell(row=rr, column=col_total).value
                        alt_raw = ws.cell(row=rr, column=col_alt).value if col_alt else None

                        if nm_raw is None and tot_raw is None:
                            rr += 1
                            if rr > r + 60: break
                            continue
                        if isinstance(nm_raw, str) and nm_raw.strip().upper() == "TOTAL":
                            break
                        # Modelo: aceptar número o string. Lo normalizamos a string.
                        if nm_raw is None:
                            rr += 1
                            continue
                        if isinstance(nm_raw, (int, float)):
                            modelo_str = str(int(nm_raw))
                        else:
                            modelo_str = str(nm_raw).strip()
                        if not modelo_str or modelo_str.startswith("("):  # placeholders tipo "(modelo)"
                            rr += 1
                            continue
                        if not isinstance(tot_raw, (int, float)) or tot_raw <= 0:
                            rr += 1
                            continue
                        # Combinar modelo + alternativa
                        alt_str = ""
                        if alt_raw and isinstance(alt_raw, str) and not alt_raw.strip().startswith("("):
                            alt_str = alt_raw.strip()
                        nombre_completo = f"{modelo_str} {alt_str}".strip()
                        modelos.append((nombre_completo, int(tot_raw)))
                        rr += 1
                        if rr > r + 60: break
                    if modelos:
                        semanas = list(range(sem_ini, sem_fin + 1))
                        return modelos, semanas
    # === Formato B: legacy (matriz modelo×semana) ===
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(ws.max_row + 1, 10)):
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str) and "semana" in v.lower():
                    for rr in range(r, r + 3):
                        nums = []
                        for cc in range(c, ws.max_column + 1):
                            val = ws.cell(row=rr, column=cc).value
                            if isinstance(val, (int, float)) and 1 <= val <= 53:
                                nums.append((cc, int(val)))
                        if len(nums) >= 3:
                            return _read_modelos(ws, rr + 1, nums)
    raise RuntimeError(f"No se encontró tabla de backlog en {path}")

def _read_modelos(ws, start_row, sem_cols):
    """Desde start_row hacia abajo, lee filas con modelo+pares."""
    modelos = []
    semanas = [s for _, s in sem_cols]
    cols = [c for c, _ in sem_cols]
    # columna de modelo: la inmediatamente anterior a la primera col de semana
    col_modelo = cols[0] - 1
    while col_modelo > 0:
        v = ws.cell(row=start_row, column=col_modelo).value
        if v and isinstance(v, str):
            break
        col_modelo -= 1
    if col_modelo < 1:
        col_modelo = cols[0] - 1
    r = start_row
    while r <= ws.max_row + 5:
        nm = ws.cell(row=r, column=col_modelo).value
        if not nm or not isinstance(nm, str) or nm.strip() in ("", "TOTAL"):
            r += 1
            if r > start_row + 30:
                break
            continue
        nm = nm.strip()
        # sumar pares de las columnas semana
        total = 0
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                total += int(v)
        if total > 0:
            modelos.append((nm, total))
        r += 1
        if len(modelos) >= 50:
            break
    return modelos, semanas

# ============================================================
# MATCHING modelo_str -> modelo_num
# ============================================================
def match_modelo(nm, catalogo_db):
    """Extrae el modelo_num del string '68127 NE/RO SLI'."""
    import re
    m = re.match(r"(\d{4,6})", nm.strip())
    if not m: return None
    num = m.group(1)
    return num if num in catalogo_db else None

# ============================================================
# ALGORITMO GREEDY (playbook)
# ============================================================
def n_sem_por_volumen(t, max_sem):
    # EXPERIMENTO: regla heurística desactivada. Todo modelo ocupa 1 semana.
    # La carga de robot por semana se controla en el scoring del greedy.
    return 1
    # if t <= 1500: return 1
    # if t <= 2200: return min(2, max_sem)
    # return min(3, max_sem)

def horas_robot_total(seg_par_rob, total):
    return total * seg_par_rob / 3600

def distribuir(modelos, semanas, catalogo_db, cap_robot_sem, opts):
    """
    modelos: [(nombre_str, total_pares)]
    Retorna: dict {nombre: {sem: pares}}, dict {sem: carga_robot}
    """
    asig = {n: {s: 0 for s in semanas} for n, _ in modelos}
    carga = {s: 0.0 for s in semanas}
    mods = {s: 0 for s in semanas}

    # info por modelo
    info = {}
    for n, t in modelos:
        num = match_modelo(n, catalogo_db)
        if num and num in catalogo_db:
            cat = catalogo_db[num]
            info[n] = {
                "ROB": cat["ROB"],
                "PRE": cat["PRE"],
                "POST": cat["POST"],
                "NA": cat["NA"],
                "MAQ": cat["MAQ"],
                "robot_restringido": cat["robot_restringido"],
                "neutro": cat["ROB"] == 0,
                "sin_catalogo": False,
            }
        else:
            info[n] = {"ROB":0,"PRE":0,"POST":0,"NA":0,"MAQ":0,
                       "robot_restringido": False, "neutro": True, "sin_catalogo": True}

    # aplicar fijaciones
    fijar = opts.get("fijar", {})
    for n, _ in modelos:
        if n in fijar:
            asig[n] = {s: 0 for s in semanas}

    # orden: restringidos -> robot-share por h-robot desc -> neutros
    def prio(item):
        n, t = item
        i = info[n]
        if n in fijar: return (-1, 0)
        if i["robot_restringido"] and n not in opts.get("no_aislar", set()):
            return (0, 0)
        if i["neutro"]:
            return (2, -t)
        return (1, -horas_robot_total(i["ROB"], t))

    orden = sorted(modelos, key=prio)
    max_sem = opts.get("max_sem", MAX_SEM_MOD_DEF)
    max_mod = opts.get("max_mod", MAX_MOD_SEM_DEF)
    print(f"[backlog.distribuir] max_mod={max_mod} max_sem={max_sem} "
          f"n_modelos={len(modelos)} n_semanas={len(semanas)}", flush=True)

    # Calcular N por modelo (usado por el loop principal). El cap max_mod se
    # trata como SOFT: el algoritmo lo respeta donde puede, y solo lo excede
    # en las semanas donde es matemáticamente imposible cumplirlo.
    n_por_modelo = {}
    for n, total in orden:
        if n in fijar:
            n_por_modelo[n] = len([s for s in fijar[n] if s in semanas]) or 1
        else:
            n_por_modelo[n] = n_sem_por_volumen(total, max_sem)
    demanda = sum(n_por_modelo.values())
    n_sem = len(semanas)
    max_mod_pedido = max_mod
    min_factible = max(1, math.ceil(demanda / n_sem)) if n_sem > 0 else max_mod
    # AUTO-BUMP iterativo: si max_mod pedido no permite distribución completa,
    # subir +1 hasta encontrar el menor valor donde TODOS los modelos caben
    # respetando el cap. Equivale a math.ceil(demanda/n_sem).
    intentos = []
    if max_mod < min_factible:
        for v in range(max_mod_pedido, min_factible + 1):
            intentos.append(v)
        max_mod = min_factible
        print(f"[backlog.distribuir] AUTO-BUMP: intentos={intentos} -> "
              f"usado={max_mod} (demanda={demanda}, n_sem={n_sem})", flush=True)
    pendientes = []  # modelos que no caben (no debería pasar con auto-bump)

    for n, total in orden:
        # si está fijado, distribuir parejo en sus semanas
        if n in fijar:
            sems_fijas = [s for s in fijar[n] if s in semanas]
            if not sems_fijas:
                continue
            base = total // len(sems_fijas)
            resto = total - base * len(sems_fijas)
            for i, s in enumerate(sems_fijas):
                v = base + (1 if i < resto else 0)
                v = round(v / 100) * 100
                asig[n][s] = v
                if not info[n]["neutro"]:
                    carga[s] += v * info[n]["ROB"] / 3600
                mods[s] += 1
            # ajuste exacto
            diff = total - sum(asig[n].values())
            if diff:
                asig[n][sems_fijas[0]] += diff
            continue

        N = n_por_modelo.get(n, n_sem_por_volumen(total, max_sem))
        base = total // N
        bloques = [base] * N
        for i in range(total - base * N): bloques[i] += 1
        rb = [round(b / 100) * 100 for b in bloques]
        diff = total - sum(rb)
        if diff: rb[0] += diff
        bloques = rb

        is_neutro = info[n]["neutro"]
        h_par = info[n]["ROB"] / 3600 if not is_neutro else 0
        es_restringido = info[n]["robot_restringido"] and n not in opts.get("no_aislar", set())

        # FASE 1: max_mod es DURO. Intentar meter el modelo SOLO en ventanas
        # que respeten el cap. Si no hay, dejarlo pendiente para FASE 2.
        mejor = None
        mejor_score = float("inf")
        for st in range(len(semanas) - N + 1):
            vent = semanas[st:st + N]
            if any(mods[s] >= max_mod for s in vent):
                continue
            if is_neutro:
                score = sum(mods[s] for s in vent) * 1000 - sum(carga[s] for s in vent)
            elif es_restringido:
                score = sum(carga[s] for s in vent)
            else:
                pico = max(carga[vent[k]] + bloques[k] * h_par for k in range(N))
                score = pico * (10 if pico > cap_robot_sem else 1)
            if score < mejor_score:
                mejor_score = score; mejor = vent

        if mejor is None:
            # No cupo en FASE 1: dejarlo pendiente
            pendientes.append((n, total, N, bloques, h_par, is_neutro, es_restringido))
            continue

        for k, s in enumerate(mejor):
            asig[n][s] += bloques[k]
            if not is_neutro:
                carga[s] += bloques[k] * h_par
            mods[s] += 1

    # FASE 2: colocar modelos pendientes (los que no cupieron con cap duro).
    # Estrategia: concentrar excesos en POCAS semanas en vez de repartirlos.
    # Para cada pendiente, elegir las N semanas que YA tienen más exceso
    # (o si ninguna, las menos cargadas en robot).
    if pendientes:
        print(f"[backlog.distribuir] FASE 2: {len(pendientes)} modelos pendientes "
              f"a colocar fuera del cap", flush=True)
    for n, total, N, bloques, h_par, is_neutro, es_restringido in pendientes:
        # Preferir semanas que ya excedieron el cap (concentración),
        # luego menos cargadas en robot
        ranked = sorted(
            semanas,
            key=lambda s: (
                -(max(0, mods[s] - max_mod_pedido + 1)),  # más exceso primero (negativo)
                carga[s] if not is_neutro else 0,
                mods[s],
            )
        )
        # Tomar N semanas, preferir contiguas si es posible
        elegidas = sorted(ranked[:N])
        for k, s in enumerate(elegidas):
            asig[n][s] += bloques[k]
            if not is_neutro:
                carga[s] += bloques[k] * h_par
            mods[s] += 1

    print(f"[backlog.distribuir] FINAL mods por semana: "
          f"{ {s: mods[s] for s in semanas} } (cap={max_mod})", flush=True)
    max_real = max(mods.values()) if mods else 0
    info["__meta__"] = {
        "max_mod_pedido": max_mod_pedido,
        "max_mod_usado": max_mod,
        "max_mod_real": max_real,
        "min_factible": min_factible,
        "ajustado": max_mod != max_mod_pedido,
        "intentos": intentos,
        "demanda_model_weeks": demanda,
    }
    return asig, carga, info

# ============================================================
# ESCRITURA EXCEL
# ============================================================
def safe_set(ws, r, c, v):
    try:
        ws.cell(row=r, column=c).value = v
    except AttributeError:
        coord = f"{get_column_letter(c)}{r}"
        for rng in list(ws.merged_cells.ranges):
            if coord in rng:
                ws.unmerge_cells(str(rng)); break
        ws.cell(row=r, column=c).value = v

def _modelo_num(nombre: str) -> str:
    """Extrae los primeros 5 dígitos del nombre del modelo."""
    import re
    if not nombre: return ""
    m = re.match(r"(\d{5})", str(nombre).strip())
    return m.group(1) if m else ""


def _descargar_imagen(url: str):
    """Descarga una imagen desde URL y devuelve un BytesIO listo para openpyxl.Image.
    Devuelve None si falla."""
    if not url:
        return None
    try:
        import io
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            data = r.read()
        return io.BytesIO(data)
    except Exception:
        return None


def _descargar_imagenes_paralelo(urls: list) -> dict:
    """Descarga múltiples URLs en paralelo. Retorna {url: BytesIO|None}."""
    from concurrent.futures import ThreadPoolExecutor
    out = {}
    if not urls:
        return out
    with ThreadPoolExecutor(max_workers=8) as ex:
        results = list(ex.map(_descargar_imagen, urls))
    for url, res in zip(urls, results):
        out[url] = res
    return out


def _insertar_imagen_en_celda(ws, row, col, img_bytes, max_h_px=60, max_w_px=60):
    """Inserta una imagen embebida anclada a la celda (row, col).
    Redimensiona FÍSICAMENTE con Pillow antes de embeber."""
    if img_bytes is None:
        return False
    try:
        import io as _io
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter as _gcl
        from PIL import Image as PILImage

        img_bytes.seek(0)
        pil = PILImage.open(img_bytes).convert("RGBA")
        w, h = pil.size
        ratio = min(max_w_px / w, max_h_px / h, 1.0)
        new_w, new_h = max(1, int(w * ratio)), max(1, int(h * ratio))
        pil_resized = pil.resize((new_w, new_h), PILImage.LANCZOS)

        out_bytes = _io.BytesIO()
        pil_resized.save(out_bytes, format="PNG")
        out_bytes.seek(0)

        img = XLImage(out_bytes)
        img.width = new_w
        img.height = new_h
        img.anchor = f"{_gcl(col)}{row}"
        ws.add_image(img)
        return True
    except Exception:
        return False


def _quitar_imagenes_de_filas(ws, filas_a_quitar: set):
    """Elimina de ws._images todas las imágenes ancladas a filas en el set.
    Las filas se pasan en 1-indexed (como Excel); el anchor las maneja 0-indexed."""
    if not hasattr(ws, "_images") or not ws._images:
        return
    filas_0idx = {r - 1 for r in filas_a_quitar}
    nuevas = []
    for img in ws._images:
        anchor = getattr(img, "anchor", None)
        fr = getattr(anchor, "_from", None) if anchor is not None else None
        if fr is None:
            nuevas.append(img)
            continue
        if fr.row in filas_0idx:
            continue  # eliminar
        nuevas.append(img)
    ws._images = nuevas


def _escribir_total_general_row(ws, r, nombre, asig_modelo, info_modelo):
    """Escribe una fila de la hoja Total General.
    Cols: B=Modelo, C=Total Pares,
          D=SegPar PRE, E=H, F=Pers, G=Días
          H=SegPar ROB, I=H, J=Pers, K=Días
          L=SegPar POST, M=H, N=Pers, O=Días
          P=SegPar NA, Q=H
          R=SegPar MAQ, S=H
          T=TOTAL HRS
    """
    safe_set(ws, r, 2, nombre)
    total_pares = sum(asig_modelo.values())
    safe_set(ws, r, 3, total_pares)
    if info_modelo.get("sin_catalogo"):
        safe_set(ws, r, 4, "SIN CATALOGO")
        safe_set(ws, r, 20, "—")
        return
    if info_modelo.get("PRE", 0) > 0:
        safe_set(ws, r, 4, info_modelo["PRE"])
        safe_set(ws, r, 5, f"=C{r}*D{r}/3600")
        safe_set(ws, r, 6, f"=E{r}/$F$2")
        safe_set(ws, r, 7, f"=F{r}/$C$2")
    else:
        safe_set(ws, r, 4, "—"); safe_set(ws, r, 5, "—")
        safe_set(ws, r, 6, "—"); safe_set(ws, r, 7, "—")
    if info_modelo.get("ROB", 0) > 0:
        safe_set(ws, r, 8, info_modelo["ROB"])
        safe_set(ws, r, 9, f"=C{r}*H{r}/3600")
        safe_set(ws, r, 10, f"=I{r}/$J$2")
        safe_set(ws, r, 11, f"=J{r}/$C$2")
    else:
        safe_set(ws, r, 8, "—"); safe_set(ws, r, 9, "—")
        safe_set(ws, r, 10, "—"); safe_set(ws, r, 11, "—")
    if info_modelo.get("POST", 0) > 0:
        safe_set(ws, r, 12, info_modelo["POST"])
        safe_set(ws, r, 13, f"=C{r}*L{r}/3600")
        safe_set(ws, r, 14, f"=M{r}/$N$2")
        safe_set(ws, r, 15, f"=N{r}/$C$2")
    else:
        safe_set(ws, r, 12, "—"); safe_set(ws, r, 13, "—")
        safe_set(ws, r, 14, "—"); safe_set(ws, r, 15, "—")
    if info_modelo.get("NA", 0) > 0:
        safe_set(ws, r, 16, info_modelo["NA"])
        safe_set(ws, r, 17, f"=C{r}*P{r}/3600")
    else:
        safe_set(ws, r, 16, "—"); safe_set(ws, r, 17, "—")
    if info_modelo.get("MAQ", 0) > 0:
        safe_set(ws, r, 18, info_modelo["MAQ"])
        safe_set(ws, r, 19, f"=C{r}*R{r}/3600")
    else:
        safe_set(ws, r, 18, "—"); safe_set(ws, r, 19, "—")
    parts = []
    for col, key in [("E","PRE"),("I","ROB"),("M","POST"),("Q","NA"),("S","MAQ")]:
        if info_modelo.get(key, 0) > 0:
            parts.append(f"{col}{r}")
    safe_set(ws, r, 20, "=" + "+".join(parts) if parts else 0)


def _escribir_seg_por_par_row(ws, r, nombre, info_modelo):
    """Escribe una fila de la hoja Seg por Par.
    Cols: B=Modelo, C=Total Seg/Par, D=PRE, E=ROB, F=POST, G=NA, H=MAQ, I=Pares/Hr
    """
    safe_set(ws, r, 2, nombre)
    if info_modelo.get("sin_catalogo"):
        safe_set(ws, r, 3, "SIN CATALOGO")
        return
    pre = info_modelo.get("PRE", 0) or 0
    rob = info_modelo.get("ROB", 0) or 0
    post = info_modelo.get("POST", 0) or 0
    na = info_modelo.get("NA", 0) or 0
    maq = info_modelo.get("MAQ", 0) or 0
    safe_set(ws, r, 3, f"=D{r}+E{r}+F{r}+G{r}+H{r}")
    safe_set(ws, r, 4, pre)
    safe_set(ws, r, 5, rob)
    safe_set(ws, r, 6, post)
    safe_set(ws, r, 7, na)
    safe_set(ws, r, 8, maq)
    safe_set(ws, r, 9, f"=IF(C{r}>0,3600/C{r},0)")


def escribir_excel(template_path, output_path, asig, info, semanas, catalogo=None):
    wb = openpyxl.load_workbook(template_path)

    # Modelos del input indexados por número (4-6 dígitos)
    input_modelos = list(asig.keys())  # nombres completos del input
    input_por_num = {_modelo_num(n): n for n in input_modelos if _modelo_num(n)}
    asignados_input = set()  # nombres ya colocados

    # Eliminar TODAS las imágenes del workbook (a petición del usuario)
    for sn_img in wb.sheetnames:
        ws_img = wb[sn_img]
        if hasattr(ws_img, "_images"):
            ws_img._images = []

    # ===== Hoja "Backlog" =====
    # Cols: B=Modelo, C..=Sem (dinámicas según semanas), última=Total
    if "Backlog" in wb.sheetnames:
        from openpyxl.styles import Font as FtB, PatternFill as PFB, Alignment as AlB, Border as BdB, Side as SdB
        NO_FILL = PFB(fill_type=None)
        NO_BORDER = BdB()
        ws = wb["Backlog"]
        safe_set(ws, 1, 1, "PROPUESTA OPTIMIZADA - data-driven")
        safe_set(ws, 2, 1, "Generado por backlog_tool/generar_propuesta.py")

        # Estilos
        B_HDR_FILL = PFB(fill_type="solid", fgColor="FF1E3A5F")
        B_HDR_FONT = FtB(bold=True, color="FFFFFFFF", size=11)
        B_TOT_FILL = PFB(fill_type="solid", fgColor="FF2E5BA8")
        B_TOT_FONT = FtB(bold=True, color="FFFFFFFF", size=11)
        B_PINK = PFB(fill_type="solid", fgColor="FFFECACA")
        B_SINCAT_FONT = FtB(italic=True, color="FFB91C1C", size=10)
        B_BORDER = BdB(
            left=SdB(border_style="thin", color="FFCBD5E1"),
            right=SdB(border_style="thin", color="FFCBD5E1"),
            top=SdB(border_style="thin", color="FFCBD5E1"),
            bottom=SdB(border_style="thin", color="FFCBD5E1"),
        )
        B_CENTER = AlB(horizontal="center", vertical="center")
        B_LEFT = AlB(horizontal="left", vertical="center", indent=1)
        B_RIGHT = AlB(horizontal="right", vertical="center", indent=1)

        # Capturar nombres del template ANTES de limpiar
        template_rows_b = {}
        for r in range(5, 35):
            nm = ws.cell(row=r, column=2).value
            if not nm or str(nm).strip().upper() == "TOTAL":
                continue
            num = _modelo_num(str(nm))
            if num:
                template_rows_b[r] = num

        # LIMPIAR valores Y fills Y borders
        for r in range(4, 41):
            for c in range(2, 16):
                safe_set(ws, r, c, None)
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.fill = NO_FILL
                    cell.border = NO_BORDER
                except AttributeError:
                    pass
            ws.row_dimensions[r].hidden = False

        # Headers (row 4)
        for j, s in enumerate(semanas):
            cell = ws.cell(row=4, column=3 + j, value=f"Sem {s}")
            cell.fill = B_HDR_FILL
            cell.font = B_HDR_FONT
            cell.alignment = B_CENTER
            cell.border = B_BORDER
        col_total_b = 3 + len(semanas)
        cell_modelo = ws.cell(row=4, column=2, value="Modelo")
        cell_modelo.fill = B_HDR_FILL
        cell_modelo.font = B_HDR_FONT
        cell_modelo.alignment = B_CENTER
        cell_modelo.border = B_BORDER
        cell_total_h = ws.cell(row=4, column=col_total_b, value="Total")
        cell_total_h.fill = B_HDR_FILL
        cell_total_h.font = B_HDR_FONT
        cell_total_h.alignment = B_CENTER
        cell_total_h.border = B_BORDER
        ws.row_dimensions[4].height = 26

        def _aplicar_estilo_fila_b(ws_, r_, sin_cat=False):
            """Aplica estilo a una fila de datos en Backlog."""
            ws_.row_dimensions[r_].height = 24
            for c in range(2, col_total_b + 1):
                try:
                    cell = ws_.cell(row=r_, column=c)
                    cell.border = B_BORDER
                    if c == 2:
                        cell.alignment = B_LEFT
                        cell.font = B_SINCAT_FONT if sin_cat else FtB(bold=True, size=10, color="FF1E3A5F")
                    elif c == col_total_b:
                        cell.alignment = B_RIGHT
                        cell.font = FtB(bold=True, size=10, color="FF1E3A5F")
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = B_CENTER
                        cell.font = B_SINCAT_FONT if sin_cat else FtB(size=10)
                        cell.number_format = '#,##0'
                    if sin_cat:
                        cell.fill = B_PINK
                except AttributeError:
                    pass

        # Escribir filas matcheadas por num
        filas_huerfanas_b = set()
        for r, num in template_rows_b.items():
            input_match = input_por_num.get(num)
            if input_match:
                safe_set(ws, r, 2, input_match)
                for j, s in enumerate(semanas):
                    v = asig[input_match][s]
                    safe_set(ws, r, 3 + j, v if v > 0 else None)
                safe_set(ws, r, col_total_b, f"=SUM(C{r}:{get_column_letter(col_total_b - 1)}{r})")
                _aplicar_estilo_fila_b(ws, r, sin_cat=info[input_match].get("sin_catalogo", False))
                asignados_input.add(input_match)
            else:
                ws.row_dimensions[r].hidden = True
                filas_huerfanas_b.add(r)
        _quitar_imagenes_de_filas(ws, filas_huerfanas_b)

        # Agregar modelos nuevos del input al final
        last_row_b = max(template_rows_b.keys()) if template_rows_b else 4
        for nm in input_modelos:
            if nm in asignados_input: continue
            last_row_b += 1
            safe_set(ws, last_row_b, 2, nm)
            for j, s in enumerate(semanas):
                v = asig[nm][s]
                safe_set(ws, last_row_b, 3 + j, v if v > 0 else None)
            safe_set(ws, last_row_b, col_total_b, f"=SUM(C{last_row_b}:{get_column_letter(col_total_b - 1)}{last_row_b})")
            _aplicar_estilo_fila_b(ws, last_row_b, sin_cat=info[nm].get("sin_catalogo", False))

        # Fila TOTAL
        rt = last_row_b + 1
        safe_set(ws, rt, 2, "TOTAL")
        for j in range(len(semanas)):
            col = 3 + j
            letter = get_column_letter(col)
            safe_set(ws, rt, col, f"=SUM({letter}5:{letter}{rt - 1})")
        letter_t = get_column_letter(col_total_b)
        safe_set(ws, rt, col_total_b, f"=SUM({letter_t}5:{letter_t}{rt - 1})")
        # Estilo TOTAL
        ws.row_dimensions[rt].height = 26
        for c in range(2, col_total_b + 1):
            try:
                cell = ws.cell(row=rt, column=c)
                cell.fill = B_TOT_FILL
                cell.font = B_TOT_FONT
                cell.alignment = B_CENTER if c == 2 else (B_RIGHT if c == col_total_b else B_CENTER)
                cell.border = B_BORDER
                if c >= 3:
                    cell.number_format = '#,##0'
            except AttributeError:
                pass
        # Ocultar filas vacías post-TOTAL
        for r_hide in range(rt + 1, max(ws.max_row, rt + 30) + 1):
            ws.row_dimensions[r_hide].hidden = True

    # ===== Hoja "Backlog Original" =====
    # El usuario solo da TOTALES en el input, así que esta hoja muestra solo
    # Modelo + Alternativa + Total Pares (sin desglose por semana).
    if "Backlog Original" in wb.sheetnames:
        ws = wb["Backlog Original"]

        # Capturar nombres del template para limpiar imágenes huérfanas
        template_rows_o = {}
        for r in range(4, 35):
            nm = ws.cell(row=r, column=3).value
            if not nm or str(nm).strip().upper() == "TOTAL":
                continue
            num = _modelo_num(str(nm))
            if num:
                template_rows_o[r] = num

        from openpyxl.styles import (
            PatternFill as PFI, Border as BdI, Font as FtI,
            Alignment as AlI, Side as SdI,
        )
        NO_FILL_I = PFI(fill_type=None)
        NO_BORDER_I = BdI()

        # Limpiar TODO en cols 3..22 filas 1..40
        for r in range(1, 41):
            for c in range(3, 22):
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.fill = NO_FILL_I
                    cell.border = NO_BORDER_I
                except AttributeError:
                    pass
        for rng in list(ws.merged_cells.ranges):
            if rng.min_col >= 3:
                ws.unmerge_cells(str(rng))
        for r in range(1, 41):
            ws.row_dimensions[r].hidden = False

        # Quitar imágenes de filas template
        _quitar_imagenes_de_filas(ws, set(template_rows_o.keys()))

        # === Estilos ===
        HDR_FILL = PFI(fill_type="solid", fgColor="FF1E3A5F")
        HDR_FONT = FtI(bold=True, color="FFFFFFFF", size=11)
        TOT_FILL = PFI(fill_type="solid", fgColor="FF2E5BA8")
        TOT_FONT = FtI(bold=True, color="FFFFFFFF", size=11)
        ALT_FILL = PFI(fill_type="solid", fgColor="FFF8FAFC")
        BORDER_O = BdI(
            left=SdI(border_style="thin", color="FFCBD5E1"),
            right=SdI(border_style="thin", color="FFCBD5E1"),
            top=SdI(border_style="thin", color="FFCBD5E1"),
            bottom=SdI(border_style="thin", color="FFCBD5E1"),
        )
        CENTER_O = AlI(horizontal="center", vertical="center")
        LEFT_O = AlI(horizontal="left", vertical="center", indent=1)
        RIGHT_O = AlI(horizontal="right", vertical="center", indent=1)

        # === Título ===
        ws.merge_cells("B1:E1")
        cell_title = ws.cell(row=1, column=2, value="BACKLOG — TOTALES POR MODELO")
        cell_title.font = FtI(bold=True, size=14, color="FF1E3A5F")
        cell_title.alignment = CENTER_O
        ws.row_dimensions[1].height = 28

        # === Headers (row 3) ===
        for col, txt in [(2, ""), (3, "Modelo"), (4, "Alternativa"), (5, "Total Pares")]:
            cell = ws.cell(row=3, column=col, value=txt)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER_O
            cell.border = BORDER_O
        ws.row_dimensions[3].height = 26

        # === Pre-cargar imágenes en paralelo ===
        urls_a_descargar = []
        modelo_url = {}  # nm -> url
        if catalogo is not None:
            import re as _re_pre
            for nm in input_modelos:
                m_ = _re_pre.match(r"(\d{5})", nm.strip())
                if not m_: continue
                num_ = m_.group(1)
                cat_m = catalogo.get(num_)
                if not cat_m: continue
                url = cat_m.get("imagen_url") or ""
                if not url:
                    alts = cat_m.get("alternativas_imagenes") or {}
                    if isinstance(alts, dict):
                        for v in alts.values():
                            if v:
                                url = v
                                break
                if url:
                    modelo_url[nm] = url
                    urls_a_descargar.append(url)
        # Descargar todas las URLs únicas en paralelo
        urls_unicas = list(set(urls_a_descargar))
        imgs_cache = _descargar_imagenes_paralelo(urls_unicas) if urls_unicas else {}

        # === Datos ordenados por modelo num ===
        import re as _re_o
        modelos_ordenados = sorted(
            input_modelos,
            key=lambda x: _modelo_num(x) or x,
        )
        r_out = 4
        for nm in modelos_ordenados:
            mtch = _re_o.match(r"^(\d{5})\s*(.*)$", nm.strip())
            mod_num = mtch.group(1) if mtch else nm
            alt = mtch.group(2) if mtch else ""
            tot = sum(asig[nm].values())
            row_fill = ALT_FILL if (r_out - 4) % 2 == 1 else None
            # Celda imagen (col B)
            cell_img = ws.cell(row=r_out, column=2)
            cell_img.border = BORDER_O
            if row_fill is not None:
                cell_img.fill = row_fill
            for col, val, align in [
                (3, mod_num, CENTER_O),
                (4, alt, LEFT_O),
                (5, tot, RIGHT_O),
            ]:
                cell = ws.cell(row=r_out, column=col, value=val)
                cell.border = BORDER_O
                cell.alignment = align
                if row_fill is not None:
                    cell.fill = row_fill
            ws.cell(row=r_out, column=3).font = FtI(bold=True, size=11, color="FF1E3A5F")
            ws.cell(row=r_out, column=5).number_format = '#,##0'
            ws.row_dimensions[r_out].height = 50
            # Insertar imagen si existe
            url_nm = modelo_url.get(nm)
            if url_nm:
                img_bytes = imgs_cache.get(url_nm)
                if img_bytes is not None:
                    # Crear copia del BytesIO porque cada Image lo "consume"
                    import io as _io
                    img_bytes.seek(0)
                    copia = _io.BytesIO(img_bytes.read())
                    img_bytes.seek(0)
                    _insertar_imagen_en_celda(ws, r_out, 2, copia, max_h_px=58, max_w_px=58)
            r_out += 1

        # === Fila TOTAL ===
        rt_o = r_out
        for col, val, align in [
            (2, "", CENTER_O),
            (3, "TOTAL", CENTER_O),
            (4, "", CENTER_O),
            (5, f"=SUM(E4:E{rt_o - 1})", RIGHT_O),
        ]:
            cell = ws.cell(row=rt_o, column=col, value=val)
            cell.fill = TOT_FILL
            cell.font = TOT_FONT
            cell.alignment = align
            cell.border = BORDER_O
        ws.cell(row=rt_o, column=5).number_format = '#,##0'
        ws.row_dimensions[rt_o].height = 26

        # === Anchos de columna ===
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 16

        # Ocultar filas posteriores
        for r_hide in range(rt_o + 1, max(ws.max_row, rt_o + 30) + 1):
            ws.row_dimensions[r_hide].hidden = True

    # Hojas semanales
    # IMPORTANTE: limpiar TODAS las hojas Sem* del template (no solo las del rango)
    # para evitar datos huérfanos que generen referencias circulares.
    # Además: eliminar imágenes (redundantes con Backlog) y limpiar fills de filas vacías.
    # OJO: row 5 es decorativa (banda de colores por etapa) — NO tocarla.
    from openpyxl.styles import PatternFill, Border
    NO_FILL = PatternFill(fill_type=None)
    NO_BORDER = Border()
    for sn_clean in [s for s in wb.sheetnames if s.startswith("Sem ")]:
        ws_c = wb[sn_clean]
        # eliminar todas las imágenes (las que existen son del template original)
        if hasattr(ws_c, "_images"):
            ws_c._images = []
        # limpiar valores, fills y borders DESDE row 6 (row 5 es decorativa)
        for r in range(6, ws_c.max_row + 5):
            for c in range(2, 21):
                safe_set(ws_c, r, c, None)
                try:
                    cell = ws_c.cell(row=r, column=c)
                    cell.fill = NO_FILL
                    cell.border = NO_BORDER
                except AttributeError:
                    pass
    # Estilos del template
    from openpyxl.styles import PatternFill as PF, Font as Ft, Alignment as Al, Border as Bd, Side as Sd
    PINK_FILL = PF(fill_type="solid", fgColor="FFFECACA")
    TOTAL_FILL = PF(fill_type="solid", fgColor="FF1E3A5F")
    TOTAL_FONT = Ft(bold=True, color="FFFFFFFF")
    SIN_CAT_FONT = Ft(italic=True, color="FFB91C1C")
    THIN_BORDER = Bd(
        left=Sd(border_style="thin", color="FFE5E7EB"),
        right=Sd(border_style="thin", color="FFE5E7EB"),
        top=Sd(border_style="thin", color="FFE5E7EB"),
        bottom=Sd(border_style="thin", color="FFE5E7EB"),
    )
    CENTER_AL = Al(horizontal="center", vertical="center")
    LEFT_AL = Al(horizontal="left", vertical="center")

    # Tints suaves por columna (etapa) - colores del template original
    PRE_TINT = PF(fill_type="solid", fgColor="FFFEF3C7")    # cream para PRELIM
    ROB_TINT = PF(fill_type="solid", fgColor="FFD1FAE5")    # verde claro para ROBOT
    POST_TINT = PF(fill_type="solid", fgColor="FFFCE7F3")   # rosa claro para POST
    NA_TINT = PF(fill_type="solid", fgColor="FFE2E8F0")     # gris claro para N/A
    MAQ_TINT = PF(fill_type="solid", fgColor="FFEDE9FE")    # purpura muy claro para MAQUILA
    # Mapping col -> fill
    COL_TINTS = {
        4: PRE_TINT, 5: PRE_TINT, 6: PRE_TINT, 7: PRE_TINT,
        8: ROB_TINT, 9: ROB_TINT, 10: ROB_TINT, 11: ROB_TINT,
        12: POST_TINT, 13: POST_TINT, 14: POST_TINT, 15: POST_TINT,
        16: NA_TINT, 17: NA_TINT,
        18: MAQ_TINT, 19: MAQ_TINT,
    }

    def _estilo_fila(ws_, r_, fill=None, font=None, align=None, border=THIN_BORDER, cols=range(2, 21)):
        for c_ in cols:
            try:
                cell = ws_.cell(row=r_, column=c_)
                if fill is not None: cell.fill = fill
                if font is not None: cell.font = font
                if align is not None: cell.alignment = align
                if border is not None: cell.border = border
            except AttributeError:
                pass

    for s in semanas:
        sn = f"Sem {s}"
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        activos = [(n, asig[n][s]) for n in asig if asig[n][s] > 0]
        r = 6
        for n, pares in activos:
            safe_set(ws, r, 2, n)
            safe_set(ws, r, 3, pares)
            i = info.get(n, {})
            if i.get("sin_catalogo"):
                safe_set(ws, r, 4, "SIN CATALOGO")
                safe_set(ws, r, 20, "—")
                _estilo_fila(ws, r, fill=PINK_FILL, font=SIN_CAT_FONT, align=CENTER_AL)
                ws.cell(row=r, column=2).alignment = LEFT_AL
            else:
                if i.get("PRE", 0) > 0:
                    safe_set(ws, r, 4, i["PRE"])
                    safe_set(ws, r, 5, f"=C{r}*D{r}/3600")
                    safe_set(ws, r, 6, f"=E{r}/$F$2")
                    safe_set(ws, r, 7, f"=F{r}/$C$2")
                else: safe_set(ws, r, 4, "—")
                if i.get("ROB", 0) > 0:
                    safe_set(ws, r, 8, i["ROB"])
                    safe_set(ws, r, 9, f"=C{r}*H{r}/3600")
                    safe_set(ws, r, 10, f"=I{r}/$J$2")
                    safe_set(ws, r, 11, f"=J{r}/$C$2")
                else: safe_set(ws, r, 8, "—")
                if i.get("POST", 0) > 0:
                    safe_set(ws, r, 12, i["POST"])
                    safe_set(ws, r, 13, f"=C{r}*L{r}/3600")
                    safe_set(ws, r, 14, f"=M{r}/$N$2")
                    safe_set(ws, r, 15, f"=N{r}/$C$2")
                else: safe_set(ws, r, 12, "—")
                if i.get("NA", 0) > 0:
                    safe_set(ws, r, 16, i["NA"])
                    safe_set(ws, r, 17, f"=C{r}*P{r}/3600")
                else: safe_set(ws, r, 16, "—")
                if i.get("MAQ", 0) > 0:
                    safe_set(ws, r, 18, i["MAQ"])
                    safe_set(ws, r, 19, f"=C{r}*R{r}/3600")
                else: safe_set(ws, r, 18, "—")
                parts = []
                for col, key in [("E","PRE"),("I","ROB"),("M","POST"),("Q","NA"),("S","MAQ")]:
                    if i.get(key, 0) > 0: parts.append(f"{col}{r}")
                safe_set(ws, r, 20, "=" + "+".join(parts) if parts else 0)
                # Estilo de fila normal: forzar font negro normal (size 10), border, alineación
                from openpyxl.styles import Font as _FtN
                NORMAL_FONT = _FtN(bold=False, italic=False, color="FF000000", size=10)
                MODELO_FONT = _FtN(bold=True, italic=False, color="FF000000", size=10)
                _estilo_fila(ws, r, font=NORMAL_FONT, align=CENTER_AL)
                ws.cell(row=r, column=2).alignment = LEFT_AL
                ws.cell(row=r, column=2).font = MODELO_FONT
                ws.cell(row=r, column=20).font = MODELO_FONT
                # Aplicar tints suaves por columna (PRELIM/ROBOT/POST/NA/MAQ)
                for col_idx, tint in COL_TINTS.items():
                    try:
                        ws.cell(row=r, column=col_idx).fill = tint
                    except AttributeError:
                        pass
            # Forzar altura uniforme para TODAS las filas de datos
            ws.row_dimensions[r].height = 33
            r += 1
        rt = r; first = 6
        safe_set(ws, rt, 2, "TOTAL")
        if rt > first:
            # Hay al menos un modelo: SUM normal
            safe_set(ws, rt, 3, f"=SUM(C{first}:C{rt-1})")
            for col_letter in ["E","F","G","I","J","K","M","N","O","Q","S","T"]:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                safe_set(ws, rt, col_idx, f"=SUM({col_letter}{first}:{col_letter}{rt-1})")
        else:
            # Semana sin modelos: valores fijos en 0 para evitar referencia circular
            safe_set(ws, rt, 3, 0)
            for col_letter in ["E","F","G","I","J","K","M","N","O","Q","S","T"]:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                safe_set(ws, rt, col_idx, 0)
        # Estilo TOTAL: fondo azul oscuro + texto blanco bold + altura uniforme
        _estilo_fila(ws, rt, fill=TOTAL_FILL, font=TOTAL_FONT, align=CENTER_AL)
        ws.row_dimensions[rt].height = 33
        # Ocultar filas vacías después del TOTAL hasta el max_row del template
        for r_hide in range(rt + 1, max(ws.max_row, rt + 30) + 1):
            ws.row_dimensions[r_hide].hidden = True

    # ===== Hoja "Total General" =====
    if "Total General" in wb.sheetnames:
        ws = wb["Total General"]
        template_rows_tg = {}
        for r in range(6, 35):
            nm = ws.cell(row=r, column=2).value
            if not nm or str(nm).strip().upper() == "TOTAL":
                continue
            num = _modelo_num(str(nm))
            if num:
                template_rows_tg[r] = num

        # Estilo del template para filas de Total General
        from openpyxl.styles import Font as FtTG
        TG_FONT_NORMAL = FtTG(bold=False, size=10)
        TG_FONT_MODELO = FtTG(bold=True, size=10)
        TG_FONT_TOTAL_HRS = FtTG(bold=True, size=10)
        # Tints por columna (mismos colores del template)
        TG_TINTS = {
            4: PRE_TINT, 5: PRE_TINT, 6: PRE_TINT, 7: PRE_TINT,
            8: ROB_TINT, 9: ROB_TINT, 10: ROB_TINT, 11: ROB_TINT,
            12: POST_TINT, 13: POST_TINT, 14: POST_TINT, 15: POST_TINT,
            16: NA_TINT, 17: NA_TINT,
            18: MAQ_TINT, 19: MAQ_TINT,
        }

        def _aplicar_estilo_tg(ws_, r_, sin_cat=False):
            """Aplica el estilo idéntico al template a una fila de Total General."""
            ws_.row_dimensions[r_].height = 33
            for c in range(2, 21):
                try:
                    cell = ws_.cell(row=r_, column=c)
                    if c == 2:
                        cell.alignment = LEFT_AL
                        cell.font = TG_FONT_MODELO
                    elif c == 20:
                        cell.alignment = CENTER_AL
                        cell.font = TG_FONT_TOTAL_HRS
                    else:
                        cell.alignment = CENTER_AL
                        cell.font = TG_FONT_NORMAL
                    # Tint por columna
                    if sin_cat:
                        cell.fill = PINK_FILL
                    elif c in TG_TINTS:
                        cell.fill = TG_TINTS[c]
                    else:
                        cell.fill = NO_FILL
                except AttributeError:
                    pass

        # Limpiar valores Y fills
        for r in range(6, 41):
            for c in range(2, 21):
                safe_set(ws, r, c, None)
                try:
                    ws.cell(row=r, column=c).fill = NO_FILL
                except AttributeError:
                    pass
            ws.row_dimensions[r].hidden = False

        asignados_tg = set()
        for r, num in template_rows_tg.items():
            input_match = input_por_num.get(num)
            if input_match:
                _escribir_total_general_row(ws, r, input_match, asig[input_match], info[input_match])
                _aplicar_estilo_tg(ws, r, sin_cat=info[input_match].get("sin_catalogo", False))
                asignados_tg.add(input_match)
            else:
                ws.row_dimensions[r].hidden = True
        last_row_tg = max(template_rows_tg.keys()) if template_rows_tg else 5
        for nm in input_modelos:
            if nm in asignados_tg: continue
            last_row_tg += 1
            _escribir_total_general_row(ws, last_row_tg, nm, asig[nm], info[nm])
            _aplicar_estilo_tg(ws, last_row_tg, sin_cat=info[nm].get("sin_catalogo", False))
        rt_tg = last_row_tg + 1
        safe_set(ws, rt_tg, 2, "TOTAL")
        for col_letter in ["C","E","F","G","I","J","K","M","N","O","Q","S","T"]:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            safe_set(ws, rt_tg, col_idx, f"=SUM({col_letter}6:{col_letter}{rt_tg - 1})")
        # Estilo TOTAL: azul oscuro + bold blanco
        ws.row_dimensions[rt_tg].height = 33
        for c in range(2, 21):
            try:
                cell = ws.cell(row=rt_tg, column=c)
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
                cell.alignment = CENTER_AL
            except AttributeError:
                pass
        # Formato 1 decimal en horas/personas/dias
        for r in range(6, rt_tg + 1):
            for col_letter in ["E","F","G","I","J","K","M","N","O","Q","S","T"]:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "0.0"
        # Ocultar filas vacías post-TOTAL
        for r_hide in range(rt_tg + 1, max(ws.max_row, rt_tg + 30) + 1):
            ws.row_dimensions[r_hide].hidden = True

    # ===== Hoja "Seg por Par" =====
    if "Seg por Par" in wb.sheetnames:
        ws = wb["Seg por Par"]
        template_rows_sp = {}
        for r in range(4, 35):
            nm = ws.cell(row=r, column=2).value
            if not nm or str(nm).strip().upper() == "TOTAL":
                continue
            num = _modelo_num(str(nm))
            if num:
                template_rows_sp[r] = num
        for r in range(4, 41):
            for c in range(2, 10):
                safe_set(ws, r, c, None)
            ws.row_dimensions[r].hidden = False

        asignados_sp = set()
        for r, num in template_rows_sp.items():
            input_match = input_por_num.get(num)
            if input_match:
                _escribir_seg_por_par_row(ws, r, input_match, info[input_match])
                asignados_sp.add(input_match)
            else:
                ws.row_dimensions[r].hidden = True
        last_row_sp = max(template_rows_sp.keys()) if template_rows_sp else 3
        for nm in input_modelos:
            if nm in asignados_sp: continue
            last_row_sp += 1
            _escribir_seg_por_par_row(ws, last_row_sp, nm, info[nm])

    # Resumen Semanal
    if "Resumen Semanal" in wb.sheetnames:
        ws_r = wb["Resumen Semanal"]
        for r in range(5, ws_r.max_row + 2):
            for c in range(2, 21):
                safe_set(ws_r, r, c, None)
        r = 5
        for s in semanas:
            activos_n = sum(1 for n in asig if asig[n][s] > 0)
            rt = 6 + activos_n
            sn_q = f"'Sem {s}'"
            safe_set(ws_r, r, 2, f"Sem {s}")
            for col, letter in [(3,"C"),(5,"E"),(6,"F"),(7,"G"),(9,"I"),(10,"J"),(11,"K"),(13,"M"),(14,"N"),(15,"O"),(17,"Q"),(19,"S"),(20,"T")]:
                safe_set(ws_r, r, col, f"={sn_q}!{letter}{rt}")
            r += 1
        safe_set(ws_r, r, 2, "TOTAL")
        for col, letter in [(3,"C"),(5,"E"),(6,"F"),(7,"G"),(9,"I"),(10,"J"),(11,"K"),(13,"M"),(14,"N"),(15,"O"),(17,"Q"),(19,"S"),(20,"T")]:
            safe_set(ws_r, r, col, f"=SUM({letter}5:{letter}{r-1})")

    # formato 1 decimal
    COLS_DEC = [5,6,7,9,10,11,13,14,15,17,19,20]
    for sn in [f"Sem {s}" for s in semanas] + ["Resumen Semanal"]:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        for r in range(5, ws.max_row + 1):
            for c in COLS_DEC:
                cell = ws.cell(row=r, column=c)
                if cell.value is not None: cell.number_format = "0.0"

    wb.save(output_path)

# ============================================================
# MAIN
# ============================================================
def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input", required=True, help="Excel del backlog (ruta absoluta o nombre dentro de backlog_tool/inputs/)")
    p.add_argument("--output", help="Excel salida (default: backlog_tool/outputs/<nombre>)")
    p.add_argument("--template", help="Plantilla visual (default: backlog_tool/template_visual.xlsx)")
    p.add_argument("--max-modelos", type=int)
    p.add_argument("--max-sem", type=int)
    p.add_argument("--excluir", default="")
    p.add_argument("--fijar", action="append", default=[], help="MODELO_NUM:S1,S2 (repetible)")
    p.add_argument("--no-aislar", default="")
    p.add_argument("--no-overrides", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    TOOL_DIR = Path(__file__).parent
    INPUTS_DIR = TOOL_DIR / "inputs"
    OUTPUTS_DIR = TOOL_DIR / "outputs"
    OUTPUTS_DIR.mkdir(exist_ok=True)

    # Resolver input: ruta absoluta -> tal cual; si no, buscar en inputs/, luego en cwd
    inp = Path(args.input)
    if not inp.is_absolute():
        candidates = [INPUTS_DIR / args.input, Path.cwd() / args.input, inp]
        for c in candidates:
            if c.exists():
                inp = c; break
    if not inp.exists():
        print(f"[ERROR] No existe: {args.input}", file=sys.stderr)
        print(f"        Busqué en: {INPUTS_DIR}, {Path.cwd()}", file=sys.stderr)
        sys.exit(1)

    if args.output:
        out = Path(args.output)
    else:
        import re
        stem = re.sub(r"[_\- ]?MANUAL", "", inp.stem, flags=re.IGNORECASE).strip("_- ")
        now = datetime.now()
        iso_week = now.isocalendar().week
        out = OUTPUTS_DIR / f"{stem}_OPTIMIZADO_S{iso_week:02d}_{now:%Y%m%d_%H%M}.xlsx"

    # 1) Leer overrides yaml
    ov = {} if args.no_overrides else load_overrides()

    # 2) Combinar opts (CLI > YAML > defaults)
    opts = {
        "max_mod": args.max_modelos or ov.get("max_modelos_semana") or MAX_MOD_SEM_DEF,
        "max_sem": args.max_sem or ov.get("max_semanas_modelo") or MAX_SEM_MOD_DEF,
        "excluir": set([s.strip() for s in args.excluir.split(",") if s.strip()]) | set(ov.get("excluir", [])),
        "no_aislar": set([s.strip() for s in args.no_aislar.split(",") if s.strip()]) | set(ov.get("no_aislar", [])),
        "fijar": dict(ov.get("fijar", {})),
    }
    for f in args.fijar:
        if ":" in f:
            k, v = f.split(":", 1)
            opts["fijar"][k.strip()] = [int(x) for x in v.split(",") if x.strip().isdigit()]

    cap_extra = ov.get("capacidad_robot_extra", 0)

    # 3) Leer backlog
    print(f"[1/5] Leyendo backlog: {inp.name}")
    modelos, semanas = parse_backlog(inp)
    print(f"      {len(modelos)} modelos, semanas {semanas}")
    # filtrar excluidos
    if opts["excluir"]:
        modelos = [(n, t) for n, t in modelos if not any(e in n for e in opts["excluir"])]
        print(f"      excluidos: {opts['excluir']} -> {len(modelos)} restantes")

    # 4) Consultar Supabase
    print("[2/5] Consultando Supabase...")
    robots = fetch_robots_activos()
    catalogo = fetch_catalogo()
    restricciones = fetch_restricciones_activas()
    n_robots_compartidos = sum(1 for r in robots if not r["nombre"].lower().startswith(("maq.","remach","perfor","desheb","cabina","m-cod")))
    cap_robot_sem = max(1, n_robots_compartidos * HRS_SEM_POR_ROBOT + cap_extra)
    print(f"      {n_robots_compartidos} robots compartidos activos -> capacidad {cap_robot_sem} h-robot/sem")
    print(f"      {len(catalogo)} modelos en catálogo, {len(restricciones)} restricciones activas (mayoría precedencias intra-modelo)")

    # 5) Distribuir
    print("[3/5] Aplicando algoritmo greedy con reglas del playbook...")
    asig, carga, info = distribuir(modelos, semanas, catalogo, cap_robot_sem, opts)

    # 6) Validar
    print("[4/5] Validando...")
    errores = []
    for n, t in modelos:
        s = sum(asig[n].values())
        if s != t:
            errores.append(f"  FAIL {n}: {s} != {t}")
    for sem in semanas:
        if carga[sem] > cap_robot_sem * 1.05:
            errores.append(f"  FAIL Sem {sem}: {carga[sem]:.1f} h > capacidad {cap_robot_sem}")
    if errores:
        print("\n".join(errores), file=sys.stderr)

    # 7) Reporte
    print("[5/5] Distribución resultante:")
    print(f"  {'Sem':<5} {'Pares':>7} {'h-Robot':>10} {'%Cap':>6} {'#Mod':>6}")
    for s in semanas:
        nm = sum(1 for n in asig if asig[n][s] > 0)
        pct = carga[s] / cap_robot_sem * 100
        print(f"  {s:<5} {sum(asig[n][s] for n in asig):>7} {carga[s]:>10.1f} {pct:>5.0f}% {nm:>6}")
    print(f"\n  Modelos:")
    for n, t in modelos:
        sus = [s for s in semanas if asig[n][s] > 0]
        print(f"    {n:<22} -> {sus}")

    # 8) Escribir Excel
    if args.dry_run:
        print("\n[dry-run] No se escribe Excel.")
        return
    template = Path(args.template) if args.template else (TOOL_DIR / "template_visual.xlsx")
    if not template.exists():
        print(f"[ERROR] Template no existe: {template}", file=sys.stderr); sys.exit(1)
    print(f"\nEscribiendo: {out}")
    escribir_excel(template, out, asig, info, semanas, catalogo=catalogo)
    print("OK.")

if __name__ == "__main__":
    main()
