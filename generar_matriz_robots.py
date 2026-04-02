"""
Genera Excel con matriz de programas de robot por fracción.
X = programa ya existe en ese robot
Amarillo = falta (impacto alto)
(vacío) = no aplica
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Todos los robots con programas de costura en el sistema (15 robots)
ROBOTS = [
    "2A-3020-M1", "2A-3020-M2",
    "3020-M1", "3020-M2", "3020-M3", "3020-M4", "3020-M5", "3020-M6",
    "6040-M1", "6040-M2", "6040-M3", "6040-M4", "6040-M5",
    "M048-CHACHE", "M049-CHACHE"
]

# (modelo, fracc, operacion, rate, [robots_actuales], [falta_impacto_alto])
# Sin columna "OTROS" (tachados) ni notas a mano
DATA = [
    ("61747", 4, "COSTURA DE FELPA Y GANCHO", 60,
     ["2A-3020-M1", "2A-3020-M2", "M048-CHACHE"],
     ["M049-CHACHE"]),

    ("61748", 4, "COSTURA DE TALON", 60,
     ["6040-M4", "6040-M5", "M048-CHACHE"],
     ["M049-CHACHE"]),

    ("61748", 5, "COSTURA DE CHINELA INTERNA", 60,
     ["2A-3020-M1"],
     ["2A-3020-M2"]),

    ("61748", 6, "COSTURA DE CHINELA EXTERNA", 60,
     ["3020-M4"],
     ["2A-3020-M1", "2A-3020-M2"]),

    ("64197", 1, "COSTURA DE CHINELA", 77,
     ["2A-3020-M1", "2A-3020-M2", "3020-M4", "3020-M6", "6040-M4", "6040-M5"],
     ["M048-CHACHE", "M049-CHACHE"]),

    ("65413", 1, "COSTURA DE CHINELA", 144,
     ["2A-3020-M1", "M049-CHACHE", "6040-M5"],
     ["M048-CHACHE"]),

    ("65413", 2, "COSTURA DE LATIGO", 80,
     ["2A-3020-M1", "6040-M4", "6040-M5"],
     ["2A-3020-M2"]),

    ("65413", 3, "COSTURA DE HEBILLA", 66,
     ["2A-3020-M1", "2A-3020-M2", "6040-M4"],
     ["6040-M5"]),

    ("65422", 1, "COSER CHINELAS CHICAS", 78,
     ["2A-3020-M1", "3020-M6"],
     ["2A-3020-M2"]),

    ("65422", 3, "COSTURA DE CHINELA LARGA", 55,
     ["2A-3020-M1", "2A-3020-M2"],
     []),

    ("65422", 8, "COSTURA DE CHINELA LASER", 82,
     ["2A-3020-M1", "6040-M4"],
     ["2A-3020-M2", "6040-M5"]),

    ("65568", 1, "COSTURA CHINELA", 100,
     ["2A-3020-M1", "2A-3020-M2", "3020-M4"],
     ["6040-M4", "6040-M5"]),

    ("65568", 5, "COSTURA TALON", 65,
     ["6040-M4", "6040-M5", "M048-CHACHE", "M049-CHACHE"],
     []),  # completo

    ("88186", 1, "COSTURA DE CHINELA", 60,
     ["2A-3020-M1", "2A-3020-M2", "6040-M4", "6040-M5", "M049-CHACHE"],
     ["M048-CHACHE"]),

    ("88186", 4, "COSTURA DE GANCHOS", 120,
     ["2A-3020-M1", "2A-3020-M2", "3020-M4", "3020-M6", "6040-M4", "6040-M5"],
     ["M048-CHACHE", "M049-CHACHE"]),

    ("88186", 7, "COSTURA DE TALON", 65,
     ["2A-3020-M1", "2A-3020-M2", "6040-M4", "6040-M5", "M048-CHACHE", "M049-CHACHE"],
     []),  # completo

    ("94750", 4, "COSTURA DE TALON", 90,
     ["2A-3020-M2", "M048-CHACHE"],
     ["2A-3020-M1", "M049-CHACHE", "6040-M4", "6040-M5"]),

    ("94750", 5, "COSTURA DE CHINELA", 60,
     ["2A-3020-M1", "2A-3020-M2"],
     []),
]

# Estilos
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Matriz Robots"

# ── Headers ──
headers = ["MODELO", "FRACC", "OPERACIÓN", "RATE"] + ROBOTS + ["TIENE", "FALTA"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# ── Data rows ──
for i, (modelo, fracc, op, rate, actuales, falta_alto) in enumerate(DATA, 2):
    ws.cell(row=i, column=1, value=modelo).font = BOLD_FONT
    ws.cell(row=i, column=2, value=fracc).font = NORMAL_FONT
    ws.cell(row=i, column=3, value=op).font = NORMAL_FONT
    ws.cell(row=i, column=4, value=rate).font = NORMAL_FONT

    for c in range(1, 5):
        ws.cell(row=i, column=c).border = THIN_BORDER
        ws.cell(row=i, column=c).alignment = CENTER

    count_tiene = 0
    count_falta = 0

    for j, robot in enumerate(ROBOTS):
        col = 5 + j
        cell = ws.cell(row=i, column=col)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        if robot in actuales:
            cell.value = "X"
            cell.fill = GREEN_FILL
            cell.font = Font(bold=True, size=11, color="006100")
            count_tiene += 1
        elif robot in falta_alto:
            cell.value = ""
            cell.fill = YELLOW_FILL
            count_falta += 1
        else:
            cell.value = ""
            cell.font = NORMAL_FONT

    # Columnas resumen
    tiene_cell = ws.cell(row=i, column=5 + len(ROBOTS), value=count_tiene)
    tiene_cell.font = BOLD_FONT
    tiene_cell.alignment = CENTER
    tiene_cell.border = THIN_BORDER
    tiene_cell.fill = GREEN_FILL

    falta_cell = ws.cell(row=i, column=6 + len(ROBOTS), value=count_falta)
    falta_cell.font = BOLD_FONT
    falta_cell.alignment = CENTER
    falta_cell.border = THIN_BORDER
    if count_falta > 0:
        falta_cell.fill = YELLOW_FILL

# ── Resumen por robot (fila inferior) ──
summary_row = len(DATA) + 3
ws.cell(row=summary_row, column=3, value="TOTAL PROGRAMAS:").font = BOLD_FONT
ws.cell(row=summary_row + 1, column=3, value="TOTAL FALTANTES:").font = BOLD_FONT

for j, robot in enumerate(ROBOTS):
    col = 5 + j
    tiene = sum(1 for d in DATA if robot in d[4])
    falta = sum(1 for d in DATA if robot in d[5])

    cell_t = ws.cell(row=summary_row, column=col, value=tiene)
    cell_t.font = BOLD_FONT
    cell_t.alignment = CENTER
    cell_t.border = THIN_BORDER
    cell_t.fill = GREEN_FILL

    cell_f = ws.cell(row=summary_row + 1, column=col, value=falta)
    cell_f.font = BOLD_FONT
    cell_f.alignment = CENTER
    cell_f.border = THIN_BORDER
    if falta > 0:
        cell_f.fill = YELLOW_FILL

# ── Leyenda ──
leg = summary_row + 3
ws.cell(row=leg, column=1, value="LEYENDA:").font = BOLD_FONT
c1 = ws.cell(row=leg + 1, column=1, value="X")
c1.font = Font(bold=True, color="006100")
c1.fill = GREEN_FILL
ws.cell(row=leg + 1, column=2, value="Programa ya existe en el robot").font = NORMAL_FONT

c2 = ws.cell(row=leg + 2, column=1)
c2.fill = YELLOW_FILL
ws.cell(row=leg + 2, column=2, value="Falta programa (impacto alto)").font = NORMAL_FONT

ws.cell(row=leg + 3, column=1, value="").font = NORMAL_FONT
ws.cell(row=leg + 3, column=2, value="No aplica / sin impacto").font = NORMAL_FONT

# ── Ajustar anchos ──
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 7
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 7
for j in range(len(ROBOTS)):
    col_letter = openpyxl.utils.get_column_letter(5 + j)
    ws.column_dimensions[col_letter].width = 15
ws.column_dimensions[openpyxl.utils.get_column_letter(5 + len(ROBOTS))].width = 8
ws.column_dimensions[openpyxl.utils.get_column_letter(6 + len(ROBOTS))].width = 8

ws.freeze_panes = "E2"

OUTPUT = "Matriz_Robots_Programas.xlsx"
wb.save(OUTPUT)
print(f"Archivo generado: {OUTPUT}")
print(f"  {len(DATA)} fracciones × {len(ROBOTS)} robots")
