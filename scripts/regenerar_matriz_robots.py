"""Regenera `Matriz_Robots_Programas.xlsx` desde cero con el formato
original + las 24 nuevas filas para llenar. Ordenado por (modelo, fraccion).

Colores:
- Header azul oscuro con texto blanco
- Columnas de robot: fondo verde claro base; X = verde medio; ! = amarillo medio
- Columnas M048/M049-CHACHE: fondo amarillo claro base
- TIENE / FALTA calculados con formula COUNTIF
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = 'Matriz_Robots_Programas.xlsx'

ROBOTS = [
    '2A-3020-M1', '2A-3020-M2',
    '3020-M1', '3020-M2', '3020-M3', '3020-M4', '3020-M5', '3020-M6',
    '6040-M1', '6040-M2', '6040-M3', '6040-M4', '6040-M5',
    'M048-CHACHE', 'M049-CHACHE',
]
# Primeras cols = robots de costura; ultimas 2 = maquinas chachera (amarillo)
CHACHE_ROBOTS = {'M048-CHACHE', 'M049-CHACHE'}

# Datos: (modelo, fraccion, operacion, rate, {robot: 'X' | '!'})
ORIGINAL = [
    ('61747', 4, 'COSTURA DE FELPA Y GANCHO', 60,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', 'M048-CHACHE': 'X', 'M049-CHACHE': '!'}),
    ('61748', 4, 'COSTURA DE TALON', 60,
     {'6040-M4': 'X', '6040-M5': 'X', 'M048-CHACHE': 'X', 'M049-CHACHE': '!'}),
    ('61748', 5, 'COSTURA DE CHINELA INTERNA', 60,
     {'2A-3020-M1': 'X', '2A-3020-M2': '!'}),
    ('61748', 6, 'COSTURA DE CHINELA EXTERNA', 60,
     {'2A-3020-M1': '!', '2A-3020-M2': '!', '3020-M4': 'X'}),
    ('64197', 1, 'COSTURA DE CHINELA', 77,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', '3020-M4': 'X', '3020-M6': 'X',
      '6040-M4': 'X', '6040-M5': 'X', 'M048-CHACHE': '!', 'M049-CHACHE': '!'}),
    ('65413', 1, 'COSTURA DE CHINELA', 144,
     {'2A-3020-M1': 'X', '6040-M5': 'X', 'M048-CHACHE': '!', 'M049-CHACHE': 'X'}),
    ('65413', 2, 'COSTURA DE LATIGO', 80,
     {'2A-3020-M1': 'X', '2A-3020-M2': '!', '6040-M4': 'X', '6040-M5': 'X'}),
    ('65413', 3, 'COSTURA DE HEBILLA', 66,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', '6040-M4': 'X', '6040-M5': '!'}),
    ('65422', 1, 'COSER CHINELAS CHICAS', 76,
     {'2A-3020-M1': 'X', '2A-3020-M2': '!', '3020-M6': 'X'}),
    ('65422', 3, 'COSTURA DE CHINELA LARGA', 55,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X'}),
    ('65422', 8, 'COSTURA DE CHINELA LASER', 82,
     {'2A-3020-M1': 'X', '2A-3020-M2': '!', '6040-M4': 'X', '6040-M5': '!'}),
    ('65568', 1, 'COSTURA CHINELA', 100,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', '3020-M4': 'X', '6040-M4': '!', '6040-M5': '!'}),
    ('65568', 5, 'COSTURA TALON', 65,
     {'6040-M4': 'X', '6040-M5': 'X', 'M048-CHACHE': 'X', 'M049-CHACHE': 'X'}),
    ('88186', 1, 'COSTURA DE CHINELA', 60,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', '6040-M4': 'X', '6040-M5': 'X',
      'M048-CHACHE': '!', 'M049-CHACHE': 'X'}),
    ('88186', 4, 'COSTURA DE GANCHOS', 120,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', '3020-M4': 'X', '3020-M6': 'X',
      '6040-M4': 'X', '6040-M5': 'X', 'M048-CHACHE': '!', 'M049-CHACHE': '!'}),
    ('88186', 7, 'COSTURA DE TALON', 65,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X', '6040-M4': 'X', '6040-M5': 'X',
      'M048-CHACHE': 'X', 'M049-CHACHE': '!'}),
    ('94750', 4, 'COSTURA DE TALON', 90,
     {'2A-3020-M1': '!', '2A-3020-M2': 'X', '6040-M4': '!', '6040-M5': '!',
      'M048-CHACHE': 'X', 'M049-CHACHE': '!'}),
    ('94750', 5, 'COSTURA DE CHINELA', 60,
     {'2A-3020-M1': 'X', '2A-3020-M2': 'X'}),
]

NEW = [
    ('61747', 3, 'COSTURA DE FELPA Y GANCHO', 60, {}),
    ('62100', 3, 'COSTURA DE INPUT 1 EN CHINELA', 25, {}),
    ('62100', 4, 'COSTURA DE LENGUA', 45, {}),
    ('68127', 1, 'COSTURA CHINELA INTERNA', 82, {}),
    ('68127', 2, 'COSTURA CHINELA EXTERNA', 82, {}),
    ('68127', 4, 'COSTURA GANCHOS', 120, {}),
    ('68127', 6, 'COSTURA TALON', 100, {}),
    ('68127', 7, 'COSTURA COMPLEMENTO DE TALON', 118, {}),
    ('68127', 9, 'COSTURA TALON EXTERNO', 110, {}),
    ('68127', 10, 'COSTURA TALON APLICACION', 80, {}),
    ('69906', 4, 'COSTURA DE INPUT 1 EN CHINELA DERECHA', 68, {}),
    ('69906', 5, 'COSTURA DE INPUT 1 EN CHINELA IZQUIERDA', 68, {}),
    ('69906', 6, 'COSTURA DE LENGUA', 111, {}),
    ('69906', 16, 'CERRADO DE CORTE CON CHINELA (FUERA DE LINEA)', 158, {}),
    ('93346', 3, 'COSTURA DE CHINELA', 51, {}),
    ('93346', 4, 'COSTURA DE EMPEINE', 51, {}),
    ('93346', 5, 'COSTURA DE LATIGO', 150, {}),
    ('93347', 5, 'COSTURA DE TALON', 46, {}),
    ('93347', 6, 'COSTURA DE CHINELA', 73, {}),
    ('93349', 2, 'COSER CHINELA EXTERNA E INTERNA', 67, {}),
    ('94749', 1, 'COSTURA DE CHINELA A TALON', 100, {}),
    ('94751', 3, 'COSTURA DE PIEZA HEBILLA', 100, {}),
    ('94751', 4, 'COSTURA DE CHINELA', 80, {}),
    ('94751', 5, 'COSTURA DE AHORCAPOLLO', 50, {}),
]

# Colores estilo Excel estandar
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

TITLE_FILL = PatternFill('solid', fgColor='1F4E79')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')

# Base tenue por columna
GREEN_BASE = PatternFill('solid', fgColor='E2EFDA')  # robots
YELLOW_BASE = PatternFill('solid', fgColor='FFF2CC')  # chachera

# Celdas con X (programa existe) o ! (falta programa)
X_FILL = PatternFill('solid', fgColor='C6EFCE')
X_FONT = Font(name='Calibri', size=11, bold=True, color='375623')
BANG_FILL = PatternFill('solid', fgColor='FFEB9C')
BANG_FONT = Font(name='Calibri', size=11, bold=True, color='9C5700')

META_FONT = Font(name='Calibri', size=11, bold=True)
DATA_FONT = Font(name='Calibri', size=11)
TOTALS_FONT = Font(name='Calibri', size=11, bold=True)

# Estilo para marcar modelos NUEVOS por completar
NEW_BADGE_FILL = PatternFill('solid', fgColor='DEEBF7')  # azul claro
NEW_BADGE_FONT = Font(name='Calibri', size=11, bold=True, color='1F4E79')
SEPARATOR_FILL = PatternFill('solid', fgColor='BDD7EE')  # azul medio para fila separadora
SEPARATOR_FONT = Font(name='Calibri', size=11, bold=True, color='1F4E79', italic=True)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_chache_col(robot: str) -> bool:
    return robot in CHACHE_ROBOTS


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Matriz Robots'

    total_cols = 4 + len(ROBOTS) + 2  # 21

    # Fila 1: titulo (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1).value = 'MATRIZ DE PROGRAMAS DE ROBOTS — ÁREA PESPUNTE'
    ws.cell(1, 1).fill = TITLE_FILL
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(1, 1).alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Fila 2: headers
    headers = ['MODELO', 'FRACC', 'OPERACIÓN', 'RATE'] + ROBOTS + ['TIENE', 'FALTA']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(2, c)
        cell.value = h
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 36

    # ORIGINAL primero (ordenado por modelo+fraccion), luego separador,
    # luego NEW (ordenado por modelo+fraccion) con badge azul en MODELO
    originales = sorted(ORIGINAL, key=lambda r: (r[0], r[1]))
    nuevos = sorted(NEW, key=lambda r: (r[0], r[1]))

    data_first = 3

    def write_data_row(r: int, modelo, frac, operacion, rate, marks, is_new: bool):
        ws.row_dimensions[r].height = 26

        # MODELO — con badge azul si es nuevo
        c = ws.cell(r, 1)
        c.value = modelo
        c.alignment = CENTER
        c.border = BORDER
        if is_new:
            c.font = NEW_BADGE_FONT
            c.fill = NEW_BADGE_FILL
        else:
            c.font = META_FONT
        # FRACC
        c = ws.cell(r, 2); c.value = frac; c.alignment = CENTER; c.border = BORDER
        if is_new:
            c.font = NEW_BADGE_FONT
            c.fill = NEW_BADGE_FILL
        else:
            c.font = META_FONT
        # OPERACION
        c = ws.cell(r, 3); c.value = operacion; c.font = DATA_FONT; c.alignment = LEFT; c.border = BORDER
        # RATE
        c = ws.cell(r, 4); c.value = rate; c.font = DATA_FONT; c.alignment = CENTER; c.border = BORDER

        # Columnas de robots
        for j, robot in enumerate(ROBOTS):
            col = 5 + j
            cell = ws.cell(r, col)
            cell.alignment = CENTER
            cell.border = BORDER
            mark = marks.get(robot)
            if mark == 'X':
                cell.value = 'X'; cell.fill = X_FILL; cell.font = X_FONT
            elif mark == '!':
                cell.value = '!'; cell.fill = BANG_FILL; cell.font = BANG_FONT
            else:
                cell.font = DATA_FONT  # sin fill: celda blanca

        # TIENE / FALTA (formula por fila)
        row_first_col = get_column_letter(5)
        row_last_col = get_column_letter(4 + len(ROBOTS))
        tiene = ws.cell(r, 4 + len(ROBOTS) + 1)
        tiene.value = f'=COUNTIF({row_first_col}{r}:{row_last_col}{r},"X")'
        tiene.font = TOTALS_FONT; tiene.alignment = CENTER; tiene.border = BORDER
        tiene.fill = GREEN_BASE
        falta = ws.cell(r, 4 + len(ROBOTS) + 2)
        falta.value = f'=COUNTIF({row_first_col}{r}:{row_last_col}{r},"!")'
        falta.font = TOTALS_FONT; falta.alignment = CENTER; falta.border = BORDER
        falta.fill = YELLOW_BASE

    # Escribir originales
    cur = data_first
    for (modelo, frac, op, rate, marks) in originales:
        write_data_row(cur, modelo, frac, op, rate, marks, is_new=False)
        cur += 1

    # Fila separadora
    sep_row = cur
    ws.row_dimensions[sep_row].height = 22
    ws.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=total_cols)
    sep_cell = ws.cell(sep_row, 1)
    sep_cell.value = '▼ MODELOS NUEVOS — POR COMPLETAR ▼'
    sep_cell.alignment = CENTER
    sep_cell.fill = SEPARATOR_FILL
    sep_cell.font = SEPARATOR_FONT
    sep_cell.border = BORDER
    cur += 1

    # Escribir nuevos
    for (modelo, frac, op, rate, marks) in nuevos:
        write_data_row(cur, modelo, frac, op, rate, marks, is_new=True)
        cur += 1

    data_last = cur - 1

    # Filas de totales (una en blanco + TOTAL PROGRAMAS + TOTAL FALTANTES)
    totales_first = data_last + 2

    # TOTAL PROGRAMAS
    r = totales_first
    ws.cell(r, 3).value = 'TOTAL PROGRAMAS:'
    ws.cell(r, 3).font = TOTALS_FONT; ws.cell(r, 3).alignment = Alignment(horizontal='right', vertical='center')
    for j, robot in enumerate(ROBOTS):
        col = 5 + j
        col_letter = get_column_letter(col)
        cell = ws.cell(r, col)
        cell.value = f'=COUNTIF({col_letter}{data_first}:{col_letter}{data_last},"X")'
        cell.font = TOTALS_FONT; cell.alignment = CENTER; cell.border = BORDER
        cell.fill = YELLOW_BASE if is_chache_col(robot) else GREEN_BASE
    # Columna total (T) con total general de X
    col = 4 + len(ROBOTS) + 1  # TIENE
    col_letter = get_column_letter(col)
    cell = ws.cell(r, col)
    cell.value = f'=SUM({col_letter}{data_first}:{col_letter}{data_last})'
    cell.font = TOTALS_FONT; cell.alignment = CENTER; cell.border = BORDER
    cell.fill = GREEN_BASE
    ws.row_dimensions[r].height = 24

    # TOTAL FALTANTES
    r = totales_first + 1
    ws.cell(r, 3).value = 'TOTAL FALTANTES:'
    ws.cell(r, 3).font = TOTALS_FONT; ws.cell(r, 3).alignment = Alignment(horizontal='right', vertical='center')
    for j, robot in enumerate(ROBOTS):
        col = 5 + j
        col_letter = get_column_letter(col)
        cell = ws.cell(r, col)
        cell.value = f'=COUNTIF({col_letter}{data_first}:{col_letter}{data_last},"!")'
        cell.font = TOTALS_FONT; cell.alignment = CENTER; cell.border = BORDER
        cell.fill = YELLOW_BASE if is_chache_col(robot) else GREEN_BASE
    # Columna total (U) con total general de !
    col = 4 + len(ROBOTS) + 2  # FALTA
    col_letter = get_column_letter(col)
    cell = ws.cell(r, col)
    cell.value = f'=SUM({col_letter}{data_first}:{col_letter}{data_last})'
    cell.font = TOTALS_FONT; cell.alignment = CENTER; cell.border = BORDER
    cell.fill = YELLOW_BASE
    ws.row_dimensions[r].height = 24

    # Leyenda
    leyenda_first = totales_first + 3
    ws.cell(leyenda_first, 1).value = 'LEYENDA:'
    ws.cell(leyenda_first, 1).font = TOTALS_FONT

    items = [
        ('X', 'Programa ya existe en el robot', X_FILL, X_FONT),
        ('!', 'Falta programa (impacto alto)', BANG_FILL, BANG_FONT),
        ('', 'No aplica / sin impacto', None, None),
        ('NEW', 'Modelo/fraccion nueva — pendiente de completar', NEW_BADGE_FILL, NEW_BADGE_FONT),
    ]
    for i, (sym, desc, fill, font) in enumerate(items):
        r = leyenda_first + 1 + i
        cell = ws.cell(r, 1)
        cell.value = sym
        cell.alignment = CENTER
        cell.border = BORDER
        if fill: cell.fill = fill
        if font: cell.font = font
        ws.cell(r, 2).value = desc
        ws.cell(r, 2).font = DATA_FONT
        ws.cell(r, 2).alignment = LEFT

    # Anchos de columna
    widths = {
        1: 10, 2: 7, 3: 38, 4: 7,
    }
    for i, _ in enumerate(ROBOTS):
        widths[5 + i] = 12
    widths[5 + len(ROBOTS)] = 8      # TIENE
    widths[5 + len(ROBOTS) + 1] = 8  # FALTA
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze panes (filas hasta 2 y columnas hasta D)
    ws.freeze_panes = 'E3'

    wb.save(OUTPUT)
    print(f'Guardado: {OUTPUT} | {len(ORIGINAL) + len(NEW)} filas de datos ({len(ORIGINAL)} originales + {len(NEW)} nuevas)')


if __name__ == '__main__':
    build()
