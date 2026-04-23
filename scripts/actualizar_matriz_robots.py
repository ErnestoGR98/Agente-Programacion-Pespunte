"""Agrega nuevas operaciones al Excel `Matriz_Robots_Programas.xlsx`
preservando el formato original.

Reglas:
- NO modifica filas existentes (valores, estilos, alturas).
- Copia altura de fila desde una fila de datos existente.
- Para cada celda de robot en las nuevas filas, copia el fill DE UNA CELDA
  VACIA de la misma columna en las filas existentes (ese fill representa el
  color base: verde claro para robots con programas o amarillo claro para
  las chacheras). Si no encuentra celda vacia, toma cualquiera como
  fallback.
- Inserta las nuevas filas al final de los datos (no reordena los
  existentes).
"""
from copy import copy
import openpyxl
from openpyxl.styles import Alignment

PATH = 'Matriz_Robots_Programas.xlsx'

# (modelo_num, fraccion, operacion, rate)
NEW_OPS = [
    ('61747', 3, 'COSTURA DE FELPA Y GANCHO', 60),
    ('62100', 3, 'COSTURA DE INPUT 1 EN CHINELA', 25),
    ('62100', 4, 'COSTURA DE LENGUA', 45),
    ('68127', 1, 'COSTURA CHINELA INTERNA', 82),
    ('68127', 2, 'COSTURA CHINELA EXTERNA', 82),
    ('68127', 4, 'COSTURA GANCHOS', 120),
    ('68127', 6, 'COSTURA TALON', 100),
    ('68127', 7, 'COSTURA COMPLEMENTO DE TALON', 118),
    ('68127', 9, 'COSTURA TALON EXTERNO', 110),
    ('68127', 10, 'COSTURA TALON APLICACION', 80),
    ('69906', 4, 'COSTURA DE INPUT 1 EN CHINELA DERECHA', 68),
    ('69906', 5, 'COSTURA DE INPUT 1 EN CHINELA IZQUIERDA', 68),
    ('69906', 6, 'COSTURA DE LENGUA', 111),
    ('69906', 16, 'CERRADO DE CORTE CON CHINELA (FUERA DE LINEA)', 158),
    ('93346', 3, 'COSTURA DE CHINELA', 51),
    ('93346', 4, 'COSTURA DE EMPEINE', 51),
    ('93346', 5, 'COSTURA DE LATIGO', 150),
    ('93347', 5, 'COSTURA DE TALON', 46),
    ('93347', 6, 'COSTURA DE CHINELA', 73),
    ('93349', 2, 'COSER CHINELA EXTERNA E INTERNA', 67),
    ('94749', 1, 'COSTURA DE CHINELA A TALON', 100),
    ('94751', 3, 'COSTURA DE PIEZA HEBILLA', 100),
    ('94751', 4, 'COSTURA DE CHINELA', 80),
    ('94751', 5, 'COSTURA DE AHORCAPOLLO', 50),
]


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb['Matriz Robots']

    # Detectar rango de filas de datos: empiezan en fila 2 hasta que
    # MODELO deje de ser numerico
    data_first = 2
    data_last = 1
    existing_keys: set[tuple[str, int]] = set()
    for r in range(2, ws.max_row + 1):
        modelo = ws.cell(r, 1).value
        frac = ws.cell(r, 2).value
        if modelo is None:
            continue
        if not str(modelo).strip().isdigit():
            break
        data_last = r
        try:
            existing_keys.add((str(modelo), int(frac)))
        except (TypeError, ValueError):
            pass

    print(f'Filas de datos existentes: {data_first}..{data_last} ({data_last - data_first + 1} filas)')

    # Para cada columna (1..21), encontrar celda vacia sin X/! para copiar su fill
    # (color base de la columna). Tambien capturar font/border/alignment/number_format.
    col_styles: dict[int, dict] = {}
    for c in range(1, 22):
        chosen = None
        # Prioridad 1: celda vacia (sin value)
        for r in range(data_first, data_last + 1):
            cell = ws.cell(r, c)
            if cell.value is None:
                chosen = cell
                break
        # Prioridad 2: cualquier celda (fallback)
        if chosen is None:
            chosen = ws.cell(data_first, c)
        col_styles[c] = {
            'font': copy(chosen.font),
            'border': copy(chosen.border),
            'alignment': copy(chosen.alignment),
            'fill': copy(chosen.fill),
            'number_format': chosen.number_format,
        }

    # Altura de fila: tomar del promedio de las filas existentes, o del primera
    ref_height = None
    for r in range(data_first, data_last + 1):
        dim = ws.row_dimensions.get(r)
        if dim is not None and dim.height:
            ref_height = dim.height
            break

    # Determinar cuantas nuevas filas realmente se agregan (las que no existen ya)
    new_rows = [(m, f, op, rate) for (m, f, op, rate) in NEW_OPS if (m, f) not in existing_keys]
    print(f'Nuevas filas a agregar: {len(new_rows)}')

    # Insertar filas despues de data_last, empujando totales/leyenda hacia abajo
    insert_at = data_last + 1
    ws.insert_rows(insert_at, amount=len(new_rows))

    # Escribir cada fila nueva
    for i, (m, f, op, rate) in enumerate(new_rows):
        r = insert_at + i
        values = [m, f, op, rate] + [None] * 15 + [0, 0]  # 4 meta + 15 robots + TIENE + FALTA
        for c, val in enumerate(values, start=1):
            cell = ws.cell(r, c)
            cell.value = val
            st = col_styles[c]
            cell.font = copy(st['font'])
            cell.border = copy(st['border'])
            cell.alignment = copy(st['alignment'])
            cell.fill = copy(st['fill'])
            cell.number_format = st['number_format']
        if ref_height is not None:
            ws.row_dimensions[r].height = ref_height

    wb.save(PATH)
    print(f'Guardado con {len(new_rows)} filas nuevas. Altura ref: {ref_height}')


if __name__ == '__main__':
    main()
