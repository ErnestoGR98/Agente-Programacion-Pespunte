"""
data_manager.py - Persistencia de datos y generacion de templates Excel.

Maneja:
  - Catalogo de operaciones: guardado/carga en JSON, importacion desde Excel
  - Pedidos semanales: guardado/carga en JSON, importacion desde Excel template
  - Resultados de optimizacion: guardado/carga en JSON para acceso multi-equipo
  - Generacion de templates Excel vacios para que el usuario los llene
"""

import json
import re
import io
from pathlib import Path
from copy import deepcopy
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Directorio base de datos (pespunte-agent/data/)
DATA_DIR = Path(__file__).parent.parent.parent / "data"
CATALOG_JSON = DATA_DIR / "catalogo.json"
PEDIDOS_DIR = DATA_DIR / "pedidos"
RESULTADOS_DIR = DATA_DIR / "resultados"
OPERARIOS_JSON = DATA_DIR / "operarios.json"
RESTRICCIONES_JSON = DATA_DIR / "restricciones.json"
AVANCE_JSON = DATA_DIR / "avance.json"

# Tipos de recurso validos (categorias fisicas base)
VALID_RESOURCES = {"MESA", "ROBOT", "PLANA", "POSTE", "MAQUILA"}

# Configuraciones validas (informativas, no afectan optimizacion)
VALID_CONFIGURACIONES = {"LINEA", "INDIVIDUAL", ""}

# Mapeo legacy: tipos compuestos -> tipo base
_COMPOUND_TO_BASE = {
    "MESA-LINEA": "MESA",
    "PLANA-LINEA": "PLANA",
    "POSTE-LINEA": "POSTE",
}

# Mapeo legacy: tipos compuestos -> configuracion
_COMPOUND_TO_CONFIG = {
    "MESA-LINEA": "LINEA",
    "PLANA-LINEA": "LINEA",
    "POSTE-LINEA": "LINEA",
}


def _get_valid_robots() -> set:
    """Lee robots validos desde config.json."""
    from config_manager import get_physical_robots
    return set(get_physical_robots())


def _get_robot_aliases() -> dict:
    """Lee aliases de robots desde config.json."""
    from config_manager import get_robot_aliases
    return get_robot_aliases()

# Estilos comunes para templates Excel
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Catalogo: persistencia JSON
# ---------------------------------------------------------------------------

def save_catalog(catalog: dict):
    """Guarda el catalogo completo a JSON."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(CATALOG_JSON, "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)


def load_catalog() -> dict | None:
    """Carga el catalogo desde JSON. Migra tipos compuestos si los encuentra."""
    if not CATALOG_JSON.exists():
        return None
    with open(CATALOG_JSON, "r", encoding="utf-8") as f:
        catalog = json.load(f)
    if _migrate_catalog_resources(catalog):
        save_catalog(catalog)
    return catalog


def _migrate_catalog_resources(catalog: dict) -> bool:
    """Migra tipos compuestos de recurso a base + configuracion. Retorna True si migro."""
    migrated = False
    for model_data in catalog.values():
        for op in model_data.get("operations", []):
            recurso = op.get("recurso", "")
            if recurso in _COMPOUND_TO_BASE:
                op["configuracion"] = _COMPOUND_TO_CONFIG[recurso]
                op["recurso"] = _COMPOUND_TO_BASE[recurso]
                migrated = True
            elif "configuracion" not in op:
                if recurso in ("MESA", "PLANA", "POSTE"):
                    op["configuracion"] = "INDIVIDUAL"
                else:
                    op["configuracion"] = ""
                migrated = True
        # Recomputar resource_summary
        if migrated and "operations" in model_data:
            resource_summary = {}
            for op in model_data["operations"]:
                r = op["recurso"]
                resource_summary[r] = resource_summary.get(r, 0) + 1
            model_data["resource_summary"] = resource_summary
    return migrated


def delete_catalog_model(model_num: str) -> bool:
    """Elimina un modelo del catalogo. Retorna True si se elimino."""
    catalog = load_catalog()
    if catalog and model_num in catalog:
        del catalog[model_num]
        save_catalog(catalog)
        return True
    return False


def save_catalog_model(model_num: str, model_data: dict):
    """Guarda o actualiza un modelo individual en el catalogo."""
    catalog = load_catalog() or {}
    catalog[model_num] = model_data
    save_catalog(catalog)


# ---------------------------------------------------------------------------
# Catalogo: importacion desde Excel existente (formato CATALOGO DE FRACCIONES)
# ---------------------------------------------------------------------------

def import_catalog_from_existing_excel(file_or_path) -> dict:
    """
    Importa catalogo desde el Excel existente (formato CATALOGO DE FRACCIONES).
    Usa la misma logica de catalog_loader.py pero guarda como JSON.

    Args:
        file_or_path: ruta a archivo o UploadedFile de Streamlit

    Returns:
        dict con el catalogo parseado
    """
    import sys
    src_dir = str(Path(__file__).parent.parent)
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)
    from catalog_loader import load_catalog_v2

    # Si es un UploadedFile de Streamlit, guardar temporalmente
    if hasattr(file_or_path, "read"):
        import tempfile, os
        suffix = os.path.splitext(file_or_path.name)[1] if hasattr(file_or_path, "name") else ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_or_path.read())
            tmp_path = tmp.name
        file_or_path.seek(0)  # Reset para posibles lecturas futuras
        catalog = load_catalog_v2(tmp_path)
        os.unlink(tmp_path)
    else:
        catalog = load_catalog_v2(str(file_or_path))

    # Guardar como JSON
    save_catalog(catalog)
    return catalog


# ---------------------------------------------------------------------------
# Catalogo: importacion desde template propio
# ---------------------------------------------------------------------------

def import_catalog_from_template(file_or_bytes) -> tuple:
    """
    Importa catalogo desde el template Excel propio.

    Template tiene columnas: MODELO, FRACCION, OPERACION, RECURSO, RATE,
    ROBOT_1, ROBOT_2, ..., ROBOT_8

    Returns:
        (catalog_dict, errors_list)
    """
    if hasattr(file_or_bytes, "read"):
        wb = openpyxl.load_workbook(file_or_bytes, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_or_bytes), data_only=True)

    ws = wb.active
    errors = []
    raw_ops = {}

    # Detectar formato de robots: header=robot_name con OK, o ROBOT_1..8 con nombres
    robot_col_map = {}  # col -> robot_name (nuevo formato)
    is_new_format = False
    valid_robots = _get_valid_robots()
    for col in range(7, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            h = str(header).strip()
            if h in valid_robots:
                robot_col_map[col] = h
                is_new_format = True
            else:
                normalized = _normalize_robot(h)
                if normalized and not h.startswith("ROBOT_"):
                    robot_col_map[col] = normalized
                    is_new_format = True

    for row in range(2, ws.max_row + 1):
        modelo = ws.cell(row=row, column=1).value
        fabrica_val = ws.cell(row=row, column=2).value
        fraccion = ws.cell(row=row, column=3).value
        operacion = ws.cell(row=row, column=4).value
        recurso = ws.cell(row=row, column=5).value
        rate = ws.cell(row=row, column=6).value

        if not modelo or not fraccion:
            continue

        modelo_str = str(modelo).strip()
        model_num = re.match(r"^(\d+)", modelo_str)
        if not model_num:
            errors.append(f"Fila {row}: MODELO '{modelo_str}' no inicia con numero")
            continue
        model_num = model_num.group(1)

        fabrica_str = str(fabrica_val).strip() if fabrica_val else ""

        # Validar recurso (migrar legacy compuestos)
        recurso_str = str(recurso).strip().upper() if recurso else ""
        recurso_str = _COMPOUND_TO_BASE.get(recurso_str, recurso_str)
        if not recurso_str:
            operacion_name = str(operacion).strip() if operacion else f"OP-{fraccion}"
            errors.append(f"⚠ Fila {row}: '{model_num} / {operacion_name}' sin RECURSO asignado (usando GENERAL)")
        elif recurso_str not in VALID_RESOURCES:
            errors.append(f"Fila {row}: RECURSO '{recurso_str}' no valido. "
                          f"Debe ser: {', '.join(sorted(VALID_RESOURCES))}")
            continue

        # Validar rate
        try:
            rate_val = float(rate)
            if rate_val <= 0:
                raise ValueError()
        except (TypeError, ValueError):
            errors.append(f"Fila {row}: RATE '{rate}' debe ser numerico > 0")
            continue

        # Parsear robots: nuevo formato (header=robot, celda=OK) o legacy (celda=nombre)
        robots = []
        if is_new_format:
            for col, robot_name in robot_col_map.items():
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().upper() in ("OK", "SI", "X", "1"):
                    if robot_name not in robots:
                        robots.append(robot_name)
        else:
            for col in range(7, 15):
                val = ws.cell(row=row, column=col).value
                robot = _normalize_robot(val)
                if robot and robot not in robots:
                    robots.append(robot)

        sec_per_pair = round(3600.0 / rate_val)

        if model_num not in raw_ops:
            raw_ops[model_num] = {"codigo_full": modelo_str, "fabrica": fabrica_str, "ops": []}
        elif fabrica_str and not raw_ops[model_num].get("fabrica"):
            raw_ops[model_num]["fabrica"] = fabrica_str

        raw_ops[model_num]["ops"].append({
            "fraccion": int(fraccion),
            "operacion": str(operacion).strip() if operacion else f"OP-{fraccion}",
            "etapa": "",
            "recurso": recurso_str or "GENERAL",
            "recurso_raw": recurso_str,
            "robots": robots,
            "rate": round(rate_val, 2),
            "sec_per_pair": sec_per_pair,
        })

    # Construir catalogo final
    catalog = {}
    for model_num, data in raw_ops.items():
        ops = sorted(data["ops"], key=lambda x: x["fraccion"])
        # Deduplicar por fraccion
        seen = set()
        unique_ops = []
        for op in ops:
            if op["fraccion"] not in seen:
                seen.add(op["fraccion"])
                unique_ops.append(op)

        total_sec = sum(op["sec_per_pair"] for op in unique_ops)
        resource_summary = {}
        for op in unique_ops:
            r = op["recurso"]
            resource_summary[r] = resource_summary.get(r, 0) + 1

        robot_ops = sum(1 for op in unique_ops if op.get("robots"))
        all_robots = set()
        for op in unique_ops:
            for r in op.get("robots", []):
                all_robots.add(r)

        catalog[model_num] = {
            "codigo_full": data["codigo_full"],
            "fabrica": data.get("fabrica", ""),
            "operations": unique_ops,
            "total_sec_per_pair": total_sec,
            "num_ops": len(unique_ops),
            "resource_summary": resource_summary,
            "robot_ops": robot_ops,
            "robots_used": sorted(all_robots),
        }

    if catalog:
        save_catalog(catalog)

    return catalog, errors


def _normalize_robot(val):
    """Normaliza nombre de robot usando config."""
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    aliases = _get_robot_aliases()
    valid = _get_valid_robots()
    if s in aliases:
        s = aliases[s]
    if s in valid:
        return s
    s_upper = s.upper().replace(" ", "")
    for robot in valid:
        if robot.upper().replace(" ", "") == s_upper:
            return robot
    return None


# ---------------------------------------------------------------------------
# Pedidos: persistencia JSON
# ---------------------------------------------------------------------------

def save_pedido(name: str, pedido: list):
    """
    Guarda un pedido semanal.

    Args:
        name: nombre del pedido (ej: "sem_8_2025")
        pedido: lista de dicts con keys: modelo, fabrica, volumen
    """
    PEDIDOS_DIR.mkdir(parents=True, exist_ok=True)
    filepath = PEDIDOS_DIR / f"{name}.json"
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(pedido, f, ensure_ascii=False, indent=2)


def load_pedido(name: str) -> list | None:
    """Carga un pedido por nombre. Migra formato viejo si es necesario."""
    filepath = PEDIDOS_DIR / f"{name}.json"
    if not filepath.exists():
        return None
    with open(filepath, "r", encoding="utf-8") as f:
        pedido = json.load(f)
    # Migrar formato viejo: si 'modelo' contiene "77525 NE" y 'color' es "NE",
    # separar para que 'modelo' sea solo "77525"
    migrated = False
    for item in pedido:
        modelo = item.get("modelo", "")
        color = item.get("color", "")
        m = re.match(r"^(\d{5})\s+([A-Z]{2})$", modelo)
        if m and m.group(2) == color:
            item["modelo"] = m.group(1)
            migrated = True
    if migrated:
        save_pedido(name, pedido)
    return pedido


def list_pedidos() -> list:
    """Lista los nombres de pedidos guardados."""
    PEDIDOS_DIR.mkdir(parents=True, exist_ok=True)
    return sorted([
        f.stem for f in PEDIDOS_DIR.glob("*.json")
    ])


def delete_pedido(name: str) -> bool:
    """Elimina un pedido guardado."""
    filepath = PEDIDOS_DIR / f"{name}.json"
    if filepath.exists():
        filepath.unlink()
        return True
    return False


_DRAFT_NAME = "_borrador"


def save_pedido_draft(pedido: list):
    """Guarda el borrador actual del pedido (auto-save)."""
    save_pedido(_DRAFT_NAME, pedido)


def load_pedido_draft() -> list:
    """Carga el borrador del pedido. Retorna lista vacia si no existe."""
    return load_pedido(_DRAFT_NAME) or []


def clear_pedido_draft():
    """Elimina el borrador del pedido."""
    delete_pedido(_DRAFT_NAME)


# ---------------------------------------------------------------------------
# Pedido: importacion desde template Excel
# ---------------------------------------------------------------------------

def import_pedido_from_template(file_or_bytes) -> tuple:
    """
    Importa pedido semanal desde template Excel.

    Template: MODELO | ALTERNATIVA | CLAVE MATERIAL | FABRICA | VOLUMEN

    Returns:
        (pedido_list, errors_list)
    """
    if hasattr(file_or_bytes, "read"):
        wb = openpyxl.load_workbook(file_or_bytes, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_or_bytes), data_only=True)

    ws = wb.active
    errors = []
    pedido = []

    for row in range(2, ws.max_row + 1):
        modelo = ws.cell(row=row, column=1).value
        color = ws.cell(row=row, column=2).value
        clave = ws.cell(row=row, column=3).value
        fabrica = ws.cell(row=row, column=4).value
        volumen = ws.cell(row=row, column=5).value

        if not modelo:
            continue

        modelo_str = str(modelo).strip()
        if not modelo_str:
            continue

        color_str = str(color).strip() if color else ""
        clave_str = str(clave).strip() if clave else ""

        # Validar fabrica
        fabrica_str = str(fabrica).strip() if fabrica else "FABRICA 1"

        # Validar volumen
        try:
            vol = int(float(volumen))
            if vol <= 0:
                raise ValueError()
        except (TypeError, ValueError):
            errors.append(f"Fila {row}: VOLUMEN '{volumen}' debe ser entero > 0")
            continue

        pedido.append({
            "modelo": modelo_str,
            "color": color_str,
            "clave_material": clave_str,
            "fabrica": fabrica_str,
            "volumen": vol,
        })

    return pedido, errors


# ---------------------------------------------------------------------------
# Conversion: pedido + catalogo -> modelos listos para optimizador
# ---------------------------------------------------------------------------

def build_matched_models(pedido: list, catalog: dict) -> tuple:
    """
    Cruza pedido con catalogo para construir la lista de modelos
    con la misma estructura que espera el optimizador.

    Returns:
        (matched, unmatched) donde:
        - matched: lista de dicts compatibles con optimizer_weekly/v2
        - unmatched: lista de dicts de modelos sin match en catalogo
    """
    matched = []
    unmatched = []

    for item in pedido:
        modelo_str = item["modelo"]
        color = item.get("color", "")
        clave_material = item.get("clave_material", "")
        # Extraer numero de modelo (por si viene con texto extra)
        m = re.match(r"^(\d+)", modelo_str)
        model_num = m.group(1) if m else modelo_str
        # Construir codigo de display: numero + alternativa
        codigo_display = f"{model_num} {color}".strip() if color else model_num

        if model_num in catalog:
            cat = catalog[model_num]
            model = {
                "codigo": codigo_display,
                "modelo_num": model_num,
                "color": color,
                "clave_material": clave_material,
                "suela": "",
                "volumen_declarado": item["volumen"],
                "total_producir": item["volumen"],
                "fabrica": item["fabrica"],
                "daily_prs_original": {},
                "operations": deepcopy(cat["operations"]),
                "total_sec_per_pair": cat["total_sec_per_pair"],
                "num_ops": cat["num_ops"],
                "catalog_code": cat["codigo_full"],
                "resource_summary": cat.get("resource_summary", {}),
            }
            matched.append(model)
        else:
            unmatched.append({
                "codigo": codigo_display,
                "modelo_num": model_num,
                "color": color,
                "clave_material": clave_material,
                "total_producir": item["volumen"],
                "fabrica": item["fabrica"],
            })

    return matched, unmatched


# ---------------------------------------------------------------------------
# Generacion de Templates Excel
# ---------------------------------------------------------------------------

def generate_template_pedido() -> bytes:
    """Genera template Excel vacio para pedido semanal. Retorna bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido Semanal"

    headers = ["MODELO", "ALTERNATIVA", "CLAVE MATERIAL", "FABRICA", "VOLUMEN"]
    col_widths = [15, 12, 16, 15, 12]

    # Encabezados
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Filas de ejemplo
    examples = [
        ("65413", "NE", "SLI", "FABRICA 1", 1700),
        ("95420", "BL", "SLI", "FABRICA 2", 800),
        ("91721", "CA", "SLI", "FABRICA 1", 450),
    ]
    for row, (modelo, color, clave, fab, vol) in enumerate(examples, 2):
        ws.cell(row=row, column=1, value=modelo).font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=color).font = Font(italic=True, color="999999")
        ws.cell(row=row, column=3, value=clave).font = Font(italic=True, color="999999")
        ws.cell(row=row, column=4, value=fab).font = Font(italic=True, color="999999")
        ws.cell(row=row, column=5, value=vol).font = Font(italic=True, color="999999")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = _THIN_BORDER

    # Instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    ws_inst.column_dimensions["A"].width = 80
    instructions = [
        "TEMPLATE - PEDIDO SEMANAL",
        "",
        "Llene la hoja 'Pedido Semanal' con los modelos a producir esta semana.",
        "",
        "Columnas:",
        "  MODELO  - Codigo del modelo (debe existir en el catalogo). Ej: 65413 NE",
        "  FABRICA - Fabrica asignada. Ej: FABRICA 1, FABRICA 2, FABRICA 3",
        "  VOLUMEN - Pares totales a producir en la semana. Debe ser > 0",
        "",
        "Notas:",
        "  - Las filas de ejemplo (en gris) se pueden sobreescribir",
        "  - El MODELO debe coincidir con el catalogo de operaciones",
        "  - No dejar filas vacias entre modelos",
    ]
    for i, line in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=line)
    ws_inst.cell(row=1, column=1).font = Font(bold=True, size=14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_template_catalogo() -> bytes:
    """Genera template Excel vacio para catalogo de operaciones. Retorna bytes."""
    from config_manager import get_physical_robots

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogo"

    robots_sorted = sorted(get_physical_robots())
    headers_fixed = ["MODELO", "FABRICA", "FRACCION", "OPERACION", "RECURSO", "RATE"]
    widths_fixed = [20, 15, 12, 30, 15, 10]
    headers = headers_fixed + robots_sorted
    col_widths = widths_fixed + [5] * len(robots_sorted)

    robot_header_align = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, text_rotation=90)
    robot_header_font = Font(bold=True, color="FFFFFF", size=9)

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        if col <= len(headers_fixed):
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        else:
            cell.font = robot_header_font
            cell.alignment = robot_header_align

    # Filas de ejemplo con OK en robots
    example_font = Font(italic=True, color="999999")
    ok_font = Font(bold=True, color="2E8B57", size=10, italic=True)
    ok_align = Alignment(horizontal="center", vertical="center")
    examples = [
        ("65413 NE", "FABRICA 1", 1, "PEGAR FELPA", "MESA", 120, set()),
        ("65413 NE", "FABRICA 1", 2, "COSER PUNTERA", "ROBOT", 80, {"3020-M4", "3020-M6"}),
        ("65413 NE", "FABRICA 1", 3, "ADORNAR COSTADOS", "PLANA", 90, set()),
        ("95420 NE", "FABRICA 2", 1, "PEGAR FORRO", "MESA", 110, set()),
        ("95420 NE", "FABRICA 2", 2, "COSER LATERAL", "ROBOT", 75, {"6040-M4", "6040-M5"}),
    ]
    for row_num, (modelo, fab, frac, op_name, rec, rate, robots_ok) in enumerate(examples, 2):
        fixed_vals = [modelo, fab, frac, op_name, rec, rate]
        for col, val in enumerate(fixed_vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = example_font
            cell.border = _THIN_BORDER
        for i, robot in enumerate(robots_sorted):
            col = len(headers_fixed) + 1 + i
            cell = ws.cell(row=row_num, column=col)
            cell.border = _THIN_BORDER
            if robot in robots_ok:
                cell.value = "OK"
                cell.font = ok_font
                cell.alignment = ok_align

    # Instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    ws_inst.column_dimensions["A"].width = 80
    instructions = [
        "TEMPLATE - CATALOGO DE OPERACIONES",
        "",
        "Llene la hoja 'Catalogo' con las operaciones de cada modelo.",
        "",
        "Columnas fijas:",
        "  MODELO    - Codigo del modelo. Ej: 65413 NE",
        "  FABRICA   - Fabrica asignada. Ej: FABRICA 1, FABRICA 2, FABRICA 3",
        "  FRACCION  - Numero secuencial de la operacion (1, 2, 3...)",
        "  OPERACION - Descripcion de la operacion. Ej: PEGAR FELPA",
        "  RECURSO   - Tipo de recurso: MESA, ROBOT, PLANA, POSTE, MAQUILA",
        "  RATE      - Pares por hora (numerico, > 0)",
        "",
        "Columnas de robots:",
        "  Cada columna es un robot fisico. Poner OK si la operacion puede",
        "  ejecutarse en ese robot. Solo aplica cuando RECURSO = ROBOT.",
        "",
        "Notas:",
        "  - Las filas de ejemplo (en gris) se pueden sobreescribir",
        "  - Las fracciones deben ser secuenciales por modelo (1, 2, 3...)",
        "  - El MODELO se repite en cada fila de fraccion",
    ]
    for i, line in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=line)
    ws_inst.cell(row=1, column=1).font = Font(bold=True, size=14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_catalog_to_template(catalog: dict) -> bytes:
    """Exporta el catalogo actual al formato template Excel (columnas robot con OK)."""
    from config_manager import get_physical_robots

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogo"

    robots_sorted = sorted(get_physical_robots())
    headers_fixed = ["MODELO", "FABRICA", "FRACCION", "OPERACION", "RECURSO", "RATE"]
    widths_fixed = [20, 15, 12, 30, 15, 10]
    headers = headers_fixed + robots_sorted
    col_widths = widths_fixed + [5] * len(robots_sorted)

    robot_header_align = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, text_rotation=90)
    robot_header_font = Font(bold=True, color="FFFFFF", size=9)
    ok_font = Font(bold=True, color="2E8B57", size=10)
    ok_align = Alignment(horizontal="center", vertical="center")

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        if col <= len(headers_fixed):
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        else:
            cell.font = robot_header_font
            cell.alignment = robot_header_align

    row = 2
    for model_num in sorted(catalog.keys()):
        model_data = catalog[model_num]
        fabrica = model_data.get("fabrica", "")
        for op in model_data["operations"]:
            ws.cell(row=row, column=1, value=model_data["codigo_full"])
            ws.cell(row=row, column=2, value=fabrica)
            ws.cell(row=row, column=3, value=op["fraccion"])
            ws.cell(row=row, column=4, value=op["operacion"])
            ws.cell(row=row, column=5, value=op["recurso"])
            ws.cell(row=row, column=6, value=op["rate"])
            # Robots: OK en columnas correspondientes
            op_robots = set(op.get("robots", []))
            for i, robot in enumerate(robots_sorted):
                col = len(headers_fixed) + 1 + i
                cell = ws.cell(row=row, column=col)
                cell.border = _THIN_BORDER
                if robot in op_robots:
                    cell.value = "OK"
                    cell.font = ok_font
                    cell.alignment = ok_align
            for c in range(1, len(headers_fixed) + 1):
                ws.cell(row=row, column=c).border = _THIN_BORDER
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Resultados de Optimizacion: persistencia JSON
# ---------------------------------------------------------------------------

def _next_version(base_name: str) -> int:
    """Calcula la siguiente version disponible para un base_name."""
    RESULTADOS_DIR.mkdir(parents=True, exist_ok=True)
    max_v = 0
    for f in RESULTADOS_DIR.glob(f"{base_name}_v*.json"):
        stem = f.stem  # ej: sem_8_2026_v3
        suffix = stem[len(base_name):]  # ej: _v3
        if suffix.startswith("_v"):
            try:
                v = int(suffix[2:])
                max_v = max(max_v, v)
            except ValueError:
                pass
    return max_v + 1


def _migrate_legacy_result(base_name: str):
    """Si existe archivo sin version (legacy), renombrarlo a _v1."""
    legacy = RESULTADOS_DIR / f"{base_name}.json"
    if legacy.exists():
        target = RESULTADOS_DIR / f"{base_name}_v1.json"
        if not target.exists():
            legacy.rename(target)
            # Actualizar contenido con campos de version
            try:
                with open(target, "r", encoding="utf-8") as fh:
                    data = json.load(fh)
                data["nombre"] = f"{base_name}_v1"
                data["base_name"] = base_name
                data["version"] = 1
                data.setdefault("nota", "")
                with open(target, "w", encoding="utf-8") as fh:
                    json.dump(data, fh, ensure_ascii=False, indent=2)
            except (json.JSONDecodeError, OSError):
                pass


def parse_version_name(filename_stem: str) -> tuple:
    """Extrae (base_name, version) de un nombre de archivo.

    Returns:
        (base_name, version) o (filename_stem, 0) si no tiene version.
    """
    import re
    m = re.match(r"^(.+)_v(\d+)$", filename_stem)
    if m:
        return m.group(1), int(m.group(2))
    return filename_stem, 0


def save_optimization_results(name: str, weekly_schedule: list,
                               weekly_summary: dict, daily_results: dict,
                               pedido: list = None, params: dict = None,
                               nota: str = "") -> str:
    """
    Guarda resultados de optimizacion a JSON con versionado automatico.

    Args:
        name: nombre base del resultado (ej: "sem_7_2026")
        weekly_schedule: salida de optimizer_weekly
        weekly_summary: resumen semanal
        daily_results: salida de optimizer_v2 (dict por dia)
        pedido: pedido original (opcional, para referencia)
        params: parametros usados (opcional, para referencia)
        nota: comentario opcional sobre esta version

    Returns:
        nombre completo con version (ej: "sem_7_2026_v2")
    """
    RESULTADOS_DIR.mkdir(parents=True, exist_ok=True)

    # Migrar archivo legacy si existe
    _migrate_legacy_result(name)

    # Calcular siguiente version
    version = _next_version(name)
    full_name = f"{name}_v{version}"
    filepath = RESULTADOS_DIR / f"{full_name}.json"

    data = {
        "nombre": full_name,
        "base_name": name,
        "version": version,
        "nota": nota,
        "fecha_optimizacion": datetime.now().isoformat(),
        "weekly_schedule": weekly_schedule,
        "weekly_summary": weekly_summary,
        "daily_results": daily_results,
    }
    if pedido is not None:
        data["pedido"] = pedido
    if params is not None:
        # Solo guardar parametros serializables (excluir objetos complejos)
        safe_params = {
            "min_lot_size": params.get("min_lot_size"),
            "lot_step": params.get("lot_step"),
            "resource_capacity": params.get("resource_capacity"),
            "days": params.get("days"),
        }
        data["params"] = safe_params

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return full_name


def load_optimization_results(name: str) -> dict | None:
    """
    Carga resultados de optimizacion por nombre.

    Returns:
        dict con keys: nombre, fecha_optimizacion, weekly_schedule,
        weekly_summary, daily_results, pedido (opcional), params (opcional).
        None si no existe.
    """
    filepath = RESULTADOS_DIR / f"{name}.json"
    if not filepath.exists():
        return None
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def list_optimization_results() -> list:
    """
    Lista resultados de optimizacion guardados con metadatos basicos.

    Migra automaticamente archivos legacy (sin _vN) a formato versionado.

    Returns:
        lista de dicts con keys: nombre, base_name, version, nota,
        fecha_optimizacion, total_pares, status, tardiness.
        Ordenados por base_name desc, luego version desc.
    """
    RESULTADOS_DIR.mkdir(parents=True, exist_ok=True)

    # Migrar archivos legacy antes de listar
    for f in RESULTADOS_DIR.glob("*.json"):
        base, ver = parse_version_name(f.stem)
        if ver == 0:
            _migrate_legacy_result(f.stem)

    results = []
    for f in RESULTADOS_DIR.glob("*.json"):
        try:
            with open(f, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            summary = data.get("weekly_summary", {})
            base_name, version = parse_version_name(f.stem)
            # Si el JSON ya tiene version, usarla
            if data.get("version"):
                version = data["version"]
            if data.get("base_name"):
                base_name = data["base_name"]
            results.append({
                "nombre": f.stem,
                "base_name": base_name,
                "version": version,
                "nota": data.get("nota", ""),
                "fecha_optimizacion": data.get("fecha_optimizacion", ""),
                "total_pares": summary.get("total_pares", 0),
                "status": summary.get("status", ""),
                "tardiness": summary.get("total_tardiness", 0),
            })
        except (json.JSONDecodeError, KeyError):
            continue

    # Ordenar: base_name desc, version desc
    results.sort(key=lambda r: (r["base_name"], r["version"]), reverse=True)
    return results


def list_versions(base_name: str) -> list:
    """Lista versiones de una semana especifica, ordenadas por version desc."""
    all_results = list_optimization_results()
    return [r for r in all_results if r["base_name"] == base_name]


def delete_optimization_result(name: str) -> bool:
    """Elimina un resultado de optimizacion guardado."""
    filepath = RESULTADOS_DIR / f"{name}.json"
    if filepath.exists():
        filepath.unlink()
        return True
    return False


# ---------------------------------------------------------------------------
# Operarios: persistencia JSON
# ---------------------------------------------------------------------------

def save_operarios(operarios: list):
    """Guarda la lista completa de operarios a JSON."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(OPERARIOS_JSON, "w", encoding="utf-8") as f:
        json.dump(operarios, f, ensure_ascii=False, indent=2)


def load_operarios() -> list:
    """Carga la lista de operarios desde JSON. Migra tipos compuestos si los encuentra."""
    if not OPERARIOS_JSON.exists():
        return []
    with open(OPERARIOS_JSON, "r", encoding="utf-8") as f:
        operarios = json.load(f)
    if _migrate_operario_resources(operarios):
        save_operarios(operarios)
    return operarios


def _migrate_operario_resources(operarios: list) -> bool:
    """Migra tipos compuestos en recursos_habilitados. Retorna True si migro."""
    migrated = False
    for op in operarios:
        new_recursos = []
        for r in op.get("recursos_habilitados", []):
            mapped = _COMPOUND_TO_BASE.get(r, r)
            if mapped not in new_recursos:
                new_recursos.append(mapped)
            if mapped != r:
                migrated = True
        if migrated:
            op["recursos_habilitados"] = new_recursos
    return migrated


def _next_operario_id(operarios: list) -> str:
    """Genera el siguiente ID de operario (op_001, op_002, ...)."""
    max_num = 0
    for op in operarios:
        oid = op.get("id", "")
        if oid.startswith("op_"):
            try:
                max_num = max(max_num, int(oid[3:]))
            except ValueError:
                pass
    return f"op_{max_num + 1:03d}"


def save_operario(operario: dict) -> dict:
    """
    Agrega o actualiza un operario individual.
    Si tiene 'id' y ese id existe, actualiza. Si no, agrega nuevo.
    Retorna el operario guardado (con id asignado).
    """
    operarios = load_operarios()
    if operario.get("id"):
        for i, op in enumerate(operarios):
            if op["id"] == operario["id"]:
                operarios[i] = operario
                save_operarios(operarios)
                return operario
    operario["id"] = _next_operario_id(operarios)
    operarios.append(operario)
    save_operarios(operarios)
    return operario


def delete_operario(operario_id: str) -> bool:
    """Elimina un operario por ID. Retorna True si se elimino."""
    operarios = load_operarios()
    filtered = [op for op in operarios if op.get("id") != operario_id]
    if len(filtered) < len(operarios):
        save_operarios(filtered)
        return True
    return False


def compute_headcount_by_resource(operarios: list, day_name: str) -> dict:
    """
    Computa cuantos operarios activos pueden trabajar en cada tipo de recurso
    para un dia dado.

    Returns:
        dict {recurso: count} e.g. {"MESA": 12, "PLANA": 8, ...}
    """
    counts = {}
    for op in operarios:
        if not op.get("activo", True):
            continue
        if day_name not in op.get("dias_disponibles", []):
            continue
        for recurso in op.get("recursos_habilitados", []):
            counts[recurso] = counts.get(recurso, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Template Consolidado: genera un solo Excel con 3 hojas + instrucciones
# ---------------------------------------------------------------------------

def generate_template_consolidado() -> bytes:
    """
    Genera template Excel consolidado con 3 hojas de datos + instrucciones.

    Hojas:
      1. CATALOGO   - Operaciones por modelo (fracciones, recurso, rate, robots)
      2. PEDIDO     - Pedido semanal (modelos, volumenes, fabricas)
      3. OPERARIOS  - Personal (habilidades, robots, disponibilidad)
      4. INSTRUCCIONES - Guia detallada de como llenar cada hoja

    Returns:
        bytes del archivo Excel
    """
    from config_manager import get_physical_robots, get_fabricas

    wb = openpyxl.Workbook()

    robots_list = get_physical_robots()
    fabricas_list = get_fabricas()
    recursos_list = sorted(VALID_RESOURCES)
    dias_default = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab"]

    # Estilos adicionales
    example_font = Font(italic=True, color="999999")
    section_font = Font(bold=True, size=12, color="2E4057")
    note_font = Font(italic=True, color="666666", size=10)
    # ===== HOJA 1: CATALOGO =====
    ws_cat = wb.active
    ws_cat.title = "CATALOGO"

    # Headers fijos + columnas dinámicas de robots (ordenados)
    robots_sorted = sorted(robots_list)
    cat_headers_fixed = [
        "MODELO", "ALTERNATIVAS", "FRACCION", "OPERACION",
        "INPUT O PROCESO", "ETAPA", "RECURSO", "RATE",
    ]
    cat_widths_fixed = [14, 18, 12, 32, 20, 20, 16, 10]
    cat_headers = cat_headers_fixed + robots_sorted
    cat_widths = cat_widths_fixed + [5] * len(robots_sorted)
    robot_header_align = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, text_rotation=90)
    robot_header_font = Font(bold=True, color="FFFFFF", size=9)
    for col, (header, width) in enumerate(zip(cat_headers, cat_widths), 1):
        cell = ws_cat.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        ws_cat.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        if col <= len(cat_headers_fixed):
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        else:
            cell.font = robot_header_font
            cell.alignment = robot_header_align
    # INPUT O PROCESO: valores validos y sus colores (simbologia)
    proceso_values = ["PRELIMINARES", "ROBOT", "POST", "MAQUILA", "N/A PRELIMINAR"]
    proceso_colors = {
        "PRELIMINARES":    {"bg": "FFFF00", "fg": "000000"},  # amarillo
        "ROBOT":           {"bg": "92D050", "fg": "000000"},  # verde
        "POST":            {"bg": "FF66FF", "fg": "000000"},  # rosa
        "MAQUILA":         {"bg": "FF0000", "fg": "000000"},  # rojo
        "N/A PRELIMINAR":  {"bg": "FFFFFF", "fg": "000000"},  # blanco texto negro
    }

    # Escribir lista en hoja auxiliar oculta para referencia del dropdown
    ws_ref = wb.create_sheet("_REF")
    for i, val in enumerate(proceso_values, 1):
        cell = ws_ref.cell(row=i, column=1, value=val)
        colors = proceso_colors[val]
        cell.fill = PatternFill(start_color=colors["bg"], end_color=colors["bg"], fill_type="solid")
        cell.font = Font(bold=True, color=colors["fg"])
    ws_ref.sheet_state = "hidden"

    # Data validation: dropdown en columna E (INPUT O PROCESO)
    from openpyxl.worksheet.datavalidation import DataValidation
    proceso_dv = DataValidation(
        type="list",
        formula1="=_REF!$A$1:$A$5",
        allow_blank=True,
    )
    ws_cat.add_data_validation(proceso_dv)
    proceso_dv.add("E2:E200")

    # Formato condicional en columna E (INPUT O PROCESO)
    from openpyxl.formatting.rule import FormulaRule
    for proc_val, colors in proceso_colors.items():
        fill = PatternFill(start_color=colors["bg"], end_color=colors["bg"], fill_type="solid")
        font = Font(bold=True, color=colors["fg"])
        ws_cat.conditional_formatting.add(
            "E2:E200",
            FormulaRule(
                formula=[f'EXACT($E2,"{proc_val}")'],
                fill=fill,
                font=font,
            ),
        )

    # Simbologia visible (columna despues de robots + 1 separacion)
    simb_col = len(cat_headers) + 2
    cell = ws_cat.cell(row=1, column=simb_col, value="SIMBOLOGIA")
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center")
    ws_cat.column_dimensions[openpyxl.utils.get_column_letter(simb_col)].width = 20
    for i, val in enumerate(proceso_values, 2):
        colors = proceso_colors[val]
        cell = ws_cat.cell(row=i, column=simb_col, value=val)
        cell.fill = PatternFill(start_color=colors["bg"], end_color=colors["bg"], fill_type="solid")
        cell.font = Font(bold=True, color=colors["fg"])
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Ejemplos con nuevo formato de robots (OK en columnas correspondientes)
    # Tuplas: (modelo, alt, frac, op, input_proc, etapa, recurso, rate, {robots_ok})
    ok_font = Font(bold=True, color="2E8B57", size=10, italic=True)
    ok_align = Alignment(horizontal="center", vertical="center")
    cat_examples = [
        ("65413", "NE, GC", 1, "PEGAR FELPA",      "PRELIMINARES", "PRE-ROBOT",  "MESA",  120, set()),
        ("65413", "NE, GC", 2, "COSER PUNTERA",     "ROBOT",        "ROBOT",      "ROBOT",  80, {"3020-M4", "3020-M6"}),
        ("65413", "NE, GC", 3, "ADORNAR COSTADOS",  "POST",         "POST-LINEA", "PLANA",  90, set()),
        ("65413", "NE, GC", 4, "CERRAR TALON",      "PRELIMINARES", "PRE-ROBOT",  "MESA",  100, set()),
        ("65413", "NE, GC", 5, "COSER VISTA",       "POST",         "POST-LINEA", "PLANA",  95, set()),
    ]
    for row_offset, (modelo, alt, frac, op_name, inp, etapa, rec, rate, robots_ok) in enumerate(cat_examples, 2):
        fixed_vals = [modelo, alt, frac, op_name, inp, etapa, rec, rate]
        for col, val in enumerate(fixed_vals, 1):
            cell = ws_cat.cell(row=row_offset, column=col, value=val)
            cell.font = example_font
            cell.border = _THIN_BORDER
        # Columnas de robots: poner OK donde corresponda
        for i, robot in enumerate(robots_sorted):
            col = len(cat_headers_fixed) + 1 + i
            cell = ws_cat.cell(row=row_offset, column=col)
            cell.border = _THIN_BORDER
            if robot in robots_ok:
                cell.value = "OK"
                cell.font = ok_font
                cell.alignment = ok_align

    # ===== HOJA 2: OPERARIOS =====
    ws_ops = wb.create_sheet("OPERARIOS")

    ops_headers = [
        "NOMBRE", "FABRICA", "RECURSOS_HABILITADOS",
        "ROBOTS_HABILITADOS", "EFICIENCIA", "DIAS_DISPONIBLES",
    ]
    ops_widths = [18, 16, 35, 35, 12, 35]
    for col, (header, width) in enumerate(zip(ops_headers, ops_widths), 1):
        cell = ws_ops.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws_ops.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ops_examples = [
        ("ARACELI", "FABRICA 1", "MESA, PLANA", "", 1.0, "Lun, Mar, Mie, Jue, Vie"),
        ("DIANA", "FABRICA 1", "MESA, PLANA", "", 1.1, "Lun, Mar, Mie, Jue, Vie"),
        ("HUGO", "FABRICA 1", "ROBOT, MESA", "3020-M4, 6040-M5", 1.0, "Lun, Mar, Mie, Jue, Vie, Sab"),
        ("CARLOS", "FABRICA 2", "MESA, PLANA, POSTE", "", 0.9, "Lun, Mar, Mie, Jue, Vie"),
    ]
    for row_offset, (nombre, fab, recursos, robots, eff, dias) in enumerate(ops_examples, 2):
        ws_ops.cell(row=row_offset, column=1, value=nombre).font = example_font
        ws_ops.cell(row=row_offset, column=2, value=fab).font = example_font
        ws_ops.cell(row=row_offset, column=3, value=recursos).font = example_font
        ws_ops.cell(row=row_offset, column=4, value=robots).font = example_font
        ws_ops.cell(row=row_offset, column=5, value=eff).font = example_font
        ws_ops.cell(row=row_offset, column=6, value=dias).font = example_font
        for col in range(1, 7):
            ws_ops.cell(row=row_offset, column=col).border = _THIN_BORDER

    # ===== HOJA 3: CONFIGURACION =====
    from config_manager import load_config
    config = load_config()

    ws_cfg = wb.create_sheet("CONFIGURACION")
    ws_cfg.column_dimensions["A"].width = 35
    ws_cfg.column_dimensions["B"].width = 18
    ws_cfg.column_dimensions["C"].width = 18
    ws_cfg.column_dimensions["D"].width = 18
    ws_cfg.column_dimensions["E"].width = 18
    ws_cfg.column_dimensions["F"].width = 18
    ws_cfg.column_dimensions["G"].width = 18

    section_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    section_font_w = Font(bold=True, color="FFFFFF", size=11)
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)

    cfg_row = 1

    def _cfg_section(title, cols=7):
        nonlocal cfg_row
        ws_cfg.merge_cells(start_row=cfg_row, start_column=1,
                           end_row=cfg_row, end_column=cols)
        cell = ws_cfg.cell(row=cfg_row, column=1, value=title)
        cell.font = section_font_w
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        for c in range(1, cols + 1):
            ws_cfg.cell(row=cfg_row, column=c).border = _THIN_BORDER
        cfg_row += 1

    def _cfg_kv(label, value):
        nonlocal cfg_row
        c1 = ws_cfg.cell(row=cfg_row, column=1, value=label)
        c1.font = label_font
        c1.border = _THIN_BORDER
        c2 = ws_cfg.cell(row=cfg_row, column=2, value=value)
        c2.font = value_font
        c2.border = _THIN_BORDER
        cfg_row += 1

    # --- Horario Laboral ---
    schedule = config.get("schedule", {})
    _cfg_section("HORARIO LABORAL (Lunes a Viernes)")
    _cfg_kv("Hora entrada", schedule.get("entrada", "08:00"))
    _cfg_kv("Hora salida", schedule.get("salida", "19:00"))
    _cfg_kv("Comida inicio", schedule.get("comida_inicio", "13:10"))
    _cfg_kv("Comida fin", schedule.get("comida_fin", "13:50"))
    _cfg_kv("Duracion bloque (min)", schedule.get("bloque_min", 60))
    cfg_row += 1

    # --- Horario Fin de Semana ---
    schedule_fs = config.get("schedule_finsemana", {})
    _cfg_section("HORARIO FIN DE SEMANA (Sabado)")
    _cfg_kv("Hora entrada", schedule_fs.get("entrada", "08:00"))
    _cfg_kv("Hora salida", schedule_fs.get("salida", "13:00"))
    _cfg_kv("Comida inicio", schedule_fs.get("comida_inicio", ""))
    _cfg_kv("Comida fin", schedule_fs.get("comida_fin", ""))
    cfg_row += 1

    # --- Dias Laborales ---
    _cfg_section("DIAS LABORALES")
    day_headers = ["DIA", "MIN. REGULAR", "PLANTILLA REG.", "MIN. OVERTIME", "PLANTILLA OT", "ES SABADO"]
    for col, h in enumerate(day_headers, 1):
        cell = ws_cfg.cell(row=cfg_row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill(start_color="4A6B8A", end_color="4A6B8A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    cfg_row += 1

    for day in config.get("days", []):
        ws_cfg.cell(row=cfg_row, column=1, value=day["name"]).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=2, value=day["minutes"]).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=3, value=day["plantilla"]).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=4, value=day.get("minutes_ot", 0)).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=5, value=day.get("plantilla_ot", 0)).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=6, value="SI" if day.get("is_saturday") else "NO").border = _THIN_BORDER
        for c in range(1, 7):
            ws_cfg.cell(row=cfg_row, column=c).font = value_font
        cfg_row += 1
    cfg_row += 1

    # --- Capacidades por Recurso ---
    _cfg_section("CAPACIDADES POR RECURSO")
    cap_headers = ["RECURSO", "CAPACIDAD (personas)"]
    for col, h in enumerate(cap_headers, 1):
        cell = ws_cfg.cell(row=cfg_row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill(start_color="4A6B8A", end_color="4A6B8A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    cfg_row += 1

    for res, cap in sorted(config.get("resource_capacity", {}).items()):
        ws_cfg.cell(row=cfg_row, column=1, value=res).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=1).font = value_font
        ws_cfg.cell(row=cfg_row, column=2, value=cap).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=2).font = value_font
        cfg_row += 1
    cfg_row += 1

    # --- Robots Fisicos ---
    _cfg_section("ROBOTS FISICOS")
    for robot in robots_list:
        ws_cfg.cell(row=cfg_row, column=1, value=robot).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=1).font = value_font
        cfg_row += 1
    cfg_row += 1

    # --- Fabricas ---
    _cfg_section("FABRICAS")
    for fab in fabricas_list:
        ws_cfg.cell(row=cfg_row, column=1, value=fab).border = _THIN_BORDER
        ws_cfg.cell(row=cfg_row, column=1).font = value_font
        cfg_row += 1
    cfg_row += 1

    # --- Modelos por Fabrica ---
    modelo_fabrica = config.get("modelo_fabrica", {})
    all_cols = fabricas_list + ["SIN FABRICA ASIGNADA"]
    _cfg_section("MODELOS POR FABRICA", cols=len(all_cols))
    # Headers: una columna por fabrica + sin asignar
    sub_header_fill = PatternFill(start_color="4A6B8A", end_color="4A6B8A", fill_type="solid")
    warn_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    for col, fab in enumerate(all_cols, 1):
        cell = ws_cfg.cell(row=cfg_row, column=col, value=fab)
        cell.font = _HEADER_FONT
        cell.fill = warn_fill if fab == "SIN FABRICA ASIGNADA" else sub_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    cfg_row += 1
    # Filas: modelos asignados a cada fabrica
    fab_models = {fab: modelo_fabrica.get(fab, []) for fab in fabricas_list}
    # Detectar modelos sin fabrica (del catalogo)
    assigned = set()
    for models in fab_models.values():
        assigned.update(models)
    sin_fabrica = sorted(set(modelo_fabrica.get("SIN FABRICA ASIGNADA", [])))
    fab_models["SIN FABRICA ASIGNADA"] = sin_fabrica
    max_rows = max((len(v) for v in fab_models.values()), default=0)
    max_rows = max(max_rows, 5)  # minimo 5 filas vacias para que el usuario llene
    for r in range(max_rows):
        for col, fab in enumerate(all_cols, 1):
            models = fab_models.get(fab, [])
            val = models[r] if r < len(models) else None
            cell = ws_cfg.cell(row=cfg_row, column=col, value=val)
            cell.border = _THIN_BORDER
            cell.font = value_font
        cfg_row += 1
    cfg_row += 1

    # --- Parametros del Optimizador ---
    _cfg_section("PARAMETROS DEL OPTIMIZADOR")
    opt_params = config.get("optimizer_params", {})
    _cfg_kv("Lote minimo (pares)", opt_params.get("lote_minimo", 50))
    _cfg_kv("Lote preferido (pares)", opt_params.get("lote_preferido", 100))
    _cfg_kv("Factor eficiencia", opt_params.get("factor_eficiencia", 0.90))
    _cfg_kv("Factor contiguidad", opt_params.get("factor_contiguidad", 0.80))
    _cfg_kv("Timeout solver (seg)", opt_params.get("timeout_solver", 90))
    cfg_row += 1

    # --- Pesos de Priorizacion (ordenados de mayor a menor) ---
    _cfg_section("PESOS DE PRIORIZACION")
    weights = config.get("weights", {})
    weight_items = [
        ("Tardiness (atraso)", weights.get("tardiness", 100000)),
        ("Balance (sin dias muertos)", weights.get("balance", 30000)),
        ("Span (consolidar dias)", weights.get("span", 20000)),
        ("Changeover (cambio modelo)", weights.get("changeover", 10000)),
        ("Odd lot (lote impar)", weights.get("odd_lot", 5000)),
        ("Saturday (sabado)", weights.get("saturday", 500)),
        ("Uniformity (diario)", weights.get("uniformity", 100)),
        ("Overtime (por segundo)", weights.get("overtime", 10)),
        ("Early start (dias tempranos)", weights.get("early_start", 5)),
    ]
    weight_items.sort(key=lambda x: x[1], reverse=True)
    for label, val in weight_items:
        _cfg_kv(label, val)

    # ===== HOJA 4: INSTRUCCIONES =====
    ws_inst = wb.create_sheet("INSTRUCCIONES")
    ws_inst.column_dimensions["A"].width = 90

    instructions = [
        "TEMPLATE CONSOLIDADO - AGENTE DE PROGRAMACION PESPUNTE",
        "",
        "Este archivo tiene 3 hojas de datos + instrucciones.",
        "Las filas en gris son ejemplos - sobreescribalas con datos reales.",
        "Todos los campos son requeridos.",
        "",
        "=" * 70,
        "HOJA 1: CATALOGO  (operaciones de cada modelo)",
        "=" * 70,
        "",
        "Una fila por cada operacion (fraccion) de cada modelo.",
        "Un modelo tipico tiene 3-8 operaciones/fracciones.",
        "",
        "  MODELO          - Numero del modelo (solo digitos). Ej: 65413",
        "                    Se repite en cada fila de fraccion del mismo modelo.",
        "  ALTERNATIVAS    - Colores/variantes que usan este modelo.",
        "                    Separar por coma o punto y coma. Ej: NE, GC  o  NE; GC",
        "                    Se repite igual en cada fila del modelo.",
        "  FRACCION        - Numero secuencial de la operacion (1, 2, 3...)",
        "                    Indica el orden en que se ejecutan.",
        "  OPERACION       - Descripcion de lo que se hace. Ej: PEGAR FELPA",
        "  INPUT O PROCESO - Clasificacion general de la operacion (dropdown con colores).",
        "                    Valores:",
        "                    PRELIMINARES (amarillo) - Operaciones previas al robot",
        "                    ROBOT (verde) - Operaciones en robot/maquina",
        "                    POST (rosa) - Operaciones posteriores al robot",
        "                    MAQUILA (rojo) - Operaciones de maquila externa",
        "                    N/A PRELIMINAR (blanco) - No aplica como preliminar",
        "  ETAPA           - Etapa especifica de produccion. Texto libre.",
        "                    Ejemplos: PRE-ROBOT, ROBOT, POST-LINEA, POST-PLANA-LINEA",
        f"  RECURSO         - Tipo de recurso/estacion. DEBE SER UNO DE:",
        f"                    {', '.join(recursos_list)}",
        "  RATE            - Pares por hora que produce 1 persona en esta operacion.",
        "                    Numero > 0. Ej: 120 = 120 pares/hora.",
        "  COLUMNAS ROBOT  - Cada columna es un robot fisico (ordenados).",
        "                    Poner OK si la operacion puede ejecutarse en ese robot.",
        "                    Solo aplica cuando RECURSO = ROBOT.",
        "",
        "TIPS:",
        "  - Si un modelo usa robot en alguna fraccion, poner RECURSO=ROBOT",
        "    y marcar OK en las columnas de los robots elegibles.",
        "  - MESA = trabajo manual en mesa",
        "  - PLANA = maquina plana (costura)",
        "  - POSTE = maquina de poste",
        "  - MAQUILA = operacion externa",
        "",
        "=" * 70,
        "HOJA 2: OPERARIOS  (personal y habilidades)",
        "=" * 70,
        "",
        "Una fila por cada operario.",
        "",
        "  NOMBRE               - Nombre del operario. Ej: ARACELI",
        f"  FABRICA              - Fabrica donde trabaja. Ej: {', '.join(fabricas_list)}",
        f"  RECURSOS_HABILITADOS - Tipos de estacion que sabe operar, separados por coma.",
        f"                         Valores: {', '.join(recursos_list)}",
        f"  ROBOTS_HABILITADOS   - Robots en los que esta certificado, separados por coma.",
        f"                         Valores: {', '.join(robots_list)}",
        "  EFICIENCIA           - Velocidad relativa. 1.0 = estandar.",
        "                         >1.0 mas rapido, <1.0 mas lento. Rango: 0.5 a 1.5",
        f"  DIAS_DISPONIBLES     - Dias que trabaja, separados por coma.",
        f"                         Valores: {', '.join(dias_default)}",
        f"                         Default si vacio: Lun, Mar, Mie, Jue, Vie",
        "",
        "TIPS:",
        "  - Los OPERARIOS cambian poco. Solo actualizar al haber altas/bajas.",
        "  - Un operario puede tener multiples recursos (ej: MESA, PLANA).",
        "  - Solo poner robots si el operario esta CERTIFICADO para operarlos.",
        "",
        "=" * 70,
        "HOJA 3: CONFIGURACION  (parametros del sistema)",
        "=" * 70,
        "",
        "Contiene toda la configuracion del sistema organizada por secciones:",
        "",
        "  HORARIO LABORAL      - Entrada, salida, comida, duracion de bloques",
        "  DIAS LABORALES       - Minutos, plantilla y overtime por dia",
        "  CAPACIDADES          - Personas simultaneas por tipo de recurso",
        "  ROBOTS FISICOS       - Lista de maquinas disponibles",
        "  FABRICAS             - Lista de fabricas",
        "  PARAMETROS OPTIM.    - Lotes, eficiencia, timeouts",
        "  PESOS PRIORIZACION   - Prioridades del optimizador",
        "",
        "Edite directamente los valores en la columna B de cada parametro.",
        "Las tablas (dias, capacidades) se editan en sus filas correspondientes.",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="2E4057")
        elif "=====" in line:
            cell.font = Font(color="999999")
        elif line.startswith("HOJA") or line.startswith("CONFIGURACION"):
            cell.font = section_font
        elif line.startswith("TIPS"):
            cell.font = Font(bold=True, color="2E4057")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def import_consolidado(file_or_bytes) -> dict:
    """
    Importa datos desde el template consolidado (3 hojas).

    Lee cada hoja si existe y la importa con la logica correspondiente.

    Returns:
        dict con keys:
          - catalogo: (dict, errors) o None si hoja no existe
          - pedido: (list, errors) o None si hoja no existe
          - operarios: (list, errors) o None si hoja no existe
    """
    if hasattr(file_or_bytes, "read"):
        wb = openpyxl.load_workbook(file_or_bytes, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_or_bytes), data_only=True)

    result = {"catalogo": None, "pedido": None, "operarios": None}

    # --- CATALOGO ---
    if "CATALOGO" in wb.sheetnames:
        ws = wb["CATALOGO"]
        cat_data, cat_errors = _parse_catalogo_sheet(ws)
        if cat_data:
            save_catalog(cat_data)
        result["catalogo"] = (cat_data, cat_errors)

    # --- PEDIDO ---
    if "PEDIDO" in wb.sheetnames:
        ws = wb["PEDIDO"]
        ped_data, ped_errors, semana = _parse_pedido_sheet(ws)
        result["pedido"] = (ped_data, ped_errors, semana)

    # --- OPERARIOS ---
    if "OPERARIOS" in wb.sheetnames:
        ws = wb["OPERARIOS"]
        ops_data, ops_errors = _parse_operarios_sheet(ws)
        if ops_data:
            save_operarios(ops_data)
        result["operarios"] = (ops_data, ops_errors)

    return result


def _parse_catalogo_sheet(ws) -> tuple:
    """Parsea hoja CATALOGO del template consolidado.

    Columnas fijas: MODELO | ALTERNATIVAS | FRACCION | OPERACION | INPUT O PROCESO | ETAPA | RECURSO | RATE
    Columnas dinámicas (9+): nombres de robots en header, "OK" en celdas.
    También soporta formato legacy (ROBOT_1..8 con nombres de robot en celdas).
    """
    errors = []
    raw_ops = {}

    # Detectar formato de robots: leer headers desde col 9
    # Nuevo formato: header = nombre de robot, celda = "OK"
    # Legacy: header = "ROBOT_1", celda = nombre de robot
    robot_col_map = {}  # col -> robot_name (nuevo formato)
    is_new_format = False
    valid_robots = _get_valid_robots()
    for col in range(9, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip() in valid_robots:
            robot_col_map[col] = str(header).strip()
            is_new_format = True
        elif header and str(header).strip().startswith("ROBOT_"):
            break  # Legacy format
    # Si no se detectaron robots válidos en headers, intentar con todos los robots conocidos
    if not is_new_format:
        for col in range(9, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                normalized = _normalize_robot(str(header).strip())
                if normalized:
                    robot_col_map[col] = normalized
                    is_new_format = True

    for row in range(2, ws.max_row + 1):  # Skip header row only
        modelo = ws.cell(row=row, column=1).value
        alternativas_raw = ws.cell(row=row, column=2).value
        fraccion = ws.cell(row=row, column=3).value
        operacion = ws.cell(row=row, column=4).value
        input_proceso = ws.cell(row=row, column=5).value
        etapa = ws.cell(row=row, column=6).value
        recurso = ws.cell(row=row, column=7).value
        rate = ws.cell(row=row, column=8).value

        if not modelo or not fraccion:
            continue

        modelo_str = str(modelo).strip()
        model_match = re.match(r"^(\d+)", modelo_str)
        if not model_match:
            errors.append(f"Fila {row}: MODELO '{modelo_str}' no inicia con numero")
            continue

        model_num = model_match.group(1)

        # Parsear alternativas: separadas por coma o punto y coma
        alternativas = []
        if alternativas_raw:
            for a in re.split(r"[,;]", str(alternativas_raw)):
                a_clean = a.strip().upper()
                if a_clean:
                    alternativas.append(a_clean)

        recurso_str = str(recurso).strip().upper() if recurso else "GENERAL"
        recurso_str = _COMPOUND_TO_BASE.get(recurso_str, recurso_str)
        if recurso_str == "GENERAL":
            operacion_name = str(operacion).strip() if operacion else f"OP-{fraccion}"
            errors.append(f"⚠ Fila {row}: '{model_num} / {operacion_name}' sin RECURSO asignado (usando GENERAL)")
        elif recurso_str not in VALID_RESOURCES:
            errors.append(f"Fila {row}: RECURSO '{recurso_str}' no valido")
            continue

        try:
            rate_val = float(rate)
            if rate_val <= 0:
                raise ValueError()
        except (TypeError, ValueError):
            errors.append(f"Fila {row}: RATE '{rate}' debe ser numerico > 0")
            continue

        # Robots: nuevo formato (header=robot, celda=OK) o legacy (celda=nombre)
        robots = []
        if is_new_format:
            for col, robot_name in robot_col_map.items():
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().upper() in ("OK", "SI", "X", "1"):
                    if robot_name not in robots:
                        robots.append(robot_name)
        else:
            for col in range(9, 17):
                val = ws.cell(row=row, column=col).value
                robot = _normalize_robot(val)
                if robot and robot not in robots:
                    robots.append(robot)

        if robots and recurso_str != "ROBOT":
            recurso_str = "ROBOT"

        sec_per_pair = round(3600.0 / rate_val)

        # Construir codigo_full: "65413 NE/GC" o solo "65413"
        if alternativas:
            codigo_full = f"{model_num} {'/'.join(alternativas)}"
        else:
            codigo_full = model_num

        if model_num not in raw_ops:
            raw_ops[model_num] = {
                "codigo_full": codigo_full,
                "alternativas": alternativas,
                "ops": [],
            }
        elif alternativas and not raw_ops[model_num].get("alternativas"):
            raw_ops[model_num]["alternativas"] = alternativas
            raw_ops[model_num]["codigo_full"] = codigo_full

        raw_ops[model_num]["ops"].append({
            "fraccion": int(fraccion),
            "operacion": str(operacion).strip() if operacion else f"OP-{fraccion}",
            "input_o_proceso": str(input_proceso).strip() if input_proceso else "",
            "etapa": str(etapa).strip() if etapa else "",
            "recurso": recurso_str,
            "recurso_raw": recurso_str,
            "robots": robots,
            "rate": round(rate_val, 2),
            "sec_per_pair": sec_per_pair,
        })

    # Construir catalogo
    catalog = {}
    for model_num, data in raw_ops.items():
        ops = sorted(data["ops"], key=lambda x: x["fraccion"])
        seen = set()
        unique_ops = []
        for op in ops:
            if op["fraccion"] not in seen:
                seen.add(op["fraccion"])
                unique_ops.append(op)

        total_sec = sum(op["sec_per_pair"] for op in unique_ops)
        resource_summary = {}
        for op in unique_ops:
            r = op["recurso"]
            resource_summary[r] = resource_summary.get(r, 0) + 1

        robot_ops = sum(1 for op in unique_ops if op.get("robots"))
        all_robots = set()
        for op in unique_ops:
            for r in op.get("robots", []):
                all_robots.add(r)

        catalog[model_num] = {
            "codigo_full": data["codigo_full"],
            "alternativas": data.get("alternativas", []),
            "clave_material": "",
            "fabrica": "",
            "operations": unique_ops,
            "total_sec_per_pair": total_sec,
            "num_ops": len(unique_ops),
            "resource_summary": resource_summary,
            "robot_ops": robot_ops,
            "robots_used": sorted(all_robots),
        }

    return catalog, errors


def _parse_pedido_sheet(ws) -> tuple:
    """Parsea hoja PEDIDO del template consolidado.

    Layout: Fila 1=SEMANA, Fila 2=vacia, Fila 3=headers, Fila 4=req/opt, Fila 5+=datos
    Returns: (pedido_list, errors, semana_str)
    """
    errors = []
    pedido = []

    # Leer identificador de semana (celda B1)
    semana_raw = ws.cell(row=1, column=2).value
    semana = str(semana_raw).strip() if semana_raw else ""

    for row in range(5, ws.max_row + 1):
        modelo = ws.cell(row=row, column=1).value
        color = ws.cell(row=row, column=2).value
        clave = ws.cell(row=row, column=3).value
        fabrica = ws.cell(row=row, column=4).value
        volumen = ws.cell(row=row, column=5).value

        if not modelo:
            continue

        modelo_str = str(modelo).strip()
        if not modelo_str:
            continue

        color_str = str(color).strip() if color else ""
        clave_str = str(clave).strip() if clave else ""
        fabrica_str = str(fabrica).strip() if fabrica else "FABRICA 1"

        try:
            vol = int(float(volumen))
            if vol <= 0:
                raise ValueError()
        except (TypeError, ValueError):
            errors.append(f"Fila {row}: VOLUMEN '{volumen}' debe ser entero > 0")
            continue

        pedido.append({
            "modelo": modelo_str,
            "color": color_str,
            "clave_material": clave_str,
            "fabrica": fabrica_str,
            "volumen": vol,
        })

    return pedido, errors, semana


def _parse_operarios_sheet(ws) -> tuple:
    """Parsea hoja OPERARIOS del template consolidado."""
    errors = []
    operarios = []
    default_dias = ["Lun", "Mar", "Mie", "Jue", "Vie"]
    op_counter = 0

    for row in range(2, ws.max_row + 1):  # Skip header row only
        nombre = ws.cell(row=row, column=1).value
        fabrica = ws.cell(row=row, column=2).value
        recursos_raw = ws.cell(row=row, column=3).value
        robots_raw = ws.cell(row=row, column=4).value
        eficiencia = ws.cell(row=row, column=5).value
        dias_raw = ws.cell(row=row, column=6).value

        if not nombre:
            continue

        nombre_str = str(nombre).strip().upper()
        if not nombre_str:
            continue

        fabrica_str = str(fabrica).strip() if fabrica else "FABRICA 1"

        # Parsear recursos (separados por coma, migrar legacy compuestos)
        recursos = []
        if recursos_raw:
            for r in str(recursos_raw).split(","):
                r_clean = r.strip().upper()
                r_clean = _COMPOUND_TO_BASE.get(r_clean, r_clean)
                if r_clean in VALID_RESOURCES and r_clean not in recursos:
                    recursos.append(r_clean)
                elif r_clean and r_clean not in VALID_RESOURCES:
                    errors.append(f"Fila {row}: recurso '{r_clean}' no valido para {nombre_str}")
        if not recursos:
            errors.append(f"Fila {row}: {nombre_str} debe tener al menos 1 recurso habilitado")
            continue

        # Parsear robots (separados por coma)
        robots_hab = []
        if robots_raw:
            for r in str(robots_raw).split(","):
                robot = _normalize_robot(r.strip())
                if robot and robot not in robots_hab:
                    robots_hab.append(robot)

        # Eficiencia
        try:
            eff = float(eficiencia) if eficiencia else 1.0
            eff = max(0.5, min(1.5, eff))
        except (TypeError, ValueError):
            eff = 1.0

        # Dias disponibles
        dias = []
        if dias_raw:
            for d in str(dias_raw).split(","):
                d_clean = d.strip()
                # Normalizar: primera letra mayuscula
                if d_clean:
                    d_clean = d_clean[0].upper() + d_clean[1:].lower()
                    dias.append(d_clean)
        if not dias:
            dias = default_dias

        op_counter += 1
        operarios.append({
            "id": f"op_{op_counter:03d}",
            "nombre": nombre_str,
            "fabrica": fabrica_str,
            "recursos_habilitados": recursos,
            "robots_habilitados": robots_hab,
            "eficiencia": round(eff, 2),
            "dias_disponibles": dias,
            "activo": True,
        })

    return operarios, errors


# ---------------------------------------------------------------------------
# Restricciones: persistencia JSON
# ---------------------------------------------------------------------------

def save_restricciones(restricciones: list):
    """Guarda la lista completa de restricciones a JSON."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(RESTRICCIONES_JSON, "w", encoding="utf-8") as f:
        json.dump(restricciones, f, ensure_ascii=False, indent=2)


def load_restricciones() -> list:
    """Carga restricciones desde JSON. Retorna lista vacia si no existe."""
    if not RESTRICCIONES_JSON.exists():
        return []
    with open(RESTRICCIONES_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def _next_restriccion_id(restricciones: list) -> str:
    """Genera el siguiente ID de restriccion (r_001, r_002, ...)."""
    max_num = 0
    for r in restricciones:
        rid = r.get("id", "")
        if rid.startswith("r_"):
            try:
                max_num = max(max_num, int(rid[2:]))
            except ValueError:
                pass
    return f"r_{max_num + 1:03d}"


def save_restriccion(restriccion: dict) -> dict:
    """Agrega o actualiza una restriccion individual."""
    restricciones = load_restricciones()
    if restriccion.get("id"):
        for i, r in enumerate(restricciones):
            if r["id"] == restriccion["id"]:
                restricciones[i] = restriccion
                save_restricciones(restricciones)
                return restriccion
    restriccion["id"] = _next_restriccion_id(restricciones)
    restricciones.append(restriccion)
    save_restricciones(restricciones)
    return restriccion


def delete_restriccion(restriccion_id: str) -> bool:
    """Elimina una restriccion por ID."""
    restricciones = load_restricciones()
    filtered = [r for r in restricciones if r.get("id") != restriccion_id]
    if len(filtered) < len(restricciones):
        save_restricciones(filtered)
        return True
    return False


# ---------------------------------------------------------------------------
# Avance de produccion: persistencia JSON
# ---------------------------------------------------------------------------

def save_avance(avance: dict):
    """Guarda avance de produccion a JSON."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(AVANCE_JSON, "w", encoding="utf-8") as f:
        json.dump(avance, f, ensure_ascii=False, indent=2)


def load_avance() -> dict:
    """Carga avance desde JSON. Retorna dict vacio si no existe."""
    if not AVANCE_JSON.exists():
        return {}
    with open(AVANCE_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def update_avance_model(semana: str, modelo: str, dia: str, pares: int):
    """Actualiza avance de un modelo en un dia especifico."""
    avance = load_avance()
    if avance.get("semana") != semana:
        avance = {"semana": semana, "updated_at": "", "modelos": {}}
    avance["updated_at"] = datetime.now().isoformat()
    if modelo not in avance["modelos"]:
        avance["modelos"][modelo] = {}
    avance["modelos"][modelo][dia] = pares
    save_avance(avance)
