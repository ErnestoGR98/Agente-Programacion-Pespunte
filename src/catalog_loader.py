"""
catalog_loader.py - Carga mejorada del catalogo de fracciones (Iteracion 2).

Mejoras sobre el loader original:
  - Normaliza recursos a tipos canonicos (MESA, ROBOT, PLANA, etc.)
  - Infiere recurso de la etapa cuando el campo esta vacio
  - Parsea asignaciones de robots/maquinas de columnas 8-15
  - Retorna operaciones con recurso limpio y lista de robots para el scheduler
"""

import re
import openpyxl
from rules import RESOURCE_ALIASES


# Mapeo de recursos literales a tipos canonicos
RESOURCE_MAP = {
    "MESA": "MESA",
    "MESA-LINEA": "MESA-LINEA",
    "PLANA": "PLANA",
    "PLANA-LINEA": "PLANA-LINEA",
    "PLANA - LINEA": "PLANA-LINEA",
    "ROBOT": "ROBOT",
    "POSTE-LINEA": "POSTE-LINEA",
    "POST-LINEA": "POSTE-LINEA",
    "ZIGZAG-LINEA": "PLANA-LINEA",
    "ZIGZAG": "PLANA-LINEA",
    "PLANCHA": "MESA",
    "MEDIO POSTE": "POSTE-LINEA",
    "PLANA/POSTE MEDIO-LINEA": "PLANA-LINEA",
    "CHACHE 048": "ROBOT",
    "2A-3020-M1": "ROBOT",
}

# Nombres canonicos de robots fisicos
VALID_ROBOTS = {
    "2A-3020-M1", "2A-3020-M2", "3020-M4", "3020-M6",
    "6040-M4", "6040-M5", "CHACHE 048", "CHACHE 049",
}

# Alias de nombres de robots (variantes encontradas en el catalogo)
ROBOT_ALIASES = {
    "3020 M-4": "3020-M4",
    "6040-M5 (PARCIAL)": "6040-M5",
}

# Inferir recurso desde la etapa (columna 7) cuando el recurso esta vacio
ETAPA_TO_RESOURCE = {
    "MESA": "MESA",
    "ROBOT": "ROBOT",
    "PRE -ROBOT": "MESA",
    "PRE-ROBOT": "MESA",
    "PRE-ROBOT ": "MESA",
    "POST-PLANA-LINEA": "PLANA-LINEA",
    "POST-LINEA": "POSTE-LINEA",
    "ZIGZAG-LINEA": "PLANA-LINEA",
    "N/A": "MESA",
}


def normalize_resource(resource_val, etapa_val=None):
    """Normaliza un valor de recurso a un tipo canonico."""
    if resource_val:
        res = str(resource_val).strip().upper()
        # Buscar en mapeo directo
        if res in RESOURCE_MAP:
            return RESOURCE_MAP[res]
        # Buscar en alias de rules.py
        if res in RESOURCE_ALIASES:
            return RESOURCE_ALIASES[res]
        # Detectar formulas residuales (=+G12)
        if res.startswith("="):
            return _infer_from_etapa(etapa_val)
        # Buscar por contenido parcial
        for key in ["MESA-LINEA", "PLANA-LINEA", "POSTE-LINEA"]:
            if key in res:
                return key
        if "ROBOT" in res:
            return "ROBOT"
        if "PLANA" in res:
            return "PLANA"
        if "MESA" in res:
            return "MESA"
        if "CONFORMADORA" in res:
            return "MESA"
        if "PESPUNTADOR" in res or "JARETA" in res:
            return "PLANA"
        return "GENERAL"

    return _infer_from_etapa(etapa_val)


def _infer_from_etapa(etapa_val):
    """Infiere recurso a partir de la etapa."""
    if not etapa_val:
        return "GENERAL"
    etapa = str(etapa_val).strip().upper()
    if etapa in ETAPA_TO_RESOURCE:
        return ETAPA_TO_RESOURCE[etapa]
    if "ROBOT" in etapa:
        return "ROBOT"
    if "PLANA" in etapa:
        return "PLANA"
    if "MESA" in etapa:
        return "MESA"
    if "POSTE" in etapa:
        return "POSTE-LINEA"
    return "GENERAL"


def normalize_robot_name(val):
    """Normaliza un nombre de robot/maquina. Retorna None si es invalido."""
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Aplicar alias
    if s in ROBOT_ALIASES:
        s = ROBOT_ALIASES[s]
    # Verificar si es un robot valido
    if s in VALID_ROBOTS:
        return s
    # Intentar match parcial (sin espacios, case insensitive)
    s_upper = s.upper().replace(" ", "")
    for robot in VALID_ROBOTS:
        if robot.upper().replace(" ", "") == s_upper:
            return robot
    return None


def _parse_robots_from_row(ws, row):
    """Extrae y normaliza los robots de columnas 8-15 de una fila."""
    robots = []
    for col in range(8, 16):
        val = ws.cell(row=row, column=col).value
        robot = normalize_robot_name(val)
        if robot and robot not in robots:
            robots.append(robot)
    return robots


def load_catalog_v2(filepath: str) -> dict:
    """
    Parsea el catalogo de fracciones con recursos normalizados.

    Returns:
        dict {modelo_num: {
            codigo_full, operations, total_sec_per_pair, num_ops,
            resource_summary: {tipo: count}
        }}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["PLANTILLA MOD."]

    raw_ops = {}
    current_model_num = None

    for row in range(11, ws.max_row + 1):
        modelo_val = ws.cell(row=row, column=1).value
        fraccion = ws.cell(row=row, column=5).value
        operacion = ws.cell(row=row, column=6).value
        etapa = ws.cell(row=row, column=7).value
        recurso_raw = ws.cell(row=row, column=16).value
        tiempo_std = ws.cell(row=row, column=17).value
        rate = ws.cell(row=row, column=18).value

        # Detectar nueva seccion de modelo
        if modelo_val:
            modelo_str = str(modelo_val).strip()
            match = re.match(r"^(\d+)", modelo_str)
            if match:
                current_model_num = match.group(1)
                if current_model_num not in raw_ops:
                    raw_ops[current_model_num] = {
                        "codigo_full": modelo_str,
                        "ops": [],
                    }

        # Agregar operacion si tiene fraccion y rate
        if current_model_num and fraccion and rate:
            rate_val = float(rate)
            if rate_val <= 0:
                continue

            if tiempo_std:
                sec_per_pair = float(tiempo_std)
            else:
                sec_per_pair = 3600.0 / rate_val

            recurso = normalize_resource(recurso_raw, etapa)

            # Parsear robots asignados de columnas 8-15
            robots = _parse_robots_from_row(ws, row)

            raw_ops[current_model_num]["ops"].append({
                "fraccion": int(fraccion),
                "operacion": str(operacion).strip() if operacion else f"OP {etapa or 'AUTO'}",
                "etapa": str(etapa).strip() if etapa else "",
                "recurso": recurso,
                "recurso_raw": str(recurso_raw).strip() if recurso_raw else "",
                "robots": robots,
                "rate": round(rate_val, 2),
                "sec_per_pair": round(sec_per_pair),
            })

    # Deduplicar por fraccion y calcular totales
    catalog = {}
    for model_num, data in raw_ops.items():
        seen = set()
        unique_ops = []
        for op in data["ops"]:
            key = op["fraccion"]
            if key not in seen:
                seen.add(key)
                unique_ops.append(op)

        unique_ops.sort(key=lambda x: x["fraccion"])
        total_sec = sum(op["sec_per_pair"] for op in unique_ops)

        # Resumen de recursos por tipo
        resource_summary = {}
        for op in unique_ops:
            r = op["recurso"]
            resource_summary[r] = resource_summary.get(r, 0) + 1

        # Contar operaciones con robots asignados
        robot_ops = sum(1 for op in unique_ops if op.get("robots"))
        all_robots = set()
        for op in unique_ops:
            for r in op.get("robots", []):
                all_robots.add(r)

        catalog[model_num] = {
            "codigo_full": data["codigo_full"],
            "operations": unique_ops,
            "total_sec_per_pair": total_sec,
            "num_ops": len(unique_ops),
            "resource_summary": resource_summary,
            "robot_ops": robot_ops,
            "robots_used": sorted(all_robots),
        }

    return catalog
