"""
template_generator.py - Genera template Excel para importar catalogo y pedido.

Crea un archivo .xlsx limpio con 3 hojas:
  - INSTRUCCIONES: explicacion de columnas y valores validos
  - CATALOGO: headers + ejemplos para importar catalogo de operaciones
  - PEDIDO: headers + ejemplos para importar pedido semanal
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Estilos
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_EXAMPLE_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=12)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _style_example(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _EXAMPLE_FILL
        cell.border = _THIN_BORDER


def _build_instrucciones(wb):
    ws = wb.create_sheet("INSTRUCCIONES", 0)
    ws.sheet_properties.tabColor = "10B981"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    rows = [
        ("TEMPLATE DE IMPORTACION - PESPUNTE AGENT", ""),
        ("", ""),
        ("Este archivo tiene 2 hojas de datos:", ""),
        ("  - CATALOGO", "Operaciones de costura por modelo (una fila por operacion)"),
        ("  - PEDIDO", "Modelos y volumenes a producir en la semana"),
        ("", ""),
        ("=" * 50, ""),
        ("HOJA CATALOGO - Columnas", ""),
        ("=" * 50, ""),
        ("MODELO", "Numero del modelo (ej: 65413). Debe iniciar con digito."),
        ("ALTERNATIVAS", "Colores o variantes separados por coma (ej: NE,GC). Opcional."),
        ("FRACCION", "Numero secuencial de la operacion (1, 2, 3...). Requerido."),
        ("OPERACION", "Nombre de la operacion (ej: PEGAR FELPA). Requerido."),
        ("INPUT O PROCESO", "Tipo de proceso: PRELIMINARES, ROBOT, POST, MAQUILA, N/A PRELIMINAR"),
        ("ETAPA", "Etapa del proceso: PRE-ROBOT, ROBOT, POST-LINEA, etc. Opcional."),
        ("RECURSO", "Tipo de recurso: MESA, ROBOT, PLANA, POSTE, MAQUILA. Requerido."),
        ("RATE", "Pares por hora que produce un operario. Numerico > 0. Requerido."),
        ("Columnas de robots", "Una columna por robot fisico. Marcar con 'OK' si la operacion puede usar ese robot."),
        ("", ""),
        ("=" * 50, ""),
        ("HOJA PEDIDO - Layout", ""),
        ("=" * 50, ""),
        ("Fila 1", "Celda A1='SEMANA', celda B1=nombre de la semana (ej: sem_8_2026)"),
        ("Fila 2", "Vacia"),
        ("Fila 3", "Headers: MODELO, COLOR, CLAVE_MATERIAL, FABRICA, VOLUMEN"),
        ("Fila 4", "Indicadores: requerido, opcional, opcional, opcional, requerido"),
        ("Fila 5+", "Datos del pedido"),
        ("", ""),
        ("MODELO", "Numero del modelo (debe existir en el catalogo). Requerido."),
        ("COLOR", "Color o variante. Opcional."),
        ("CLAVE_MATERIAL", "Clave de material. Opcional."),
        ("FABRICA", "Fabrica asignada (ej: FABRICA 1). Default: FABRICA 1."),
        ("VOLUMEN", "Pares a producir. Entero > 0. Requerido."),
        ("", ""),
        ("=" * 50, ""),
        ("NOTAS", ""),
        ("=" * 50, ""),
        ("", "Las filas de ejemplo (fondo gris) deben eliminarse antes de importar."),
        ("", "El catalogo se puede importar independiente del pedido."),
        ("", "Los robots se configuran en el sistema. Las columnas del template"),
        ("", "reflejan los robots activos al momento de descargar."),
    ]

    for i, (a, b) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        if a and a.startswith("="):
            ws.cell(row=i, column=1).font = Font(color="999999")
        elif i == 1:
            ws.cell(row=i, column=1).font = Font(bold=True, size=14)
        elif a in ("MODELO", "ALTERNATIVAS", "FRACCION", "OPERACION",
                    "INPUT O PROCESO", "ETAPA", "RECURSO", "RATE",
                    "Columnas de robots", "COLOR", "CLAVE_MATERIAL",
                    "FABRICA", "VOLUMEN", "Fila 1", "Fila 2", "Fila 3",
                    "Fila 4", "Fila 5+", "NOTAS"):
            ws.cell(row=i, column=1).font = Font(bold=True)


def _build_catalogo(wb, robot_names: list):
    ws = wb.create_sheet("CATALOGO")
    ws.sheet_properties.tabColor = "3B82F6"

    # Headers fijos
    fixed_headers = [
        "MODELO", "ALTERNATIVAS", "FRACCION", "OPERACION",
        "INPUT O PROCESO", "ETAPA", "RECURSO", "RATE",
    ]
    total_cols = len(fixed_headers) + len(robot_names)

    for c, h in enumerate(fixed_headers, 1):
        ws.cell(row=1, column=c, value=h)
    for c, r in enumerate(robot_names, len(fixed_headers) + 1):
        ws.cell(row=1, column=c, value=r)

    _style_header(ws, 1, total_cols)

    # Anchos
    widths = [12, 15, 10, 25, 18, 15, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for c in range(len(fixed_headers) + 1, total_cols + 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 14

    # Filas de ejemplo
    examples = [
        ["65413", "NE,GC", 1, "COSTURA CHINELA", "PRELIMINARES", "PRE-ROBOT", "MESA", 120],
        ["65413", "NE,GC", 2, "COSTURA LATERAL", "ROBOT", "ROBOT", "ROBOT", 100],
        ["65413", "NE,GC", 3, "REMATE FINAL", "POST", "POST-LINEA", "PLANA", 90],
    ]

    # Marcar robots en ejemplo (filas ROBOT tienen OK en primer robot)
    for row_idx, data in enumerate(examples, 2):
        for c, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=c, value=val)
        # Si es operacion ROBOT, marcar primer robot como OK
        if data[6] == "ROBOT" and robot_names:
            ws.cell(row=row_idx, column=len(fixed_headers) + 1, value="OK")
        _style_example(ws, row_idx, total_cols)


def _build_pedido(wb):
    ws = wb.create_sheet("PEDIDO")
    ws.sheet_properties.tabColor = "F59E0B"

    # Row 1: SEMANA
    ws.cell(row=1, column=1, value="SEMANA")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2, value="sem_XX_2026")

    # Row 3: Headers
    headers = ["MODELO", "COLOR", "CLAVE_MATERIAL", "FABRICA", "VOLUMEN"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, len(headers))

    # Row 4: requerido/opcional
    markers = ["requerido", "opcional", "opcional", "opcional", "requerido"]
    for c, m in enumerate(markers, 1):
        cell = ws.cell(row=4, column=c, value=m)
        cell.font = Font(italic=True, color="999999", size=9)
        cell.alignment = Alignment(horizontal="center")

    # Filas de ejemplo
    examples = [
        ["65413", "NEGRO", "MAT-001", "FABRICA 1", 500],
        ["77525", "CAFE", "MAT-002", "FABRICA 2", 300],
        ["88190", "", "", "FABRICA 1", 200],
    ]
    for row_idx, data in enumerate(examples, 5):
        for c, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=c, value=val)
        _style_example(ws, row_idx, len(headers))

    # Anchos
    widths = [12, 12, 16, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def generate_template(robot_names: list[str] | None = None) -> BytesIO:
    """Genera template Excel y retorna como BytesIO buffer.

    Args:
        robot_names: lista de nombres de robots activos (para columnas del CATALOGO).
                     Si es None, usa lista default.
    """
    if robot_names is None:
        robot_names = [
            "2A-3020-M1", "2A-3020-M2", "3020-M4", "3020-M6",
            "6040-M4", "6040-M5", "CHACHE 048", "CHACHE 049",
        ]

    wb = Workbook()
    # Eliminar hoja default
    wb.remove(wb.active)

    _build_instrucciones(wb)
    _build_catalogo(wb, robot_names)
    _build_pedido(wb)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
