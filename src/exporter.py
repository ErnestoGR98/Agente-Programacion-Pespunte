"""
exporter.py - Exportacion de la programacion a Excel con formato sabana limpio.

Genera un archivo Excel con hojas:
  - Programacion: pares por modelo por dia (la sabana principal)
  - Balance_Diario: headcount necesario vs disponible por dia
  - Detalle_Modelos: informacion por modelo (volumen, completado, etc.)
  - PROG_LUNES, PROG_MARTES, ...: programa hora por hora (Iter 2)
"""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# Colores por tipo de recurso para resaltado visual
RESOURCE_COLORS = {
    "MESA": "D4E6F1",        # azul claro
    "ROBOT": "F9E79F",       # amarillo
    "PLANA": "A9DFBF",       # verde claro
    "POSTE": "F5CBA7",       # naranja claro
    "MAQUILA": "E8DAEF",     # violeta claro
    "GENERAL": "D5DBDB",     # gris
}


def export_schedule(schedule: list, summary: dict, output_path: str,
                    daily_results: dict = None):
    """Exporta la programacion optimizada a Excel."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_programacion(schedule, summary, writer)
        _write_balance(summary, writer)
        _write_detalle(summary, writer)

        # Hojas de programa diario (Iteracion 2)
        if daily_results:
            for day_name, day_data in daily_results.items():
                if day_data["schedule"]:
                    _write_daily_program(day_name, day_data, writer)
                # Hoja de cascada por operario (Iteracion 3)
                if day_data.get("operator_timelines"):
                    _write_cascade_sheet(day_name, day_data, writer)

    print(f"  Archivo exportado: {output_path}")


def _write_programacion(schedule: list, summary: dict, writer: pd.ExcelWriter):
    """Hoja principal: pares por modelo por dia (formato pivote)."""
    if not schedule:
        pd.DataFrame().to_excel(writer, sheet_name="Programacion", index=False)
        return

    df = pd.DataFrame(schedule)

    # Crear tabla pivote: filas=modelo, columnas=dias, valores=pares
    days_order = [ds["dia"] for ds in summary["days"]]
    available_days = [d for d in days_order if d in df["Dia"].values]

    pivot = df.pivot_table(
        index=["Fabrica", "Modelo", "Suela"],
        columns="Dia",
        values="Pares",
        aggfunc="sum",
        fill_value=0,
    )

    # Reordenar columnas segun orden de dias
    pivot = pivot.reindex(columns=[d for d in days_order if d in pivot.columns], fill_value=0)

    # Agregar columna de total
    pivot["TOTAL"] = pivot.sum(axis=1)

    # Agregar HC por modelo (promedio)
    hc_pivot = df.pivot_table(
        index=["Fabrica", "Modelo", "Suela"],
        columns="Dia",
        values="HC_Necesario",
        aggfunc="sum",
        fill_value=0,
    )
    hc_pivot = hc_pivot.reindex(columns=[d for d in days_order if d in hc_pivot.columns], fill_value=0)

    # Reset index para escribir
    pivot = pivot.reset_index()

    pivot.to_excel(writer, sheet_name="Programacion", index=False)

    # Marcar en rojo celdas con lotes no multiplo de 100
    ws = writer.sheets["Programacion"]
    red_font = Font(color="FF0000", bold=True)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Columnas de dias empiezan en col 4 (D) hasta la penultima (antes de TOTAL)
    num_day_cols = len([c for c in pivot.columns if c not in ["Fabrica", "Modelo", "Suela", "TOTAL"]])
    day_col_start = 4  # columna D (1-indexed)

    for row_idx in range(2, len(pivot) + 2):  # filas de datos (row 2 en adelante, row 1 es header)
        for col_offset in range(num_day_cols):
            cell = ws.cell(row=row_idx, column=day_col_start + col_offset)
            if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                if int(cell.value) % 100 != 0:
                    cell.font = red_font
                    cell.fill = red_fill

    # Escribir tabla de HC debajo
    start_row = len(pivot) + 3
    hc_pivot = hc_pivot.reset_index()
    hc_pivot.to_excel(
        writer, sheet_name="Programacion", index=False,
        startrow=start_row, header=True,
    )


def _write_balance(summary: dict, writer: pd.ExcelWriter):
    """Hoja de balance diario: headcount necesario vs disponible."""
    rows = []
    for ds in summary["days"]:
        tipo = "EXTRA" if ds["is_saturday"] else "Normal"
        rows.append({
            "Dia": ds["dia"],
            "Tipo": tipo,
            "Pares": ds["pares"],
            "HC_Necesario": ds["hc_necesario"],
            "HC_Disponible": ds["hc_disponible"],
            "Diferencia": ds["diferencia"],
            "Utilizacion_%": ds["utilizacion_pct"],
            "Overtime_hrs": ds.get("overtime_hrs", 0),
        })

    # Agregar fila de totales
    total_pares = sum(r["Pares"] for r in rows)
    total_hc_n = sum(r["HC_Necesario"] for r in rows)
    total_hc_d = sum(r["HC_Disponible"] for r in rows)
    rows.append({
        "Dia": "TOTAL",
        "Tipo": "",
        "Pares": total_pares,
        "HC_Necesario": round(total_hc_n, 1),
        "HC_Disponible": total_hc_d,
        "Diferencia": round(total_hc_d - total_hc_n, 1),
        "Utilizacion_%": "",
    })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Balance_Diario", index=False)


def _write_detalle(summary: dict, writer: pd.ExcelWriter):
    """Hoja de detalle por modelo."""
    rows = []
    for ms in summary["models"]:
        status = "OK" if ms["tardiness"] == 0 else "INCOMPLETO"
        rows.append({
            "Fabrica": ms["fabrica"],
            "Modelo": ms["codigo"],
            "Vol_Semana": ms["volumen"],
            "Producido": ms["producido"],
            "Pendiente": ms["tardiness"],
            "Completado_%": ms["pct_completado"],
            "Estado": status,
        })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Detalle_Modelos", index=False)


def _write_daily_program(day_name: str, day_data: dict, writer: pd.ExcelWriter):
    """Hoja de programa hora por hora para un dia (formato PROGRAMA SEM XX DIA)."""
    summary = day_data["summary"]
    block_labels = summary["block_labels"]

    # Usar assignments si existen, sino schedule original
    has_ops = "assignments" in day_data and day_data["assignments"]
    source = day_data["assignments"] if has_ops else day_data["schedule"]

    if not source:
        return

    sheet_name = f"PROG_{day_name.upper()[:3]}"

    # Construir filas del dataframe
    rows = []
    for entry in source:
        row = {
            "MODELO": entry["modelo"],
            "FRACC": entry.get("fraccion", ""),
            "OPERACION": entry.get("operacion", ""),
            "RECURSO": entry.get("recurso", ""),
        }
        # Agregar OPERARIO si hay asignaciones
        if has_ops:
            row["OPERARIO"] = entry.get("operario", "")
        row["RATE"] = entry.get("rate", 0)
        row["HC"] = entry.get("hc", 0)
        # Agregar columnas de bloques horarios
        block_pares = entry.get("block_pares", [])
        for b_idx, label in enumerate(block_labels):
            row[label] = block_pares[b_idx] if b_idx < len(block_pares) else 0
        row["TOTAL"] = entry.get("total_pares", 0)
        if has_ops:
            robot = entry.get("robot_asignado", "")
            if not robot and entry.get("robots_used"):
                robot = ", ".join(entry["robots_used"])
            row["ROBOT"] = robot
            row["PENDIENTE"] = entry.get("pendiente", 0)
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Calcular offset de columnas (cambia segun si tiene OPERARIO)
    ws = writer.sheets[sheet_name]
    # Encontrar la columna donde empiezan los bloques
    header_cols = list(df.columns)
    block_start_col = header_cols.index(block_labels[0]) + 1 if block_labels[0] in header_cols else 7
    total_cols = len(header_cols)

    # Agregar fila de TOTAL PARES por bloque
    total_row = len(df) + 2
    ws.cell(row=total_row, column=1, value="TOTAL PARES")
    for b_idx, label in enumerate(block_labels):
        col = block_start_col + b_idx
        total_val = summary["block_pares"][b_idx] if b_idx < len(summary["block_pares"]) else 0
        ws.cell(row=total_row, column=col, value=total_val)

    # Agregar fila de HC TOTAL por bloque
    hc_row = total_row + 1
    ws.cell(row=hc_row, column=1, value="HC TOTAL")
    for b_idx, label in enumerate(block_labels):
        col = block_start_col + b_idx
        hc_val = summary["block_hc"][b_idx] if b_idx < len(summary["block_hc"]) else 0
        ws.cell(row=hc_row, column=col, value=hc_val)

    ws.cell(row=hc_row + 1, column=1, value=f"PLANTILLA: {summary['plantilla']}")

    # Formateo: colores solo en celdas de bloques con actividad
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    recurso_col = header_cols.index("RECURSO") + 1 if "RECURSO" in header_cols else 4
    # Rango de columnas de bloques horarios + TOTAL
    total_col = header_cols.index("TOTAL") + 1 if "TOTAL" in header_cols else None
    block_col_set = set()
    for label in block_labels:
        if label in header_cols:
            block_col_set.add(header_cols.index(label) + 1)
    if total_col:
        block_col_set.add(total_col)

    for row_idx in range(2, len(df) + 2):
        recurso_cell = ws.cell(row=row_idx, column=recurso_col)
        recurso = str(recurso_cell.value or "")
        color = RESOURCE_COLORS.get(recurso, RESOURCE_COLORS.get("GENERAL", "D5DBDB"))
        res_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in block_col_set:
                # Solo colorear si tiene valor > 0
                val = cell.value
                if val and isinstance(val, (int, float)) and val > 0:
                    cell.fill = res_fill
                    cell.font = Font(color="FFFFFF", bold=True)
            elif col == recurso_col:
                # Celda RECURSO coloreada como indicador
                cell.fill = res_fill
                cell.font = Font(color="FFFFFF", bold=True)

    # Header en negrita
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Totales en negrita
    bold_font = Font(bold=True)
    for row_idx in [total_row, hc_row]:
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = bold_font


def _write_cascade_sheet(day_name: str, day_data: dict, writer: pd.ExcelWriter):
    """Hoja de cascada: timeline por operario (una fila por persona)."""
    timelines = day_data.get("operator_timelines", {})
    summary = day_data["summary"]
    block_labels = summary["block_labels"]

    if not timelines:
        return

    sheet_name = f"CASC_{day_name.upper()[:3]}"

    rows = []
    for op_name in sorted(timelines.keys()):
        row = {"OPERARIO": op_name}
        for b_idx, label in enumerate(block_labels):
            entry = next((e for e in timelines[op_name] if e["block"] == b_idx), None)
            if entry:
                robot_str = f" [{entry['robot']}]" if entry.get("robot") else ""
                row[label] = f"{entry['modelo']} {entry['operacion'][:20]} ({entry['pares']}p){robot_str}"
            else:
                row[label] = ""
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Formateo
    ws = writer.sheets[sheet_name]
    total_cols = len(df.columns)

    # Colores por modelo (para visualizar cascada)
    model_colors_hex = [
        "2E4057", "4A6F8C", "6B4E3D", "3E6B48",
        "5C3566", "8B6914", "2B5B84", "704214",
    ]
    model_list = []
    for _, row_data in df.iterrows():
        for label in block_labels:
            val = str(row_data.get(label, ""))
            if val:
                mc = val.split()[0] if val else ""
                if mc and mc not in model_list:
                    model_list.append(mc)
    model_color_map = {m: model_colors_hex[i % len(model_colors_hex)] for i, m in enumerate(model_list)}

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    white_font = Font(color="FFFFFF", size=9)

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(2, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value or "")
            if val:
                mc = val.split()[0] if val else ""
                color = model_color_map.get(mc, "2C3E50")
            else:
                color = "0E1117"
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = white_font
            cell.border = thin_border

    # Header
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Columna OPERARIO en negrita
    bold_font = Font(bold=True)
    for row_idx in range(2, len(df) + 2):
        ws.cell(row=row_idx, column=1).font = bold_font
