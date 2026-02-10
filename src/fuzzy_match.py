"""
fuzzy_match.py - Matching difuso para nombres de operarios.

Los documentos de la fabrica contienen errores de ortografia en nombres.
Este modulo normaliza y resuelve nombres a un registro canonico.

Fuente de referencia: columna 26 de hojas PROGRAMA SEM XX en la sabana.
"""

import re
import unicodedata
from difflib import SequenceMatcher

# Alias conocidos (detectados en datos reales de semana 6-8)
KNOWN_ALIASES = {
    "ARACELY": "ARACELI",
    "ARECELI": "ARACELI",
    "FABILA": "FABIOLA",
    "JAQUI": "JACQUI",
    "HUGI": "HUGO",
}

# Umbral de similitud para fuzzy matching
SIMILARITY_THRESHOLD = 0.85

# Palabras que no son nombres de operarios (etiquetas de header)
EXCLUDED_LABELS = {
    "HEADCOUNT", "HC", "PERSONAL ADICIONAL", "TOTAL", "DIFERENCIA",
    "OPERARIO", "NOMBRE", "RECURSO", "RATE", "PARES",
}


def normalize_name(name: str) -> str:
    """Normaliza un nombre: mayusculas, sin acentos, sin espacios extra."""
    if not name:
        return ""
    # Quitar acentos
    nfkd = unicodedata.normalize("NFKD", str(name))
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    # Mayusculas, colapsar espacios
    result = re.sub(r"\s+", " ", ascii_only.upper().strip())
    return result


def parse_operator_cell(cell_value) -> list:
    """
    Separa celdas con multiples operarios.
    Ej: 'VICTOR/ALVARO' -> ['VICTOR', 'ALVARO']
    Ej: 'VICTOR/ ALVARO' -> ['VICTOR', 'ALVARO']
    Ej: 'ARACELI' -> ['ARACELI']
    """
    if not cell_value:
        return []
    val = str(cell_value).strip()
    if not val:
        return []
    # Separar por /
    parts = [p.strip() for p in val.split("/")]
    return [normalize_name(p) for p in parts if p.strip()]


def resolve_operator(name: str, registry: dict) -> str:
    """
    Resuelve un nombre a su forma canonica.
    1. Busca en alias conocidos
    2. Busca en registro por nombre exacto
    3. Fuzzy match con umbral 85%
    4. Si no hay match, retorna el nombre normalizado
    """
    normalized = normalize_name(name)
    if not normalized:
        return ""

    # Paso 1: alias conocidos
    if normalized in KNOWN_ALIASES:
        return KNOWN_ALIASES[normalized]

    # Paso 2: match exacto en registro
    if normalized in registry:
        return normalized

    # Paso 3: fuzzy match
    best_match = None
    best_score = 0.0
    for canonical in registry:
        score = SequenceMatcher(None, normalized, canonical).ratio()
        if score > best_score:
            best_score = score
            best_match = canonical

    if best_score >= SIMILARITY_THRESHOLD and best_match:
        return best_match

    # Paso 4: no hay match -> retornar normalizado
    return normalized


def build_operator_registry(sabana_path: str) -> dict:
    """
    Extrae todos los nombres de operarios de las hojas PROGRAMA en la sabana.

    Returns:
        dict {nombre_canonico: {
            'days': [dias donde aparece],
            'count': total de apariciones,
            'aliases': [variantes encontradas]
        }}
    """
    import openpyxl

    wb = openpyxl.load_workbook(sabana_path, data_only=True)

    # Recolectar todos los nombres crudos (filtrar etiquetas de header)
    raw_names = []  # (nombre, hoja)
    for sheet_name in wb.sheetnames:
        if "PROGRAMA" not in sheet_name.upper():
            continue
        ws = wb[sheet_name]
        day_label = sheet_name.split()[-1] if sheet_name.split() else sheet_name
        for row in range(9, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=26).value
            if cell_val:
                names = parse_operator_cell(cell_val)
                for n in names:
                    if n and n not in EXCLUDED_LABELS:
                        raw_names.append((n, day_label))

    # Construir registro: agrupar por nombre canonico
    registry = {}
    for raw, day in raw_names:
        # Resolver alias
        canonical = KNOWN_ALIASES.get(raw, raw)
        if canonical not in registry:
            registry[canonical] = {
                "days": set(),
                "count": 0,
                "aliases": set(),
            }
        registry[canonical]["days"].add(day)
        registry[canonical]["count"] += 1
        if raw != canonical:
            registry[canonical]["aliases"].add(raw)

    # Fuzzy merge: buscar nombres muy similares no capturados por alias
    canonical_names = list(registry.keys())
    merges = {}
    for i, name_a in enumerate(canonical_names):
        for name_b in canonical_names[i + 1:]:
            score = SequenceMatcher(None, name_a, name_b).ratio()
            if score >= SIMILARITY_THRESHOLD:
                # Mantener el mas frecuente como canonico
                if registry[name_a]["count"] >= registry[name_b]["count"]:
                    merges[name_b] = name_a
                else:
                    merges[name_a] = name_b

    for src, dst in merges.items():
        if src in registry and dst in registry:
            registry[dst]["days"] |= registry[src]["days"]
            registry[dst]["count"] += registry[src]["count"]
            registry[dst]["aliases"].add(src)
            registry[dst]["aliases"] |= registry[src]["aliases"]
            del registry[src]

    # Convertir sets a listas para serializacion
    for data in registry.values():
        data["days"] = sorted(data["days"])
        data["aliases"] = sorted(data["aliases"])

    return registry
