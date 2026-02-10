"""
loader.py - Carga de datos desde la sabana real y el catalogo de fracciones.

Parsea dos archivos Excel:
  1. Sabana semanal: modelos, volumenes, fabricas, dias
  2. Catalogo de fracciones: operaciones y rates por modelo

Hace match entre ambos y construye las estructuras para el optimizador.
"""

import re
import openpyxl


# Columnas de la sabana donde inician los datos de cada dia (PRS)
# Cada dia ocupa 5 columnas: PRS, PREELIMINAR, OPERARIO, ROBOT, POST
DAY_START_COLS = [12, 17, 22, 27, 32, 37]


def load_sabana(filepath: str) -> tuple:
    """
    Parsea la sabana semanal para extraer modelos, volumenes y estructura de dias.

    Returns:
        (models, days) donde:
        - models: lista de dicts con codigo, volumen, fabrica, etc.
        - days: lista de dicts con nombre del dia y columnas asociadas
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Buscar la hoja SEM XX (hoja principal de planificacion)
    ws = None
    for name in wb.sheetnames:
        if re.match(r"SEM\s+\d+", name):
            ws = wb[name]
            break

    if ws is None:
        raise ValueError(
            f"No se encontro hoja 'SEM XX' en la sabana. "
            f"Hojas disponibles: {wb.sheetnames}"
        )

    # Parsear encabezados de dias (fila 16)
    days = []
    for col in DAY_START_COLS:
        day_name = ws.cell(row=16, column=col).value
        if day_name:
            days.append({
                "name": str(day_name).strip(),
                "prs_col": col,
                "pre_col": col + 1,
                "opr_col": col + 2,
                "robot_col": col + 3,
                "post_col": col + 4,
            })

    # Parsear modelos: buscar filas donde col 8 tiene codigo de modelo
    models = []
    current_fabrica = None

    for row in range(18, ws.max_row + 1):
        # Detectar marcador de fabrica (col 4)
        fab_val = ws.cell(row=row, column=4).value
        if fab_val and "FABRICA" in str(fab_val).upper():
            current_fabrica = str(fab_val).strip()

        # Detectar fila de modelo (col 8 tiene codigo tipo "65413 NE")
        model_code = ws.cell(row=row, column=8).value
        if not model_code:
            continue

        code_str = str(model_code).strip()
        if code_str.startswith("TOTAL") or not re.match(r"^\d{4,6}", code_str):
            continue

        # Extraer numero de modelo (primeros digitos)
        model_num = re.match(r"^(\d+)", code_str).group(1)

        # Volumen de la semana (col 11)
        volume = ws.cell(row=row, column=11).value
        volume = int(volume) if volume and isinstance(volume, (int, float)) else 0

        # Suela / cliente (col 10)
        suela = ws.cell(row=row, column=10).value or ""

        # Leer PRS asignados por dia en la fila del modelo
        daily_prs = {}
        for day in days:
            prs = ws.cell(row=row, column=day["prs_col"]).value
            if prs and isinstance(prs, (int, float)) and prs > 0:
                daily_prs[day["name"]] = int(prs)

        # Volumen real = max entre volumen declarado y suma de PRS diarios
        sum_prs = sum(daily_prs.values())
        total_producir = max(volume, sum_prs)

        if total_producir <= 0:
            continue  # Modelo sin produccion esta semana

        models.append({
            "codigo": code_str,
            "modelo_num": model_num,
            "suela": str(suela).strip(),
            "volumen_declarado": volume,
            "total_producir": total_producir,
            "fabrica": current_fabrica or "SIN FABRICA",
            "daily_prs_original": daily_prs,
        })

    return models, days


def load_catalog(filepath: str) -> dict:
    """
    Parsea el catalogo de fracciones para extraer operaciones y rates por modelo.

    Returns:
        dict {modelo_num: {codigo_full, operations, total_sec_per_pair, num_ops}}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["PLANTILLA MOD."]

    # Recolectar operaciones agrupadas por numero de modelo
    raw_ops = {}  # modelo_num -> list of ops
    current_model_num = None

    for row in range(11, ws.max_row + 1):
        modelo_val = ws.cell(row=row, column=1).value
        fraccion = ws.cell(row=row, column=5).value
        operacion = ws.cell(row=row, column=6).value
        etapa = ws.cell(row=row, column=7).value
        recurso = ws.cell(row=row, column=16).value
        tiempo_std = ws.cell(row=row, column=17).value
        rate = ws.cell(row=row, column=18).value

        # Detectar nueva seccion de modelo
        if modelo_val:
            modelo_str = str(modelo_val).strip()
            match = re.match(r"^(\d+)", modelo_str)
            if match:
                current_model_num = match.group(1)
                if current_model_num not in raw_ops:
                    raw_ops[current_model_num] = {
                        "codigo_full": modelo_str,
                        "ops": [],
                    }

        # Agregar operacion si tiene fraccion y rate
        if current_model_num and fraccion and rate:
            rate_val = float(rate)
            if rate_val <= 0:
                continue

            # Calcular segundos por par: usar TE si existe, sino calcular del rate
            if tiempo_std:
                sec_per_pair = float(tiempo_std)
            else:
                sec_per_pair = 3600.0 / rate_val

            raw_ops[current_model_num]["ops"].append({
                "fraccion": int(fraccion),
                "operacion": str(operacion).strip() if operacion else f"OP {etapa or 'AUTO'}",
                "etapa": str(etapa).strip() if etapa else "",
                "recurso": str(recurso).strip() if recurso else "",
                "rate": round(rate_val, 2),
                "sec_per_pair": round(sec_per_pair),
            })

    # Deduplicar operaciones por fraccion y calcular totales
    catalog = {}
    for model_num, data in raw_ops.items():
        seen = set()
        unique_ops = []
        for op in data["ops"]:
            key = op["fraccion"]
            if key not in seen:
                seen.add(key)
                unique_ops.append(op)

        unique_ops.sort(key=lambda x: x["fraccion"])
        total_sec = sum(op["sec_per_pair"] for op in unique_ops)

        catalog[model_num] = {
            "codigo_full": data["codigo_full"],
            "operations": unique_ops,
            "total_sec_per_pair": total_sec,
            "num_ops": len(unique_ops),
        }

    return catalog


def match_models(sabana_models: list, catalog: dict) -> tuple:
    """
    Cruza modelos de la sabana con el catalogo de fracciones.

    Returns:
        (matched, unmatched) donde:
        - matched: modelos con datos completos (sabana + catalogo)
        - unmatched: modelos sin rates en el catalogo
    """
    matched = []
    unmatched = []

    for model in sabana_models:
        model_num = model["modelo_num"]

        if model_num in catalog:
            cat = catalog[model_num]
            model["operations"] = cat["operations"]
            model["total_sec_per_pair"] = cat["total_sec_per_pair"]
            model["num_ops"] = cat["num_ops"]
            model["catalog_code"] = cat["codigo_full"]
            if "resource_summary" in cat:
                model["resource_summary"] = cat["resource_summary"]
            matched.append(model)
        else:
            unmatched.append(model)

    return matched, unmatched
