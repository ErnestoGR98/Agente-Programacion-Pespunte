"""
Endpoint de importacion de Excel + descarga de template.

Parsea archivos Excel (catalogo, pedido) y guarda en Supabase.
Genera template descargable con formato correcto.
"""

import os
import tempfile
import requests
import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse

from excel_parsers import (
    _parse_catalogo_sheet,
    _parse_pedido_sheet,
)
from catalog_loader import load_catalog_v2

router = APIRouter()

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")


def _sb_headers(prefer="return=representation"):
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": prefer,
    }


def _sb_get(table, query=""):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{table}?{query}", headers=_sb_headers())
    r.raise_for_status()
    return r.json()


def _sb_post(table, data):
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=_sb_headers(), json=data)
    r.raise_for_status()
    result = r.json()
    return result[0] if isinstance(result, list) and result else result


def _sb_upsert(table, data, on_conflict=""):
    headers = _sb_headers(prefer="return=representation,resolution=merge-duplicates")
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=headers, json=data)
    r.raise_for_status()
    result = r.json()
    return result[0] if isinstance(result, list) and result else result


def _sb_delete(table, query):
    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/{table}?{query}",
        headers=_sb_headers(),
    )
    r.raise_for_status()


# --- Template download ---

@router.get("/template")
def download_template():
    """Genera y descarga template Excel para importar pedido semanal."""
    from template_generator import generate_template

    buf = generate_template()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_pedido.xlsx"},
    )


# --- Import catalog ---

@router.post("/import-catalog")
async def import_catalog(file: UploadFile = File(...)):
    """Importa catalogo desde Excel y guarda en Supabase."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx)")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # Detectar formato: "PLANTILLA MOD." = formato de fabrica (catalog_loader)
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        is_plantilla = "PLANTILLA MOD." in wb.sheetnames
        wb.close()

        errors = []
        if is_plantilla:
            catalogo = load_catalog_v2(tmp_path)
        else:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb["CATALOGO"] if "CATALOGO" in wb.sheetnames else wb.active
            catalogo, errors = _parse_catalogo_sheet(ws)
            wb.close()

        if not catalogo:
            detail = "No se encontraron modelos en el archivo"
            if errors:
                detail += f". Errores: {'; '.join(errors[:5])}"
            raise HTTPException(400, detail)

        # Obtener mapa de robots
        robots = _sb_get("robots", "select=id,nombre")
        robot_map = {r["nombre"]: r["id"] for r in robots}

        saved = 0
        for modelo_num, data in catalogo.items():
            modelo_row = _sb_upsert("catalogo_modelos", {
                "modelo_num": modelo_num,
                "codigo_full": data.get("codigo_full", modelo_num),
                "alternativas": data.get("alternativas", []),
                "clave_material": data.get("clave_material", ""),
                "total_sec_per_pair": data.get("total_sec_per_pair", 0),
                "num_ops": len(data.get("operations", [])),
            })
            modelo_id = modelo_row["id"]

            _sb_delete("catalogo_operaciones", f"modelo_id=eq.{modelo_id}")

            for op in data.get("operations", []):
                op_row = _sb_post("catalogo_operaciones", {
                    "modelo_id": modelo_id,
                    "fraccion": op["fraccion"],
                    "operacion": op["operacion"],
                    "input_o_proceso": op.get("input_o_proceso", "PRELIMINARES"),
                    "etapa": op.get("etapa", ""),
                    "recurso": op["recurso"],
                    "recurso_raw": op.get("recurso_raw", ""),
                    "rate": op.get("rate", 0),
                    "sec_per_pair": op.get("sec_per_pair", 0),
                })

                for rname in op.get("robots", []):
                    rid = robot_map.get(rname)
                    if rid:
                        _sb_post("catalogo_operacion_robots", {
                            "operacion_id": op_row["id"],
                            "robot_id": rid,
                        })

            saved += 1

        return {
            "modelos_importados": saved,
            "total_operaciones": sum(
                len(d.get("operations", [])) for d in catalogo.values()
            ),
            "warnings": errors[:10] if errors else [],
        }

    finally:
        os.unlink(tmp_path)


# --- Import pedido ---

@router.post("/import-pedido/{nombre}")
async def import_pedido(nombre: str, file: UploadFile = File(...)):
    """Importa pedido desde Excel y guarda en Supabase."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx)")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb["PEDIDO"] if "PEDIDO" in wb.sheetnames else wb.active
        items, errors, semana = _parse_pedido_sheet(ws)
        wb.close()

        if not items:
            detail = "No se encontraron items en el archivo"
            if errors:
                detail += f". Errores: {'; '.join(errors[:5])}"
            raise HTTPException(400, detail)

        ped = _sb_upsert("pedidos", {"nombre": nombre})
        ped_id = ped["id"]

        _sb_delete("pedido_items", f"pedido_id=eq.{ped_id}")

        for it in items:
            _sb_post("pedido_items", {
                "pedido_id": ped_id,
                "modelo_num": it["modelo"],
                "color": it.get("color", ""),
                "clave_material": it.get("clave_material", ""),
                "fabrica": it.get("fabrica", ""),
                "volumen": it["volumen"],
            })

        return {
            "nombre": nombre,
            "items_importados": len(items),
            "warnings": errors[:10] if errors else [],
        }

    finally:
        os.unlink(tmp_path)
