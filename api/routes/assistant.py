"""
Endpoint del asistente LLM (Claude API) con tool use.

Recibe mensajes del chat, construye contexto base desde Supabase,
y permite a Claude consultar datos adicionales bajo demanda via tools.
Soporta attachments: imagenes (Claude Vision) y Excel (openpyxl).
"""

import os
import io
import json
import base64
import requests
from typing import Optional
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
import openpyxl

from llm_assistant import SYSTEM_PROMPT, build_context

router = APIRouter()

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "") or os.environ.get("SUPABASE_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

MAX_TOOL_ROUNDS = 5  # Max tool use iterations per request


def _sb_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }


def _sb_get(table: str, query: str = "") -> list:
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/{table}?{query}",
        headers=_sb_headers(),
    )
    if r.status_code == 400:
        return []
    r.raise_for_status()
    return r.json()


# ============================================================
# TOOL DEFINITIONS (schemas for Claude)
# ============================================================

TOOLS = [
    {
        "name": "consultar_programa_dia",
        "description": (
            "Consulta el programa diario completo de un dia especifico. "
            "Retorna cada operacion con: modelo, fraccion, operacion, recurso, "
            "operario asignado, rate, HC, pares por bloque (8-9 a 5-6), total. "
            "Usa esto cuando necesites ver el schedule detallado de un dia."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "dia": {
                    "type": "string",
                    "description": "Nombre del dia: Lun, Mar, Mie, Jue, Vie, Sab",
                }
            },
            "required": ["dia"],
        },
    },
    {
        "name": "consultar_operario",
        "description": (
            "Consulta informacion detallada de un operario: habilidades con nivel, "
            "dias disponibles, eficiencia, y su timeline bloque-a-bloque si hay "
            "resultado de optimizacion (que hace en cada bloque de cada dia)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "nombre": {
                    "type": "string",
                    "description": "Nombre (o parte del nombre) del operario",
                }
            },
            "required": ["nombre"],
        },
    },
    {
        "name": "consultar_modelo",
        "description": (
            "Consulta informacion completa de un modelo del catalogo: operaciones "
            "con fraccion, recurso, rate, robots elegibles, input_o_proceso. "
            "Tambien retorna su schedule semanal (pares por dia) y pedido."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "modelo_num": {
                    "type": "string",
                    "description": "Numero del modelo (ej: 65413, 61747)",
                }
            },
            "required": ["modelo_num"],
        },
    },
    {
        "name": "consultar_robot",
        "description": (
            "Consulta la utilizacion de un robot especifico: tipos, estado, "
            "y en que bloques/dias esta ocupado o libre segun el programa diario."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "nombre": {
                    "type": "string",
                    "description": "Nombre del robot (ej: 2A-3020-M1, M048-CHACHE, 6040-M4)",
                }
            },
            "required": ["nombre"],
        },
    },
    {
        "name": "consultar_restricciones",
        "description": (
            "Consulta todas las restricciones activas (temporales de la semana "
            "y reglas permanentes). Opcionalmente filtra por modelo o tipo."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "modelo": {
                    "type": "string",
                    "description": "Filtrar por modelo (opcional, ej: 65413)",
                },
                "tipo": {
                    "type": "string",
                    "description": "Filtrar por tipo de restriccion (opcional, ej: PRIORIDAD, FIJAR_DIA)",
                },
            },
            "required": [],
        },
    },
    {
        "name": "consultar_resumen_semanal",
        "description": (
            "Consulta el resumen completo de la optimizacion semanal: "
            "KPIs por dia (pares, HC, utilizacion, overtime), "
            "KPIs por modelo (volumen, producido, tardiness, span, dias). "
            "Usa esto para tener una vision general antes de profundizar."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
    {
        "name": "consultar_operarios_dia",
        "description": (
            "Consulta la utilizacion de TODOS los operarios en un dia: "
            "porcentaje de ocupacion, bloques ociosos, y que hacen en cada bloque."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "dia": {
                    "type": "string",
                    "description": "Nombre del dia: Lun, Mar, Mie, Jue, Vie, Sab",
                }
            },
            "required": ["dia"],
        },
    },
    {
        "name": "consultar_pedido",
        "description": (
            "Consulta el pedido completo: todos los items con modelo, color, "
            "clave_material, volumen, fabrica. Incluye asignaciones de maquila."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
]


# ============================================================
# TOOL EXECUTION FUNCTIONS
# ============================================================

def _exec_consultar_programa_dia(args: dict, state: dict) -> str:
    dia = args.get("dia", "")
    daily = state.get("daily_results") or {}
    dr = daily.get(dia)
    if not dr:
        return f"No hay programa para el dia '{dia}'. Dias disponibles: {list(daily.keys())}"

    BLOCK_LABELS = ["8-9", "9-10", "10-11", "11-12", "12-1", "1-2", "COMIDA", "3-4", "4-5", "5-6"]
    lines = [f"Programa {dia}: {dr.get('total_pares', 0)} pares, "
             f"tardiness={dr.get('total_tardiness', 0)}, plantilla={dr.get('plantilla', 0)}, "
             f"status={dr.get('status', '?')}"]
    for s in dr.get("schedule", []):
        blocks = s.get("blocks", [])
        block_str = " | ".join(
            f"{BLOCK_LABELS[i]}:{blocks[i]}" for i in range(min(len(blocks), len(BLOCK_LABELS)))
            if i < len(blocks) and blocks[i]
        )
        operario = s.get("operario", "-")
        sin_asignar = " ⚠SIN ASIGNAR" if operario == "SIN ASIGNAR" else ""
        motivo = f" ({s.get('motivo_sin_asignar', '')})" if s.get("motivo_sin_asignar") else ""
        lines.append(
            f"  {s.get('modelo','?')} F{s.get('fraccion','?')} {s.get('operacion','?')[:35]} "
            f"[{s.get('recurso','?')}] rate={s.get('rate',0)} HC={s.get('hc',0)} "
            f"-> {operario}{sin_asignar}{motivo} | {block_str} | TOTAL={s.get('total',0)}"
        )
    return "\n".join(lines)


def _exec_consultar_operario(args: dict, state: dict) -> str:
    nombre_q = args.get("nombre", "").upper()
    # Search in operarios
    operarios = state.get("operarios") or []
    matches = [op for op in operarios if nombre_q in op.get("nombre", "").upper()]
    if not matches:
        nombres = [op["nombre"] for op in operarios]
        return f"No se encontro operario '{nombre_q}'. Disponibles: {nombres}"

    BLOCK_LABELS = ["8-9", "9-10", "10-11", "11-12", "12-1", "1-2", "COMIDA", "3-4", "4-5", "5-6"]
    lines = []
    for op in matches:
        skills = op.get("habilidades", [])
        skill_str = ", ".join(f"{h['habilidad']}(N{h['nivel']})" for h in skills)
        dias = op.get("dias_disponibles", [])
        lines.append(f"{op['nombre']}: eficiencia={op.get('eficiencia',1.0)}, "
                     f"dias=[{', '.join(dias) if dias else 'todos'}], "
                     f"skills=[{skill_str}]")

        # Timeline from daily results
        daily = state.get("daily_results") or {}
        for day_name, dr in daily.items():
            timelines = dr.get("operator_timelines") or {}
            # Find matching timeline key
            for tl_name, entries in timelines.items():
                if nombre_q in tl_name.upper():
                    lines.append(f"  {day_name}:")
                    if isinstance(entries, dict):
                        for b_str in sorted(entries.keys(), key=lambda x: int(x)):
                            tasks = entries[b_str]
                            bi = int(b_str)
                            bl = BLOCK_LABELS[bi] if bi < len(BLOCK_LABELS) else b_str
                            if tasks:
                                task_str = ", ".join(str(t) for t in tasks) if isinstance(tasks, list) else str(tasks)
                                lines.append(f"    {bl}: {task_str}")
                            else:
                                lines.append(f"    {bl}: IDLE")
                    elif isinstance(entries, list):
                        for e in entries:
                            if isinstance(e, dict):
                                bi = e.get("block", 0)
                                bl = BLOCK_LABELS[bi] if bi < len(BLOCK_LABELS) else str(bi)
                                lines.append(f"    {bl}: {e.get('label', '?')} ({e.get('modelo','?')} F{e.get('fraccion','?')})")
                    break
    return "\n".join(lines) if lines else "Sin datos"


def _exec_consultar_modelo(args: dict, state: dict) -> str:
    modelo_q = args.get("modelo_num", "")
    # Catalog
    catalogo = state.get("catalogo") or []
    match = None
    for m in catalogo:
        if modelo_q in m.get("modelo_num", ""):
            match = m
            break
    if not match:
        return f"Modelo '{modelo_q}' no encontrado en catalogo. Disponibles: {[m['modelo_num'] for m in catalogo]}"

    lines = [f"MODELO {match['modelo_num']}: {match.get('num_ops',0)} ops, "
             f"{match.get('total_sec_per_pair',0)} sec/par"]
    if match.get("alternativas"):
        lines.append(f"  Alternativas: {', '.join(match['alternativas'])}")
    for op in match.get("operaciones", []):
        robots = f" [{', '.join(op['robots'])}]" if op.get("robots") else ""
        proceso = f" ({op.get('input_o_proceso','')})" if op.get("input_o_proceso") else ""
        lines.append(f"  F{op['fraccion']} {op['operacion']}: {op['recurso']} rate={op['rate']}{proceso}{robots}")

    # Weekly schedule for this model
    schedule = state.get("weekly_schedule") or []
    model_entries = [e for e in schedule if modelo_q in e.get("Modelo", "")]
    if model_entries:
        lines.append("Schedule semanal:")
        for e in model_entries:
            lines.append(f"  {e['Dia']}: {e['Pares']}p ({e.get('Fabrica','')})")

    # Pedido
    pedido = state.get("pedido_rows") or []
    ped_items = [p for p in pedido if modelo_q in p.get("modelo", "")]
    if ped_items:
        lines.append("Pedido:")
        for p in ped_items:
            lines.append(f"  {p['modelo']} {p.get('color','')}: {p['volumen']}p ({p.get('fabrica','')})")

    return "\n".join(lines)


def _exec_consultar_robot(args: dict, state: dict) -> str:
    robot_q = args.get("nombre", "").upper()
    BLOCK_LABELS = ["8-9", "9-10", "10-11", "11-12", "12-1", "1-2", "COMIDA", "3-4", "4-5", "5-6"]
    PRODUCTIVE = [i for i, lb in enumerate(BLOCK_LABELS) if lb != "COMIDA"]

    lines = []
    # Robot info from catalog
    robots_activos = state.get("robots_activos") or []
    found = [r for r in robots_activos if robot_q in r.upper()]
    if not found:
        return f"Robot '{robot_q}' no encontrado. Activos: {robots_activos}"
    lines.append(f"Robot: {', '.join(found)}")

    # Usage per day from daily results
    daily = state.get("daily_results") or {}
    for day_name, dr in daily.items():
        for s in dr.get("schedule", []):
            robot = s.get("robot") or ""
            if not robot or robot_q not in robot.upper():
                continue
            blocks = s.get("blocks", [])
            active_blocks = [BLOCK_LABELS[i] for i in range(min(len(blocks), len(BLOCK_LABELS))) if i < len(blocks) and blocks[i]]
            lines.append(
                f"  {day_name}: {s.get('modelo','?')} F{s.get('fraccion','?')} "
                f"{s.get('operacion','?')[:30]} -> bloques [{', '.join(active_blocks)}] "
                f"({s.get('total',0)}p)"
            )

    # Summarize free blocks per day
    for day_name, dr in daily.items():
        busy = set()
        for s in dr.get("schedule", []):
            robot = s.get("robot") or ""
            if robot_q not in robot.upper():
                continue
            blocks = s.get("blocks", [])
            for bi, p in enumerate(blocks):
                if p and p > 0:
                    busy.add(bi)
        if busy:
            free = [BLOCK_LABELS[i] for i in PRODUCTIVE if i not in busy]
            pct = int(100 * len(busy & set(PRODUCTIVE)) / len(PRODUCTIVE))
            lines.append(f"  {day_name} resumen: {pct}% ocupado, libre=[{', '.join(free)}]")

    return "\n".join(lines) if lines else f"Robot '{robot_q}' sin actividad en el programa actual"


def _exec_consultar_restricciones(args: dict, state: dict) -> str:
    restricciones = state.get("restricciones") or []
    modelo_f = args.get("modelo", "")
    tipo_f = args.get("tipo", "")

    filtered = restricciones
    if modelo_f:
        filtered = [r for r in filtered if modelo_f in r.get("modelo", "") or r.get("modelo") == "*"]
    if tipo_f:
        filtered = [r for r in filtered if tipo_f.upper() in r.get("tipo", "").upper()]

    if not filtered:
        return f"No hay restricciones activas{' para modelo ' + modelo_f if modelo_f else ''}{' tipo ' + tipo_f if tipo_f else ''}."

    lines = []
    temporales = [r for r in filtered if r.get("categoria") == "temporal"]
    permanentes = [r for r in filtered if r.get("categoria") == "permanente"]
    sin_cat = [r for r in filtered if not r.get("categoria")]
    temporales.extend(sin_cat)

    if temporales:
        lines.append(f"RESTRICCIONES TEMPORALES ({len(temporales)}):")
        for r in temporales:
            lines.append(f"  [{r['tipo']}] modelo={r.get('modelo','*')} "
                        f"params={json.dumps(r.get('parametros', {}), ensure_ascii=False)}"
                        f"{' nota=' + r['nota'] if r.get('nota') else ''}")
    if permanentes:
        lines.append(f"REGLAS PERMANENTES ({len(permanentes)}):")
        for r in permanentes:
            lines.append(f"  [{r['tipo']}] modelo={r.get('modelo','*')} "
                        f"params={json.dumps(r.get('parametros', {}), ensure_ascii=False)}"
                        f"{' nota=' + r['nota'] if r.get('nota') else ''}")
    return "\n".join(lines)


def _exec_consultar_resumen_semanal(args: dict, state: dict) -> str:
    summary = state.get("weekly_summary")
    if not summary:
        return "No hay resultado de optimizacion semanal."

    lines = [f"RESUMEN SEMANAL: status={summary['status']}, "
             f"total={summary['total_pares']}p, tardiness={summary['total_tardiness']}p, "
             f"solver={summary['wall_time_s']}s"]

    lines.append("\nPor dia:")
    for ds in summary.get("days", []):
        ot = f" +{ds['overtime_hrs']}h OT" if ds.get("overtime_hrs", 0) > 0 else ""
        sat = " (SAB)" if ds.get("is_saturday") else ""
        lines.append(f"  {ds['dia']}{sat}: {ds['pares']}p, "
                     f"HC {ds['hc_necesario']}/{ds['hc_disponible']}, "
                     f"{ds['utilizacion_pct']}% util{ot}")

    lines.append("\nPor modelo:")
    for ms in summary.get("models", []):
        tard = f" ⚠TARD={ms['tardiness']}p" if ms.get("tardiness", 0) > 0 else ""
        dias = ms.get("dias_produccion", [])
        dias_str = ", ".join(dias) if isinstance(dias, list) else str(dias)
        lines.append(f"  {ms['codigo']}: {ms.get('producido',0)}/{ms.get('volumen',0)}p "
                     f"({ms.get('pct_completado',0)}%), dias=[{dias_str}], span={ms.get('span_dias',0)}{tard}")
    return "\n".join(lines)


def _exec_consultar_operarios_dia(args: dict, state: dict) -> str:
    dia = args.get("dia", "")
    daily = state.get("daily_results") or {}
    dr = daily.get(dia)
    if not dr:
        return f"No hay datos para '{dia}'. Dias: {list(daily.keys())}"

    BLOCK_LABELS = ["8-9", "9-10", "10-11", "11-12", "12-1", "1-2", "COMIDA", "3-4", "4-5", "5-6"]
    PRODUCTIVE = [i for i, lb in enumerate(BLOCK_LABELS) if lb != "COMIDA"]
    timelines = dr.get("operator_timelines") or {}

    if not timelines:
        return f"No hay timelines de operarios para {dia}."

    lines = [f"OPERARIOS {dia} ({len(timelines)} operarios):"]
    for op_name, entries in sorted(timelines.items()):
        busy_blocks = set()
        block_tasks = {}
        if isinstance(entries, dict):
            for b_str, tasks in entries.items():
                bi = int(b_str)
                if tasks:
                    busy_blocks.add(bi)
                    if isinstance(tasks, list):
                        block_tasks[bi] = ", ".join(str(t) for t in tasks)
                    else:
                        block_tasks[bi] = str(tasks)
        elif isinstance(entries, list):
            for e in entries:
                if isinstance(e, dict):
                    bi = e.get("block", 0)
                    busy_blocks.add(bi)
                    block_tasks[bi] = e.get("label", "?")

        busy_prod = busy_blocks & set(PRODUCTIVE)
        pct = int(100 * len(busy_prod) / len(PRODUCTIVE)) if PRODUCTIVE else 0
        idle = [BLOCK_LABELS[i] for i in PRODUCTIVE if i not in busy_blocks]

        # Compact timeline
        timeline_parts = []
        for bi in PRODUCTIVE:
            bl = BLOCK_LABELS[bi]
            if bi in block_tasks:
                timeline_parts.append(f"{bl}:{block_tasks[bi]}")
            else:
                timeline_parts.append(f"{bl}:---")

        lines.append(f"  {op_name}: {pct}%")
        lines.append(f"    [{' | '.join(timeline_parts)}]")
        if idle:
            lines.append(f"    idle: {', '.join(idle)}")

    return "\n".join(lines)


def _exec_consultar_pedido(args: dict, state: dict) -> str:
    pedido = state.get("pedido_rows") or []
    if not pedido:
        return "No hay pedido cargado."

    lines = [f"PEDIDO ({len(pedido)} items):"]
    total = 0
    for p in sorted(pedido, key=lambda x: x.get("modelo", "")):
        color = f" {p.get('color','')}" if p.get("color") else ""
        clave = f" [{p.get('clave_material','')}]" if p.get("clave_material") else ""
        lines.append(f"  {p['modelo']}{color}{clave}: {p['volumen']}p ({p.get('fabrica','')})")
        total += p.get("volumen", 0)
    lines.append(f"TOTAL: {total:,} pares")

    # Maquila assignments
    maquila = state.get("asignaciones_maquila") or []
    if maquila:
        lines.append(f"\nASIGNACIONES MAQUILA ({len(maquila)}):")
        for a in maquila:
            lines.append(f"  {a.get('modelo','?')}: {a.get('pares',0)}p -> {a.get('maquila','?')} "
                        f"(entrega: {a.get('fecha_entrega','?')})")
    return "\n".join(lines)


# Tool dispatcher
TOOL_HANDLERS = {
    "consultar_programa_dia": _exec_consultar_programa_dia,
    "consultar_operario": _exec_consultar_operario,
    "consultar_modelo": _exec_consultar_modelo,
    "consultar_robot": _exec_consultar_robot,
    "consultar_restricciones": _exec_consultar_restricciones,
    "consultar_resumen_semanal": _exec_consultar_resumen_semanal,
    "consultar_operarios_dia": _exec_consultar_operarios_dia,
    "consultar_pedido": _exec_consultar_pedido,
}


# ============================================================
# BUILD STATE (loads data from Supabase)
# ============================================================

def _build_state_from_supabase(pedido_nombre: str, semana: str = "") -> dict:
    """Construye un dict compatible con build_context() y tool handlers desde Supabase."""
    state = {}

    # Pedido (con color y clave_material)
    items = []
    ped = _sb_get("pedidos", f"select=id&nombre=eq.{pedido_nombre}")
    if ped:
        items = _sb_get("pedido_items", f"select=*&pedido_id=eq.{ped[0]['id']}")
        state["pedido_rows"] = [
            {
                "modelo": it["modelo_num"],
                "color": it.get("color", ""),
                "clave_material": it.get("clave_material", ""),
                "volumen": it["volumen"],
                "fabrica": it.get("fabrica", ""),
            }
            for it in items
        ]

    # Resultado mas reciente
    q = "select=*&order=fecha_optimizacion.desc&limit=1"
    if semana:
        q = f"select=*&base_name=eq.{semana}&order=version.desc&limit=1"
    resultados = _sb_get("resultados", q)

    if resultados:
        res = resultados[0]
        state["weekly_schedule"] = res.get("weekly_schedule")
        state["weekly_summary"] = res.get("weekly_summary")
        state["daily_results"] = res.get("daily_results")
        state["params"] = res.get("params_snapshot")

    # Restricciones temporales (de la semana)
    restricciones = []
    if semana:
        rq = f"select=*&activa=eq.true&semana=eq.{semana}&order=created_at"
        restricciones = _sb_get("restricciones", rq)

    # Reglas permanentes (semana IS NULL)
    reglas = _sb_get("restricciones", "select=*&activa=eq.true&semana=is.null&order=created_at")

    all_constraints = []
    for r in restricciones:
        all_constraints.append({
            "tipo": r["tipo"],
            "modelo": r.get("modelo_num", ""),
            "activa": r["activa"],
            "parametros": r.get("parametros", {}),
            "nota": r.get("nota", ""),
            "categoria": "temporal",
        })
    for r in reglas:
        all_constraints.append({
            "tipo": r["tipo"],
            "modelo": r.get("modelo_num", ""),
            "activa": r["activa"],
            "parametros": r.get("parametros", {}),
            "nota": r.get("nota", ""),
            "categoria": "permanente",
        })
    state["restricciones"] = all_constraints

    # Catalogo (modelos + operaciones + robots)
    cat_modelos = _sb_get("catalogo_modelos", "select=id,modelo_num,alternativas,total_sec_per_pair,num_ops&order=modelo_num")
    if cat_modelos:
        cat_ops = _sb_get("catalogo_operaciones", "select=modelo_id,fraccion,operacion,input_o_proceso,recurso,etapa,rate,sec_per_pair&order=fraccion")
        robots_activos = _sb_get("robots", "select=id,nombre&estado=eq.ACTIVO&order=orden")
        robot_rels = _sb_get("catalogo_operacion_robots", "select=operacion_id,robot_id")

        robot_name_map = {r["id"]: r["nombre"] for r in robots_activos}
        op_robots = {}
        for rel in robot_rels:
            oid = rel["operacion_id"]
            rname = robot_name_map.get(rel["robot_id"], "")
            if rname:
                op_robots.setdefault(oid, []).append(rname)

        ops_by_modelo = {}
        for op in cat_ops:
            mid = op["modelo_id"]
            ops_by_modelo.setdefault(mid, []).append(op)

        state["catalogo"] = []
        for m in cat_modelos:
            ops = ops_by_modelo.get(m["id"], [])
            state["catalogo"].append({
                "modelo_num": m["modelo_num"],
                "alternativas": m.get("alternativas", []),
                "total_sec_per_pair": m.get("total_sec_per_pair", 0),
                "num_ops": m.get("num_ops", 0),
                "operaciones": [
                    {
                        "fraccion": op["fraccion"],
                        "operacion": op["operacion"],
                        "input_o_proceso": op.get("input_o_proceso", ""),
                        "recurso": op["recurso"],
                        "etapa": op.get("etapa", ""),
                        "rate": op.get("rate", 0),
                        "robots": op_robots.get(op.get("id", ""), []),
                    }
                    for op in ops
                ],
            })
        state["robots_activos"] = [r["nombre"] for r in robots_activos]

    # Configuracion (capacidades, dias laborales, fabricas)
    capacidades = _sb_get("capacidades_recurso", "select=tipo,pares_hora&order=tipo")
    if capacidades:
        state["capacidades"] = capacidades

    dias_lab = _sb_get("dias_laborales", "select=dia,activo,minutos,minutos_ot,plantilla&order=dia")
    if dias_lab:
        state["dias_laborales"] = dias_lab

    fabricas = _sb_get("fabricas", "select=nombre,es_maquila,orden&order=orden")
    if fabricas:
        state["fabricas"] = [
            {"nombre": f["nombre"], "es_maquila": f.get("es_maquila", False)}
            for f in fabricas
        ]

    # Pesos de priorizacion y parametros de optimizacion
    pesos = _sb_get("pesos_priorizacion", "select=nombre,valor&order=nombre")
    if pesos:
        state["pesos"] = {p["nombre"]: p["valor"] for p in pesos}
    params_opt = _sb_get("parametros_optimizacion", "select=nombre,valor&order=nombre")
    if params_opt:
        state["parametros_opt"] = {p["nombre"]: p["valor"] for p in params_opt}

    # Operarios con habilidades y disponibilidad
    operarios_raw = _sb_get("operarios", "select=id,nombre,eficiencia,activo&activo=eq.true&order=nombre")
    if operarios_raw:
        habs = _sb_get("operario_habilidades", "select=operario_id,habilidad,nivel")
        habs_by_op = {}
        for h in habs:
            habs_by_op.setdefault(h["operario_id"], []).append(
                {"habilidad": h["habilidad"], "nivel": h.get("nivel", 2)}
            )
        dias_raw = _sb_get("operario_dias", "select=operario_id,dia")
        dias_by_op = {}
        for d in dias_raw:
            dias_by_op.setdefault(d["operario_id"], []).append(d["dia"])

        state["operarios"] = [
            {
                "nombre": op["nombre"],
                "eficiencia": op.get("eficiencia", 1.0),
                "habilidades": habs_by_op.get(op["id"], []),
                "dias_disponibles": dias_by_op.get(op["id"], []),
            }
            for op in operarios_raw
        ]

    # Asignaciones maquila (si hay pedido)
    if ped:
        items_ids = [it["id"] for it in items] if items else []
        if items_ids:
            maquila_raw = _sb_get("asignaciones_maquila", f"select=*&order=id")
            item_id_set = set(items_ids)
            maquila_filtered = [a for a in maquila_raw if a.get("pedido_item_id") in item_id_set]
            if maquila_filtered:
                item_modelo = {it["id"]: it["modelo_num"] for it in items}
                state["asignaciones_maquila"] = [
                    {
                        "modelo": item_modelo.get(a["pedido_item_id"], "?"),
                        "maquila": a.get("maquila", ""),
                        "pares": a.get("pares", 0),
                        "fecha_entrega": str(a.get("fecha_entrega", "")) if a.get("fecha_entrega") else None,
                    }
                    for a in maquila_filtered
                ]

    # Avance
    if semana:
        av = _sb_get("avance", f"select=*&semana=eq.{semana}")
        if av:
            detalles = _sb_get("avance_detalle", f"select=*&avance_id=eq.{av[0]['id']}")
            modelos = {}
            for d in detalles:
                mn = d["modelo_num"]
                if mn not in modelos:
                    modelos[mn] = {}
                modelos[mn][d["dia"]] = d["pares"]
            state["avance"] = {"modelos": modelos}

    return state


# --- Request/Response models ---

class ChatAttachment(BaseModel):
    type: str  # "image" or "excel"
    filename: str
    mime_type: str
    data: Optional[str] = None  # base64
    preview: Optional[str] = None
    size: int = 0


class ChatMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str
    attachments: Optional[list[ChatAttachment]] = None


class ChatRequest(BaseModel):
    messages: list[ChatMessage]
    pedido_nombre: str = ""
    semana: str = ""
    model: str = "claude-sonnet-4-6"


class ChatResponse(BaseModel):
    response: str


# --- Helpers para attachments ---

def _parse_excel_base64(data_b64: str, filename: str, max_rows: int = 50, max_cols: int = 20) -> str:
    """Parsea un Excel base64 y retorna representacion de texto."""
    try:
        raw = base64.b64decode(data_b64)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)

        sections = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    rows.append(f"... ({ws.max_row - max_rows} filas mas)")
                    break
                cells = [str(c if c is not None else "") for c in row[:max_cols]]
                rows.append(" | ".join(cells))

            if rows:
                sections.append(f"Hoja '{sheet_name}':\n" + "\n".join(rows))

        wb.close()
        return "\n\n".join(sections) if sections else "Archivo vacio"
    except Exception as e:
        return f"Error al parsear Excel: {str(e)}"


def _build_message_content(msg: ChatMessage) -> dict:
    """Construye un mensaje para la API de Claude, manejando attachments multimodal."""
    if not msg.attachments:
        return {"role": msg.role, "content": msg.content}

    content_blocks = []

    for att in msg.attachments:
        if att.type == "image" and att.data:
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": att.mime_type,
                    "data": att.data,
                },
            })
        elif att.type == "excel" and att.data:
            excel_text = _parse_excel_base64(att.data, att.filename)
            content_blocks.append({
                "type": "text",
                "text": f"[Archivo Excel: {att.filename}]\n{excel_text}",
            })

    if msg.content:
        content_blocks.append({"type": "text", "text": msg.content})

    return {"role": msg.role, "content": content_blocks if content_blocks else msg.content}


# --- Endpoint ---

@router.post("/chat", response_model=ChatResponse)
def chat_endpoint(req: ChatRequest):
    """Chat con el asistente LLM — soporta tool use para consultas bajo demanda."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY no configurada")

    # Construir estado completo desde Supabase (una sola vez)
    state = _build_state_from_supabase(req.pedido_nombre, req.semana)

    # Context base (resumen ligero, no el detalle completo)
    context = build_context(state)
    system = SYSTEM_PROMPT + "\n\n--- DATOS ACTUALES ---\n" + context

    from anthropic import Anthropic, APIError

    client = Anthropic(api_key=ANTHROPIC_API_KEY)

    # Build messages with multimodal support
    has_attachments = any(m.attachments for m in req.messages if m.attachments)
    messages = [_build_message_content(m) for m in req.messages]
    max_tokens = 4096 if has_attachments else 3000

    try:
        # Tool use loop: Claude may call tools multiple times
        for round_num in range(MAX_TOOL_ROUNDS):
            response = client.messages.create(
                model=req.model,
                max_tokens=max_tokens,
                system=system,
                messages=messages,
                tools=TOOLS,
            )

            # If Claude wants to use tools
            if response.stop_reason == "tool_use":
                # Collect all tool calls and results
                assistant_content = response.content
                tool_results = []

                for block in assistant_content:
                    if block.type == "tool_use":
                        tool_name = block.name
                        tool_input = block.input
                        tool_id = block.id

                        print(f"[CHAT] Tool call #{round_num+1}: {tool_name}({json.dumps(tool_input, ensure_ascii=False)})")

                        handler = TOOL_HANDLERS.get(tool_name)
                        if handler:
                            try:
                                result_text = handler(tool_input, state)
                            except Exception as e:
                                result_text = f"Error ejecutando {tool_name}: {str(e)}"
                        else:
                            result_text = f"Tool '{tool_name}' no reconocida."

                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": tool_id,
                            "content": result_text,
                        })

                # Add assistant message + tool results to conversation
                messages.append({"role": "assistant", "content": assistant_content})
                messages.append({"role": "user", "content": tool_results})
                continue  # Next round

            # Claude finished (no more tool calls) — extract text response
            text_parts = [block.text for block in response.content if hasattr(block, "text")]
            return ChatResponse(response="\n".join(text_parts) if text_parts else "Sin respuesta.")

        # Max rounds exceeded
        text_parts = [block.text for block in response.content if hasattr(block, "text")]
        return ChatResponse(response="\n".join(text_parts) if text_parts else "Se excedio el limite de consultas.")

    except APIError as e:
        print(f"[CHAT] Anthropic API error: {e.status_code} {e.message}")
        raise HTTPException(e.status_code or 500, f"Claude API: {e.message}")
    except Exception as e:
        print(f"[CHAT] Error: {e}")
        raise HTTPException(500, f"Error al llamar Claude: {str(e)}")
